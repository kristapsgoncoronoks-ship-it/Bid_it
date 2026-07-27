"""Self-service pivot engine (Power BI / Tableau-style explore)."""

import pytest


async def _mk(auth_client, vendor, number, date_, category, unit, qty="1", tax="0", status="paid"):
    return await auth_client.post(
        "/api/v1/invoices",
        json={
            "vendor_name": vendor,
            "invoice_number": number,
            "issue_date": date_,
            "status": status,
            "line_items": [
                {
                    "description": "svc",
                    "category": category,
                    "quantity": qty,
                    "unit_price": unit,
                    "tax_rate": tax,
                },
            ],
        },
    )


@pytest.mark.asyncio
async def test_fields_catalog(auth_client):
    cat = (await auth_client.get("/api/v1/analytics/fields")).json()
    dims = {d["key"] for d in cat["dimensions"]}
    measures = {m["key"] for m in cat["measures"]}
    assert {"vendor", "category", "month", "quarter", "year", "status"} <= dims
    assert {"net", "tax", "gross", "quantity", "lines", "invoices"} <= measures


@pytest.mark.asyncio
async def test_pivot_by_category(auth_client):
    await _mk(auth_client, "AWS", "A1", "2026-01-10", "cloud", "100.00")
    await _mk(auth_client, "AWS", "A2", "2026-01-11", "cloud", "50.00")
    await _mk(auth_client, "Slack", "S1", "2026-02-01", "software", "30.00")

    r = (await auth_client.get("/api/v1/analytics/explore?measure=net&dim=category")).json()
    assert r["measure"]["key"] == "net"
    rows = {row["category"]: row["value"] for row in r["rows"]}
    assert rows["cloud"] == "150.00"
    assert rows["software"] == "30.00"
    # sorted by value desc
    assert r["rows"][0]["category"] == "cloud"


@pytest.mark.asyncio
async def test_two_dimensions_and_measures(auth_client):
    await _mk(auth_client, "AWS", "A1", "2026-01-10", "cloud", "100.00", qty="2", tax="21")
    await _mk(auth_client, "AWS", "A2", "2026-02-10", "cloud", "200.00", qty="1", tax="21")

    # month × category, net measure
    r = (
        await auth_client.get("/api/v1/analytics/explore?measure=net&dim=month&dim=category")
    ).json()
    assert [d["key"] for d in r["dimensions"]] == ["month", "category"]
    by_month = {row["month"]: row["value"] for row in r["rows"]}
    assert by_month["2026-01"] == "200.00"  # 2 * 100
    assert by_month["2026-02"] == "200.00"  # 1 * 200
    # temporal first dim → chronological order
    assert [row["month"] for row in r["rows"]] == ["2026-01", "2026-02"]

    # tax measure: 21% of net
    tax = (await auth_client.get("/api/v1/analytics/explore?measure=tax&dim=category")).json()
    assert tax["rows"][0]["value"] == "84.00"  # 21% of 400

    # quantity + counts
    qty = (await auth_client.get("/api/v1/analytics/explore?measure=quantity&dim=category")).json()
    assert qty["rows"][0]["value"] == "3.000"
    inv = (await auth_client.get("/api/v1/analytics/explore?measure=invoices&dim=category")).json()
    assert inv["rows"][0]["value"] == "2"


@pytest.mark.asyncio
async def test_filters_and_csv(auth_client):
    await _mk(auth_client, "AWS", "A1", "2026-01-10", "cloud", "100.00", status="paid")
    await _mk(auth_client, "AWS", "A2", "2026-03-10", "cloud", "200.00", status="pending")

    # date filter
    r = (
        await auth_client.get("/api/v1/analytics/explore?measure=net&dim=vendor&start=2026-02-01")
    ).json()
    assert r["rows"][0]["value"] == "200.00"
    # status filter
    r = (
        await auth_client.get("/api/v1/analytics/explore?measure=net&dim=vendor&status=paid")
    ).json()
    assert r["rows"][0]["value"] == "100.00"

    # CSV export
    csv = await auth_client.get("/api/v1/analytics/explore?measure=net&dim=category&format=csv")
    assert csv.status_code == 200
    assert csv.headers["content-type"].startswith("text/csv")
    assert "Category,Net spend" in csv.text
    assert "cloud" in csv.text


@pytest.mark.asyncio
async def test_bad_field_rejected(auth_client):
    r = await auth_client.get("/api/v1/analytics/explore?measure=net&dim=not_a_dim")
    assert r.status_code == 422
    r2 = await auth_client.get("/api/v1/analytics/explore?measure=bogus&dim=vendor")
    assert r2.status_code == 422


@pytest.mark.asyncio
async def test_xlsx_and_pdf_export(auth_client):
    """board I1.5 — the general report-to-Excel/PDF writers, over the same
    Explore cut CSV already exercises. Never a forked query: same measure/dim,
    three serialisations."""
    await _mk(auth_client, "AWS", "A1", "2026-01-10", "cloud", "100.00")
    await _mk(auth_client, "AWS", "A2", "2026-01-11", "cloud", "50.00")

    xlsx = await auth_client.get("/api/v1/analytics/explore?measure=net&dim=category&format=xlsx")
    assert xlsx.status_code == 200
    assert xlsx.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert len(xlsx.content) > 0
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx.content))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["Category", "Net spend"]
    assert [c.value for c in ws[2]] == ["cloud", "150.00"]

    pdf = await auth_client.get("/api/v1/analytics/explore?measure=net&dim=category&format=pdf")
    assert pdf.status_code == 200
    assert pdf.headers["content-type"].startswith("application/pdf")
    assert pdf.content[:4] == b"%PDF"


@pytest.mark.asyncio
async def test_xlsx_export_requires_report_read(role_client):
    """Same router-level REPORT_READ gate the existing json/csv formats already
    inherit — asserted explicitly for the new format value since it is new
    code path, not just a new format string on an unrelated route."""
    employee = await role_client("user")  # EMPLOYEE — no report.read
    ro = await role_client("user_free")  # READ_ONLY — holds report.read
    denied = await employee.get("/api/v1/analytics/explore?measure=net&dim=category&format=xlsx")
    assert denied.status_code == 403
    granted = await ro.get("/api/v1/analytics/explore?measure=net&dim=category&format=xlsx")
    assert granted.status_code == 200
