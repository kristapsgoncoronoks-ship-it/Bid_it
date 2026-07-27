"""General report-to-Excel/PDF writers over an Explore cut (board I1.5).

Unit-level: both writers consume the exact `explore.run()`/`to_csv()` result
shape by hand — no DB, no HTTP — so these tests characterise the writers in
isolation from the pivot engine itself (already covered by test_explore.py).
"""

from __future__ import annotations

import io

import openpyxl
import pytest

from app.services import report_writers

_RESULT = {
    "measure": {"key": "net", "label": "Net spend", "unit": "money"},
    "dimensions": [{"key": "category", "label": "Category", "temporal": False}],
    "rows": [
        {"category": "cloud", "value": "150.00"},
        {"category": "software", "value": "30.00"},
    ],
}


def test_to_xlsx_roundtrips_header_and_rows():
    xlsx = report_writers.to_xlsx(_RESULT)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == ["Category", "Net spend"]
    rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    assert rows == [["cloud", "150.00"], ["software", "30.00"]]


def test_to_xlsx_neutralises_formula_injection():
    malicious = {
        "measure": {"key": "net", "label": "Net spend", "unit": "money"},
        "dimensions": [{"key": "category", "label": "Category", "temporal": False}],
        "rows": [
            # A malicious dimension label AND a negative money value (leading
            # '-' is as live an Excel formula/arithmetic trigger as '=').
            {"category": "=cmd|'/c calc'!A1", "value": "-100.00"},
        ],
    }
    xlsx = report_writers.to_xlsx(malicious)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb.active
    dim_cell, value_cell = ws[2][0].value, ws[2][1].value
    # The stored string must never start with a live formula trigger — the
    # neutralising quote prefix survives the round-trip.
    assert dim_cell[0] not in ("=", "+", "-", "@", "\t", "\r")
    assert value_cell[0] not in ("=", "+", "-", "@", "\t", "\r")
    assert dim_cell == "'=cmd|'/c calc'!A1"
    assert value_cell == "'-100.00"


def test_to_pdf_returns_valid_pdf_bytes():
    pdf = report_writers.to_pdf(_RESULT)
    assert pdf[:4] == b"%PDF"
    assert len(pdf) > 500  # non-trivial — a real rendered document, not a stub

    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(pdf))
    assert len(reader.pages) >= 1
    text = reader.pages[0].extract_text()
    assert "Net spend" in text
    assert "cloud" in text


@pytest.mark.parametrize("writer", [report_writers.to_xlsx, report_writers.to_pdf])
def test_writers_handle_zero_dimensions(writer):
    grand_total = {
        "measure": {"key": "net", "label": "Net spend", "unit": "money"},
        "dimensions": [],
        "rows": [{"value": "180.00"}],
    }
    out = writer(grand_total)
    assert isinstance(out, bytes)
    assert len(out) > 0


@pytest.mark.parametrize("writer", [report_writers.to_xlsx, report_writers.to_pdf])
def test_writers_handle_no_rows(writer):
    empty = {
        "measure": {"key": "net", "label": "Net spend", "unit": "money"},
        "dimensions": [{"key": "category", "label": "Category", "temporal": False}],
        "rows": [],
    }
    out = writer(empty)
    assert isinstance(out, bytes)
    assert len(out) > 0


def test_safe_cell_neutralises_every_trigger():
    for trigger in ("=", "+", "-", "@", "\t", "\r"):
        assert report_writers._safe_cell(f"{trigger}EVIL").startswith("'" + trigger)
    assert report_writers._safe_cell("ordinary") == "ordinary"
    assert report_writers._safe_cell(None) == ""
