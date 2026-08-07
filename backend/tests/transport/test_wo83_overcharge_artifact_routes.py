"""WO-83 — the two artifact DOWNLOADS over HTTP (G4.5, R41).

Proves the two routes added to `api/routes/transport/overcharges.py`: the Excel
evidence packet and the formal PDF claim letter really cross the wire as those
media types, with a filename and `nosniff`, on the router-level
`TRANSPORT_READ` — plus the matrix every transport slice ships: the structural
permission pair (EMPLOYEE denied / ACCOUNTANT granted), the AUDITOR case that
gives the read/write split its meaning (generation is a READ, so an auditor gets
both artifacts), the module entitlement, the §4.4 opaque cross-tenant 404, and
the fail-closed refusals rendered as `{"detail", "code"}`.

Fixture strategy is `test_wo82_overcharge_routes.py`'s, unchanged: orgs/tokens
ride the real HTTP register flow, transport enablement is direct service setup
(the module is in no billing plan, so the HTTP toggle would 402), fuel lines
seed through `fuel_ingest.ingest_transaction`, and every assertion about what the
API returned goes through the API.
"""

from __future__ import annotations

import io
import uuid
from datetime import date
from decimal import Decimal

import pytest
from httpx import AsyncClient

from app.models.issuer import IssuerProfile
from app.services import modules
from app.services.transport import fuel_ingest
from tests.factories.transport import synthetic_iban, synthetic_vat_id, synthetic_vehicle_ref
from tests.transport.conftest import make_entity

V = "/api/v1"
TERMS = f"{V}/transport/contract-terms"
OVERCHARGES = f"{V}/transport/overcharges"
PERIOD = "2026-05"
SUPPLIER = "Q8"

XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_CT = "application/pdf"


async def _register_org(client: AsyncClient):
    suffix = uuid.uuid4().hex[:8]
    r = await client.post(
        f"{V}/auth/register",
        json={
            "organization_name": f"WO83 Org {suffix}",
            "name": "Owner",
            "email": f"owner-{suffix}@wo83.example.io",
            "password": "supersecret",
        },
    )
    assert r.status_code == 201, r.text
    body = r.json()
    return {"Authorization": f"Bearer {body['token']['access_token']}"}, body["organization"]["id"]


async def _member_with_role(client: AsyncClient, owner_headers, db_session, stored_role: str):
    from sqlalchemy import update

    from app.models.user import User, UserRole

    email = f"{stored_role}-{uuid.uuid4().hex[:8]}@wo83.example.io"
    inv = await client.post(
        f"{V}/team/invites", headers=owner_headers, json={"email": email, "role": "admin"}
    )
    assert inv.status_code == 201, inv.text
    acc = await client.post(
        f"{V}/auth/accept-invite",
        json={"token": inv.json()["token"], "name": "Member", "password": "supersecret"},
    )
    assert acc.status_code == 201, acc.text
    body = acc.json()
    await db_session.execute(
        update(User)
        .where(User.id == body["user"]["id"])
        .values(role=UserRole(stored_role), is_expense_approver=False)
    )
    await db_session.commit()
    return {"Authorization": f"Bearer {body['token']['access_token']}"}


async def _letterhead(db_session, org_id: str, *, seed: int) -> IssuerProfile:
    """A COMPLETE issuer company (the Art. 226 `issuer.REQUIRED_FIELDS`) flagged
    default — `make_entity` leaves `city`/`postal_code` blank on purpose, which
    is what the incomplete-letterhead refusal exercises."""
    profile = IssuerProfile(
        org_id=org_id,
        name="Demo Haulage Group",
        legal_name="Demo Haulage Group SIA",
        is_default=True,
        vat_number=synthetic_vat_id("LV", seed=seed),
        registration_number=f"REG-{seed:06d}",
        address_line1="1 Synthetic Test Street",
        city="Testville",
        postal_code="LV-1000",
        country="LV",
        iban=synthetic_iban("LV", seed=seed),
    )
    db_session.add(profile)
    await db_session.flush()
    return profile


async def _seed_breach(db_session, org_id: str, entity_id: str) -> None:
    """1,000 L invoiced at EUR 1,350.00, EUR 1,300.00 effective: applied
    0.05 EUR/L against an agreed 0.20 EUR/L => gap 0.1500, recover 150.00."""
    await fuel_ingest.ingest_transaction(
        db_session,
        org_id,
        entity_id=entity_id,
        supplier=SUPPLIER,
        period=PERIOD,
        line_seq=1,
        country="LV",
        vehicle_ref=synthetic_vehicle_ref(),
        txn_date=date(2026, 5, 12),
        station="Demo Station Riga",
        product="DIESEL",
        qty=Decimal("1000.000"),
        currency="EUR",
        net_local=Decimal("1350.00"),
        vat_local=Decimal("283.50"),
        gross_local=Decimal("1633.50"),
        net_eur=Decimal("1350.00"),
        vat_eur=Decimal("283.50"),
        net_eur_eff=Decimal("1300.00"),
    )


async def _org_with_claim(client: AsyncClient, db_session, *, seed: int, letterhead: bool = True):
    headers, org_id = await _register_org(client)
    await modules.set_enabled(db_session, org_id, "transport", True)
    entity = await make_entity(db_session, org_id, country="LV", seed=seed)
    if letterhead:
        await _letterhead(db_session, org_id, seed=seed)
    await db_session.commit()
    r = await client.put(
        TERMS,
        headers=headers,
        json={
            "supplier": SUPPLIER,
            "country": "LV",
            "product_group": "Diesel",
            "expected_discount_eur_l": "0.20",
        },
    )
    assert r.status_code == 200, r.text
    await _seed_breach(db_session, org_id, entity.id)
    await db_session.commit()
    r = await client.post(
        OVERCHARGES, headers=headers, json={"supplier": SUPPLIER, "period": PERIOD}
    )
    assert r.status_code == 200, r.text
    return headers, org_id, r.json()


# --------------------------------------------------------------------------- #
# The bytes really are the artifacts
# --------------------------------------------------------------------------- #


@pytest.mark.asyncio
async def test_wo83_the_packet_downloads_as_a_real_xlsx_with_a_filename(client, db_session):
    headers, _org_id, claim = await _org_with_claim(client, db_session, seed=8311)

    r = await client.get(f"{OVERCHARGES}/{claim['id']}/packet", headers=headers)

    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == XLSX_CT
    assert claim["id"] in r.headers["content-disposition"]
    assert ".xlsx" in r.headers["content-disposition"]
    assert r.headers["x-content-type-options"] == "nosniff"

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Claim-back", "Evidence"]
    rows = list(wb["Evidence"].iter_rows(values_only=True))
    assert Decimal(str(rows[1][9])) == Decimal("150.00")
    assert rows[2][0] == "TOTAL"
    assert Decimal(str(rows[2][9])) == Decimal("150.00")


@pytest.mark.asyncio
async def test_wo83_the_letter_downloads_as_a_real_pdf_with_a_filename(client, db_session):
    headers, _org_id, claim = await _org_with_claim(client, db_session, seed=8312)

    r = await client.get(f"{OVERCHARGES}/{claim['id']}/letter", headers=headers)

    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == PDF_CT
    assert ".pdf" in r.headers["content-disposition"]
    assert r.headers["x-content-type-options"] == "nosniff"
    assert r.content[:4] == b"%PDF"

    from pypdf import PdfReader

    text = "".join(p.extract_text() or "" for p in PdfReader(io.BytesIO(r.content)).pages)
    assert "150.00" in text
    assert "within 30 days" in text


@pytest.mark.asyncio
async def test_wo83_both_downloads_report_the_same_total(client, db_session):
    """R41's acceptance, end to end over HTTP: the euro the letter demands is
    the euro the packet's TOTAL row carries, and both are the claim-back's own
    frozen figure."""
    headers, _org_id, claim = await _org_with_claim(client, db_session, seed=8313)

    packet = await client.get(f"{OVERCHARGES}/{claim['id']}/packet", headers=headers)
    letter = await client.get(f"{OVERCHARGES}/{claim['id']}/letter", headers=headers)

    from openpyxl import load_workbook
    from pypdf import PdfReader

    rows = list(load_workbook(io.BytesIO(packet.content))["Evidence"].iter_rows(values_only=True))
    total = next(Decimal(str(r[9])) for r in rows if r[0] == "TOTAL")
    text = "".join(p.extract_text() or "" for p in PdfReader(io.BytesIO(letter.content)).pages)
    assert f"{total:,.2f}" in text
    assert total == Decimal(claim["detected_eur"])


# --------------------------------------------------------------------------- #
# The refusals, in the frozen wire shape
# --------------------------------------------------------------------------- #


@pytest.mark.asyncio
async def test_wo83_an_unknown_claim_id_is_404(client, db_session):
    headers, _org_id, _claim = await _org_with_claim(client, db_session, seed=8314)

    for suffix in ("packet", "letter"):
        r = await client.get(f"{OVERCHARGES}/{uuid.uuid4()}/{suffix}", headers=headers)
        assert r.status_code == 404, r.text
        assert r.json()["code"] == "overcharge_claim_not_found"


@pytest.mark.asyncio
async def test_wo83_a_cross_tenant_claim_id_is_an_opaque_404(client, db_session):
    """§4.4 — object-id guessing must yield zero information, and it must not
    leak a DOCUMENT either."""
    headers_a, _org_a, _claim_a = await _org_with_claim(client, db_session, seed=8315)
    _headers_b, _org_b, claim_b = await _org_with_claim(client, db_session, seed=8316)

    for suffix in ("packet", "letter"):
        r = await client.get(f"{OVERCHARGES}/{claim_b['id']}/{suffix}", headers=headers_a)
        assert r.status_code != 403, "a 403 would confirm the object exists (§4.4)"
        assert r.status_code == 404
        assert r.json()["code"] == "overcharge_claim_not_found"


@pytest.mark.asyncio
async def test_wo83_a_terminal_claim_back_refuses_the_letter_and_serves_the_packet(
    client, db_session
):
    headers, _org_id, claim = await _org_with_claim(client, db_session, seed=8317)
    for to_status in ("packaged", "claimed", "written_off"):
        r = await client.post(
            f"{OVERCHARGES}/{claim['id']}/advance", headers=headers, json={"to_status": to_status}
        )
        assert r.status_code == 200, r.text

    r = await client.get(f"{OVERCHARGES}/{claim['id']}/letter", headers=headers)
    assert r.status_code == 409, r.text
    assert r.json()["code"] == "overcharge_claim_closed"

    r = await client.get(f"{OVERCHARGES}/{claim['id']}/packet", headers=headers)
    assert r.status_code == 200, r.text


@pytest.mark.asyncio
async def test_wo83_an_incomplete_letterhead_refuses_the_letter_over_http(client, db_session):
    headers, _org_id, claim = await _org_with_claim(client, db_session, seed=8318, letterhead=False)

    r = await client.get(f"{OVERCHARGES}/{claim['id']}/letter", headers=headers)

    assert r.status_code == 409, r.text
    assert r.json()["code"] == "issuer_profile_incomplete"
    assert "city" in r.json()["detail"]
    assert (
        await client.get(f"{OVERCHARGES}/{claim['id']}/packet", headers=headers)
    ).status_code == 200


@pytest.mark.asyncio
async def test_wo83_a_drifted_claim_refuses_409_over_http(client, db_session):
    """§4.10 — a demand whose own attached lines no longer reproduce the frozen
    figure is refused, not silently re-based."""
    headers, org_id, claim = await _org_with_claim(client, db_session, seed=8319)

    from sqlalchemy import update

    from app.models.transport.overcharge import VatOverchargeClaim

    await db_session.execute(
        update(VatOverchargeClaim)
        .where(VatOverchargeClaim.id == claim["id"])
        .values(detected_eur=Decimal("900.00"))
    )
    await db_session.commit()

    for suffix in ("packet", "letter"):
        r = await client.get(f"{OVERCHARGES}/{claim['id']}/{suffix}", headers=headers)
        assert r.status_code == 409, r.text
        assert r.json()["code"] == "overcharge_evidence_drift"


@pytest.mark.asyncio
async def test_wo83_module_disabled_refuses_403_module_not_enabled(client, db_session):
    headers, org_id, claim = await _org_with_claim(client, db_session, seed=8320)
    await modules.set_enabled(db_session, org_id, "transport", False)
    await db_session.commit()

    for suffix in ("packet", "letter"):
        r = await client.get(f"{OVERCHARGES}/{claim['id']}/{suffix}", headers=headers)
        assert r.status_code == 403, r.text
        assert r.json()["code"] == "module_not_enabled"


# --------------------------------------------------------------------------- #
# Authorization — the reads ride the router-level TRANSPORT_READ
# --------------------------------------------------------------------------- #


@pytest.mark.asyncio
async def test_wo83_employee_is_denied_and_accountant_is_granted(client, db_session):
    owner_headers, _org_id, claim = await _org_with_claim(client, db_session, seed=8321)
    employee = await _member_with_role(client, owner_headers, db_session, "user")
    accountant = await _member_with_role(client, owner_headers, db_session, "accountant")

    for suffix in ("packet", "letter"):
        url = f"{OVERCHARGES}/{claim['id']}/{suffix}"
        assert (await client.get(url, headers=employee)).status_code == 403
        assert (await client.get(url, headers=accountant)).status_code == 200


@pytest.mark.asyncio
async def test_wo83_an_auditor_can_download_both_artifacts(client, db_session):
    """Generating an artifact is a READ — nothing is persisted, mutated or
    audited — so it rides `TRANSPORT_READ` and an AUDITOR (which holds
    `TRANSPORT_READ` and NOT `VAT_WRITE`) gets both, while still being unable to
    drive the worklist."""
    owner_headers, _org_id, claim = await _org_with_claim(client, db_session, seed=8322)
    auditor = await _member_with_role(client, owner_headers, db_session, "auditor")

    for suffix in ("packet", "letter"):
        r = await client.get(f"{OVERCHARGES}/{claim['id']}/{suffix}", headers=auditor)
        assert r.status_code == 200, r.text
    assert (
        await client.post(
            f"{OVERCHARGES}/{claim['id']}/advance", headers=auditor, json={"to_status": "packaged"}
        )
    ).status_code == 403
