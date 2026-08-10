"""WO-77 — transport routes slice 2: the admin/config surfaces + the filing
artifacts over HTTP.

Proves the 27 new endpoints end to end: receipt-control waivers (R15), the
checklist-rule admin (R45 — deactivating a rule removes it from the gate),
the manual status codes (R17/R12), the engine tie-out expectations (R25
regime 2), the receipt-control cadences + persisted grid (G3.5, advisory),
the note→invoice-ref overrides (R16), the customer lifecycle + per-country
activation ladder (R44), and the G2.12 workbook / evidence-pack downloads
(content-type + a REAL parse of the bytes, manifest included).

Fixture strategy mirrors WO-76: orgs/tokens ride the real HTTP register
flow; transport enablement is direct service setup (the module is in no
billing plan yet, so the HTTP toggle would 402); domain rows seed through
the same conftest helpers the service suites use. Every assertion about what
the API DID goes through the API; direct DB reads only verify internals a
client cannot see (nothing-mutated).
"""

from __future__ import annotations

import io
import uuid
import zipfile
from datetime import date
from decimal import Decimal

import pytest
from httpx import AsyncClient
from sqlalchemy import update

from app.services import modules
from app.services.transport import fuel_ingest, receipt_control
from tests.factories.transport import synthetic_vehicle_ref
from tests.transport.conftest import (
    activate_entity,
    make_entity,
    register_documented_invoice,
    seed_fee_rate,
)

V = "/api/v1"


# --------------------------------------------------------------------------- #
# HTTP org / member helpers (the WO-76 shapes)
# --------------------------------------------------------------------------- #


async def _register_org(client: AsyncClient):
    suffix = uuid.uuid4().hex[:8]
    r = await client.post(
        f"{V}/auth/register",
        json={
            "organization_name": f"WO77 Org {suffix}",
            "name": "Owner",
            "email": f"owner-{suffix}@wo77.example.io",
            "password": "supersecret",
        },
    )
    assert r.status_code == 201, r.text
    body = r.json()
    return {"Authorization": f"Bearer {body['token']['access_token']}"}, body["organization"]["id"]


async def _enable_transport(db_session, org_id: str) -> None:
    # Direct service enablement: `transport` is deliberately in no PLANS set
    # (a commercial decision), so the HTTP module toggle would 402. Setup
    # only — every path under test stays HTTP.
    await modules.set_enabled(db_session, org_id, "transport", True)
    # WO-95 (G2.9/R13): raised past the fee gate — `submit_claim` now freezes
    # a contingency fee and REFUSES (`fee_rate_not_configured`) when no rate
    # is configured, so a submit route test would otherwise 409 on a
    # commercial setting rather than exercise the path under test. Seeded as
    # a row, not through `fee.set_rate`, so no audit event enters the window
    # this file's trail assertions read (the `activate_entity` convention).
    await seed_fee_rate(db_session, org_id)


async def _member_with_role(client: AsyncClient, owner_headers, db_session, stored_role: str):
    """Invite a member into the OWNER's org, then downgrade the stored role
    (the WO-76 pattern — invites only mint admin-tier accounts)."""
    from app.models.user import User, UserRole

    email = f"{stored_role}-{uuid.uuid4().hex[:8]}@wo77.example.io"
    inv = await client.post(
        f"{V}/team/invites", headers=owner_headers, json={"email": email, "role": "admin"}
    )
    assert inv.status_code == 201, inv.text
    acc = await client.post(
        f"{V}/auth/accept-invite",
        json={"token": inv.json()["token"], "name": "Member", "password": "supersecret"},
    )
    assert acc.status_code == 201, acc.text
    body = acc.json()
    await db_session.execute(
        update(User)
        .where(User.id == body["user"]["id"])
        .values(role=UserRole(stored_role), is_expense_approver=False)
    )
    await db_session.commit()
    return {"Authorization": f"Bearer {body['token']['access_token']}"}


# --------------------------------------------------------------------------- #
# Domain seeding helpers
# --------------------------------------------------------------------------- #


async def _make_txn(
    db_session,
    org_id: str,
    entity_id: str,
    *,
    supplier: str = "Q8",
    invoice_ref: str | None = "INV-0001",
    period: str = "2026-05",
    txn_date: date = date(2026, 5, 15),
    line_seq: int = 1,
    net_eur: Decimal = Decimal("2000.00"),
    vat_eur: Decimal = Decimal("420.00"),
):
    return await fuel_ingest.ingest_transaction(
        db_session,
        org_id,
        entity_id=entity_id,
        supplier=supplier,
        period=period,
        line_seq=line_seq,
        country="LV",
        vehicle_ref=synthetic_vehicle_ref(),
        txn_date=txn_date,
        station="Demo Station Riga",
        product="DIESEL",
        qty=Decimal("100.000"),
        currency="EUR",
        net_local=net_eur,
        vat_local=vat_eur,
        gross_local=net_eur + vat_eur,
        net_eur=net_eur,
        vat_eur=vat_eur,
        invoice_ref=invoice_ref,
    )


async def _create_claim_http(
    client, headers, entity_id: str, *, refund_country: str = "LV", ref_period: str = "2026-Q2"
) -> dict:
    r = await client.post(
        f"{V}/transport/claims",
        headers=headers,
        json={"entity_id": entity_id, "refund_country": refund_country, "ref_period": ref_period},
    )
    assert r.status_code == 200, r.text
    return r.json()


async def _draft_claim(client, db_session, headers, org_id):
    """A draft claim over an unresolved Q8 transaction (waivable supplier)."""
    entity = await make_entity(db_session, org_id)
    await _make_txn(db_session, org_id, entity.id)
    await db_session.commit()
    claim = await _create_claim_http(client, headers, entity.id)
    return claim, entity


async def _submitted_claim(client, db_session, headers, org_id):
    """The G2.12 happy-path fixture over HTTP: an activated entity, a
    registered+vaulted invoice, one resolvable transaction above the Art. 17
    minimum, lines built and the claim SUBMITTED (so its lines are frozen)."""
    entity = await make_entity(db_session, org_id)
    await activate_entity(db_session, org_id, entity.id, "LV")
    await register_documented_invoice(
        db_session, org_id, supplier="Q8", invoice_number="INV-0001", country="LV"
    )
    txn = await _make_txn(db_session, org_id, entity.id)
    txn_id = txn.id
    await db_session.commit()
    claim = await _create_claim_http(client, headers, entity.id)
    built = await client.post(f"{V}/transport/claims/{claim['id']}/lines", headers=headers)
    assert built.status_code == 200, built.text
    sub = await client.post(
        f"{V}/transport/claims/{claim['id']}/submit",
        headers=headers,
        json={
            "invoices": [
                {"supplier": "Q8", "invoice_ref": "INV-0001", "fuel_transaction_id": txn_id}
            ]
        },
    )
    assert sub.status_code == 200, sub.text
    return sub.json(), entity


# --------------------------------------------------------------------------- #
# Receipt-control waivers (R15)
# --------------------------------------------------------------------------- #


@pytest.mark.asyncio
async def test_wo77_waiver_set_list_remove_over_http(client, db_session):
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    claim, _entity = await _draft_claim(client, db_session, headers, org_id)

    r = await client.post(
        f"{V}/transport/claims/{claim['id']}/waivers", headers=headers, json={"supplier": "Q8"}
    )
    assert r.status_code == 200, r.text
    assert r.json()["supplier"] == "Q8"

    listing = await client.get(f"{V}/transport/claims/{claim['id']}/waivers", headers=headers)
    assert listing.status_code == 200, listing.text
    assert [w["supplier"] for w in listing.json()] == ["Q8"]

    # Idempotent on (claim, supplier) — a repeat returns the same row.
    again = await client.post(
        f"{V}/transport/claims/{claim['id']}/waivers", headers=headers, json={"supplier": "Q8"}
    )
    assert again.status_code == 200 and again.json()["id"] == r.json()["id"]

    gone = await client.delete(f"{V}/transport/claims/{claim['id']}/waivers/Q8", headers=headers)
    assert gone.status_code == 200 and gone.json() == {"removed": True}
    noop = await client.delete(f"{V}/transport/claims/{claim['id']}/waivers/Q8", headers=headers)
    assert noop.status_code == 200 and noop.json() == {"removed": False}
    empty = await client.get(f"{V}/transport/claims/{claim['id']}/waivers", headers=headers)
    assert empty.json() == []


@pytest.mark.asyncio
async def test_wo77_waiver_on_an_invoiced_supplier_is_422(client, db_session):
    """R15 — waiving a supplier that HAS a registered invoice for the refund
    country would drop claimable VAT; the service refuses and NOTHING is
    waived."""
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    entity = await make_entity(db_session, org_id)
    await register_documented_invoice(
        db_session, org_id, supplier="Q8", invoice_number="INV-0001", country="LV"
    )
    await _make_txn(db_session, org_id, entity.id)
    await db_session.commit()
    claim = await _create_claim_http(client, headers, entity.id)

    r = await client.post(
        f"{V}/transport/claims/{claim['id']}/waivers", headers=headers, json={"supplier": "Q8"}
    )
    assert r.status_code == 422, r.text
    assert r.json()["code"] == "waiver_supplier_has_invoices"
    listing = await client.get(f"{V}/transport/claims/{claim['id']}/waivers", headers=headers)
    assert listing.json() == []  # nothing mutated


@pytest.mark.asyncio
async def test_wo77_waiver_on_a_submitted_claim_is_409_claim_not_draft(client, db_session):
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    claim, _entity = await _submitted_claim(client, db_session, headers, org_id)

    r = await client.post(
        f"{V}/transport/claims/{claim['id']}/waivers", headers=headers, json={"supplier": "DKV"}
    )
    assert r.status_code == 409, r.text
    assert r.json()["code"] == "claim_not_draft"


# --------------------------------------------------------------------------- #
# Checklist-rule admin (R45)
# --------------------------------------------------------------------------- #


@pytest.mark.asyncio
async def test_wo77_checklist_rule_seed_list_and_deactivate_removes_it_from_the_gate(
    client, db_session
):
    """R45's own acceptance test, over HTTP: deactivate a rule ⇒ it
    disappears from the claim's checklist evaluation."""
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    claim, _entity = await _draft_claim(client, db_session, headers, org_id)

    seeded = await client.post(f"{V}/transport/checklist-rules/seed", headers=headers)
    assert seeded.status_code == 200, seeded.text
    assert {r["key"] for r in seeded.json()} == {"customer_data", "bank_account"}
    # Idempotent: a repeat seeds (and audits) nothing.
    assert (await client.post(f"{V}/transport/checklist-rules/seed", headers=headers)).json() == []

    listing = await client.get(f"{V}/transport/checklist-rules", headers=headers)
    assert listing.status_code == 200, listing.text
    assert [r["key"] for r in listing.json()] == ["customer_data", "bank_account"]
    assert all(r["active"] for r in listing.json())

    before = await client.get(f"{V}/transport/claims/{claim['id']}/checklist", headers=headers)
    assert "customer_data" in {i["key"] for i in before.json()}

    off = await client.post(
        f"{V}/transport/checklist-rules/customer_data/active",
        headers=headers,
        json={"active": False},
    )
    assert off.status_code == 200, off.text
    assert off.json()["active"] is False

    after = await client.get(f"{V}/transport/claims/{claim['id']}/checklist", headers=headers)
    assert "customer_data" not in {i["key"] for i in after.json()}
    assert "bank_account" in {i["key"] for i in after.json()}


@pytest.mark.asyncio
async def test_wo77_unknown_checklist_rule_is_404(client, db_session):
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    r = await client.post(
        f"{V}/transport/checklist-rules/no_such_rule/active",
        headers=headers,
        json={"active": False},
    )
    assert r.status_code == 404, r.text
    assert r.json()["code"] == "checklist_rule_not_found"


# --------------------------------------------------------------------------- #
# Manual status codes (R17 / R12)
# --------------------------------------------------------------------------- #


@pytest.mark.asyncio
async def test_wo77_status_code_vocabulary_read(client, db_session):
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    r = await client.get(f"{V}/transport/status-codes", headers=headers)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["auto"] == ["1A", "1B", "1C", "1E"]
    assert "2A" in body["manual"] and "3D" in body["manual"]
    assert not set(body["auto"]) & set(body["manual"])  # never settable, never overlapping


@pytest.mark.asyncio
async def test_wo77_status_code_happy_path_stamps_label_and_deadline(client, db_session):
    """R12 — the code + its SOFT action deadline land on a submitted claim
    and the coarse engine `status` is untouched."""
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    claim, _entity = await _submitted_claim(client, db_session, headers, org_id)

    r = await client.post(
        f"{V}/transport/claims/{claim['id']}/status-code",
        headers=headers,
        json={"code": "3D", "action_deadline": "2027-01-31"},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["status_code"] == "3D"
    assert body["action_deadline"] == "2027-01-31"
    assert body["status"] == "submitted"  # the engine column never moves here


@pytest.mark.asyncio
async def test_wo77_status_code_refusals_leave_the_code_unchanged(client, db_session):
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    claim, _entity = await _submitted_claim(client, db_session, headers, org_id)

    # An AUTO code is system-controlled (R17).
    auto = await client.post(
        f"{V}/transport/claims/{claim['id']}/status-code", headers=headers, json={"code": "1E"}
    )
    assert auto.status_code == 409, auto.text
    assert auto.json()["code"] == "status_code_system_controlled"

    # An unknown code is refused at the service's own vocabulary check.
    unknown = await client.post(
        f"{V}/transport/claims/{claim['id']}/status-code", headers=headers, json={"code": "ZZ"}
    )
    assert unknown.status_code == 422, unknown.text
    assert unknown.json()["code"] == "unknown_status_code"

    detail = await client.get(f"{V}/transport/claims/{claim['id']}", headers=headers)
    assert detail.json()["status_code"] == "2"  # unchanged by either refusal


@pytest.mark.asyncio
async def test_wo77_status_code_on_a_draft_claim_is_409_claim_not_submitted(client, db_session):
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    claim, _entity = await _draft_claim(client, db_session, headers, org_id)

    r = await client.post(
        f"{V}/transport/claims/{claim['id']}/status-code", headers=headers, json={"code": "2A"}
    )
    assert r.status_code == 409, r.text
    assert r.json()["code"] == "claim_not_submitted"  # D4's own wording
    detail = await client.get(f"{V}/transport/claims/{claim['id']}", headers=headers)
    assert detail.json()["status_code"] is None and detail.json()["status"] == "draft"


# --------------------------------------------------------------------------- #
# Engine tie-out expectations (R25 regime 2)
# --------------------------------------------------------------------------- #


def _expectation_body(entity_id: str, **extra) -> dict:
    body = {
        "entity_id": entity_id,
        "supplier": "Q8",
        "period": "2026-05",
        "currency": "EUR",
        "expected_lines": 3,
        "expected_gross_local": "2420.00",
        "expected_net_eur": "2000.00",
        "expected_diesel_litres": "100.000",
    }
    body.update(extra)
    return body


@pytest.mark.asyncio
async def test_wo77_tie_out_expectation_put_read_retype_and_delete(client, db_session):
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    entity = await make_entity(db_session, org_id)
    await db_session.commit()

    r = await client.put(
        f"{V}/transport/tie-out-expectations", headers=headers, json=_expectation_body(entity.id)
    )
    assert r.status_code == 200, r.text
    body = r.json()
    # §4.9 — Decimal on the wire as strings, quantized by the SERVICE.
    assert body["expected_gross_local"] == "2420.00"
    assert body["expected_net_eur"] == "2000.00"
    assert body["expected_diesel_litres"] == "100.000"
    assert body["gross_local_tolerance"] == "0.02"  # the harvested default

    listing = await client.get(
        f"{V}/transport/tie-out-expectations", headers=headers, params={"period": "2026-05"}
    )
    assert listing.status_code == 200, listing.text
    assert [e["id"] for e in listing.json()] == [body["id"]]

    # A typo is corrected by simply retyping — an UPSERT on the natural key.
    retyped = await client.put(
        f"{V}/transport/tie-out-expectations",
        headers=headers,
        json=_expectation_body(entity.id, expected_lines=4),
    )
    assert retyped.status_code == 200, retyped.text
    assert retyped.json()["id"] == body["id"] and retyped.json()["expected_lines"] == 4

    params = {
        "entity_id": entity.id,
        "supplier": "Q8",
        "period": "2026-05",
        "currency": "EUR",
    }
    gone = await client.delete(
        f"{V}/transport/tie-out-expectations", headers=headers, params=params
    )
    assert gone.status_code == 204, gone.text
    again = await client.delete(
        f"{V}/transport/tie-out-expectations", headers=headers, params=params
    )
    assert again.status_code == 404, again.text
    assert again.json()["code"] == "tieout_expectation_not_found"


@pytest.mark.asyncio
async def test_wo77_tie_out_expectation_validation_refusals(client, db_session):
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    entity = await make_entity(db_session, org_id)
    await db_session.commit()

    bad_period = await client.put(
        f"{V}/transport/tie-out-expectations",
        headers=headers,
        json=_expectation_body(entity.id, period="2026-13"),
    )
    assert bad_period.status_code == 422 and bad_period.json()["code"] == "invalid_period"

    bad_ccy = await client.put(
        f"{V}/transport/tie-out-expectations",
        headers=headers,
        json=_expectation_body(entity.id, currency="E1R"),
    )
    assert bad_ccy.status_code == 422 and bad_ccy.json()["code"] == "invalid_currency"

    bad_lines = await client.put(
        f"{V}/transport/tie-out-expectations",
        headers=headers,
        json=_expectation_body(entity.id, expected_lines=-1),
    )
    assert bad_lines.status_code == 422 and bad_lines.json()["code"] == "invalid_expected_lines"

    # The harvested tolerance band: 0.06 outside, 0.05 the inclusive maximum.
    too_wide = await client.put(
        f"{V}/transport/tie-out-expectations",
        headers=headers,
        json=_expectation_body(entity.id, gross_local_tolerance="0.06"),
    )
    assert too_wide.status_code == 422 and too_wide.json()["code"] == "invalid_gross_tolerance"
    at_max = await client.put(
        f"{V}/transport/tie-out-expectations",
        headers=headers,
        json=_expectation_body(entity.id, gross_local_tolerance="0.05"),
    )
    assert at_max.status_code == 200, at_max.text
    assert at_max.json()["gross_local_tolerance"] == "0.05"

    unknown_entity = await client.put(
        f"{V}/transport/tie-out-expectations",
        headers=headers,
        json=_expectation_body(str(uuid.uuid4())),
    )
    assert unknown_entity.status_code == 404 and unknown_entity.json()["code"] == "entity_not_found"


# --------------------------------------------------------------------------- #
# Cadences + the persisted receipt-control grid (G3.5 — advisory)
# --------------------------------------------------------------------------- #


@pytest.mark.asyncio
async def test_wo77_cadence_put_list_delete_and_invalid_value(client, db_session):
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)

    r = await client.put(
        f"{V}/transport/cadences/Q8", headers=headers, json={"cadence": "monthly-per-country"}
    )
    assert r.status_code == 200, r.text
    assert r.json()["cadence"] == "monthly-per-country"

    listing = await client.get(f"{V}/transport/cadences", headers=headers)
    assert [c["supplier"] for c in listing.json()] == ["Q8"]

    bad = await client.put(
        f"{V}/transport/cadences/Q8", headers=headers, json={"cadence": "fortnightly"}
    )
    assert bad.status_code == 422, bad.text
    assert bad.json()["code"] == "invalid_cadence"  # the harvested closed set

    gone = await client.delete(f"{V}/transport/cadences/Q8", headers=headers)
    assert gone.status_code == 200 and gone.json() == {"removed": True}
    noop = await client.delete(f"{V}/transport/cadences/Q8", headers=headers)
    assert noop.json() == {"removed": False}


@pytest.mark.asyncio
async def test_wo77_receipt_control_grid_read_and_override(client, db_session):
    """The grid the close's `run_control` stage persisted, read over HTTP;
    the override is a worklist mute that changes no claim (§4.19)."""
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    entity = await make_entity(db_session, org_id)
    await _make_txn(db_session, org_id, entity.id)
    await db_session.commit()
    # The RUN is the close's stage (R60 — never inline in a request), so it is
    # driven service-side here; the READ path under test stays HTTP.
    await receipt_control.run_receipt_control(db_session, org_id, "2026-05")
    await db_session.commit()

    grid = await client.get(
        f"{V}/transport/receipt-controls", headers=headers, params={"period": "2026-05"}
    )
    assert grid.status_code == 200, grid.text
    rows = grid.json()
    assert rows and all(r["period"] == "2026-05" for r in rows)
    row = rows[0]
    assert row["supplier"] == "Q8"
    assert row["status"] == "missing"  # activity, no registered invoice — chase it
    assert row["waived"] is False and row["note"] is None

    override = await client.post(
        f"{V}/transport/receipt-controls/{row['id']}/override",
        headers=headers,
        json={"waived": True, "note": "Supplier confirmed no invoice for this slot"},
    )
    assert override.status_code == 200, override.text
    assert override.json()["waived"] is True
    assert override.json()["status"] == "missing"  # the computed column is untouched

    again = await client.get(
        f"{V}/transport/receipt-controls", headers=headers, params={"period": "2026-05"}
    )
    persisted = next(r for r in again.json() if r["id"] == row["id"])
    assert persisted["waived"] is True
    assert persisted["note"] == "Supplier confirmed no invoice for this slot"

    bad_period = await client.get(
        f"{V}/transport/receipt-controls", headers=headers, params={"period": "2026-13"}
    )
    assert bad_period.status_code == 422 and bad_period.json()["code"] == "invalid_period"

    unknown = await client.post(
        f"{V}/transport/receipt-controls/{uuid.uuid4()}/override",
        headers=headers,
        json={"waived": True},
    )
    assert unknown.status_code == 404 and unknown.json()["code"] == "receipt_control_not_found"


# --------------------------------------------------------------------------- #
# Note→invoice-ref overrides (R16)
# --------------------------------------------------------------------------- #


@pytest.mark.asyncio
async def test_wo77_note_override_set_and_list(client, db_session):
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    entity = await make_entity(db_session, org_id)
    inv = await register_documented_invoice(
        db_session, org_id, supplier="Q8", invoice_number="INV-0001", country="LV"
    )
    inv_id = inv.id
    await db_session.commit()

    body = {
        "entity_id": entity.id,
        "supplier": "Q8",
        "refund_country": "LV",
        "invoice_ref": "STMT-REF-77",
        "target_invoice_id": inv_id,
    }
    r = await client.put(f"{V}/transport/note-overrides", headers=headers, json=body)
    assert r.status_code == 200, r.text
    assert r.json()["target_invoice_id"] == inv_id

    listing = await client.get(f"{V}/transport/note-overrides", headers=headers)
    assert listing.status_code == 200, listing.text
    assert [o["invoice_ref"] for o in listing.json()] == ["STMT-REF-77"]

    # Idempotent on the key — a second call retargets, never duplicates.
    again = await client.put(f"{V}/transport/note-overrides", headers=headers, json=body)
    assert again.json()["id"] == r.json()["id"]
    assert len((await client.get(f"{V}/transport/note-overrides", headers=headers)).json()) == 1


@pytest.mark.asyncio
async def test_wo77_note_override_refusals(client, db_session):
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    entity = await make_entity(db_session, org_id)
    inv = await register_documented_invoice(
        db_session, org_id, supplier="Q8", invoice_number="INV-0001", country="LV"
    )
    inv_id = inv.id
    await db_session.commit()

    synthetic = await client.put(
        f"{V}/transport/note-overrides",
        headers=headers,
        json={
            "entity_id": entity.id,
            "supplier": "Q8",
            "refund_country": "LV",
            "invoice_ref": "UNMATCHED",
            "target_invoice_id": inv_id,
        },
    )
    assert synthetic.status_code == 422, synthetic.text
    assert synthetic.json()["code"] == "override_note_is_synthetic"

    unregistered = await client.put(
        f"{V}/transport/note-overrides",
        headers=headers,
        json={
            "entity_id": entity.id,
            "supplier": "Q8",
            "refund_country": "LV",
            "invoice_ref": "STMT-REF-77",
            "target_invoice_id": str(uuid.uuid4()),
        },
    )
    assert unregistered.status_code == 422, unregistered.text
    assert unregistered.json()["code"] == "override_target_not_registered"
    assert (await client.get(f"{V}/transport/note-overrides", headers=headers)).json() == []


# --------------------------------------------------------------------------- #
# Customer lifecycle + per-country activation (R44)
# --------------------------------------------------------------------------- #


@pytest.mark.asyncio
async def test_wo77_customer_lifecycle_ladder_over_http(client, db_session):
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    entity = await make_entity(db_session, org_id)
    await db_session.commit()
    base = f"{V}/transport/customers/{entity.id}"

    # Never onboarded: a MEANINGFUL absence, reported — never a 404.
    initial = await client.get(f"{base}/lifecycle", headers=headers)
    assert initial.status_code == 200, initial.text
    assert initial.json() == {
        "entity_id": entity.id,
        "id": None,
        "status": None,
        "countries": [],
    }

    # Activation is never skippable: promotion first.
    early = await client.post(f"{base}/activation", headers=headers, json={"active": True})
    assert early.status_code == 409, early.text
    assert early.json()["code"] == "lifecycle_transition_invalid"

    assert (await client.post(f"{base}/prospect", headers=headers)).json()["status"] == "prospect"
    # Idempotent, and it never downgrades a real client.
    assert (await client.post(f"{base}/prospect", headers=headers)).json()["status"] == "prospect"
    assert (await client.post(f"{base}/promote", headers=headers)).json()["status"] == "pending"
    promoted_twice = await client.post(f"{base}/promote", headers=headers)
    assert promoted_twice.status_code == 409 and promoted_twice.json()["code"] == "not_a_prospect"

    activated = await client.post(f"{base}/activation", headers=headers, json={"active": True})
    assert activated.status_code == 200 and activated.json()["status"] == "active"

    # The country ladder is linear: activating a never-requested country refuses.
    never = await client.post(
        f"{base}/countries/LV/activation", headers=headers, json={"active": True}
    )
    assert never.status_code == 409 and never.json()["code"] == "country_not_requested"

    requested = await client.post(f"{base}/countries/lv/request", headers=headers)
    assert requested.status_code == 200, requested.text
    assert requested.json()["countries"] == [
        {"id": requested.json()["countries"][0]["id"], "country": "LV", "status": "requested"}
    ]

    live = await client.post(
        f"{base}/countries/LV/activation", headers=headers, json={"active": True}
    )
    assert live.status_code == 200 and live.json()["countries"][0]["status"] == "active"
    repeat = await client.post(
        f"{base}/countries/LV/activation", headers=headers, json={"active": True}
    )
    assert repeat.status_code == 409 and repeat.json()["code"] == "country_transition_invalid"

    bad_country = await client.post(f"{base}/countries/XYZ/request", headers=headers)
    assert bad_country.status_code == 422 and bad_country.json()["code"] == "invalid_country"

    overview = await client.get(f"{base}/lifecycle", headers=headers)
    assert overview.json()["status"] == "active"
    assert [c["country"] for c in overview.json()["countries"]] == ["LV"]

    # The churn edge — the commercial off-switch.
    inactive = await client.post(f"{base}/inactive", headers=headers)
    assert inactive.status_code == 200 and inactive.json()["status"] == "inactive"
    terminal = await client.post(f"{base}/activation", headers=headers, json={"active": True})
    assert terminal.status_code == 409  # `inactive` is terminal in this slice


@pytest.mark.asyncio
async def test_wo77_lifecycle_activation_is_what_the_r44_submit_gate_reads(client, db_session):
    """The routes only MOVE state; the fail-CLOSED gate stays in the service.
    Activating over HTTP is what lets a submission past R44."""
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    entity = await make_entity(db_session, org_id)
    await register_documented_invoice(
        db_session, org_id, supplier="Q8", invoice_number="INV-0001", country="LV"
    )
    txn = await _make_txn(db_session, org_id, entity.id)
    txn_id = txn.id
    await db_session.commit()
    claim = await _create_claim_http(client, headers, entity.id)
    assert (
        await client.post(f"{V}/transport/claims/{claim['id']}/lines", headers=headers)
    ).status_code == 200
    submit_body = {
        "invoices": [{"supplier": "Q8", "invoice_ref": "INV-0001", "fuel_transaction_id": txn_id}]
    }

    refused = await client.post(
        f"{V}/transport/claims/{claim['id']}/submit", headers=headers, json=submit_body
    )
    assert refused.status_code == 409 and refused.json()["code"] == "customer_not_active"

    base = f"{V}/transport/customers/{entity.id}"
    await client.post(f"{base}/prospect", headers=headers)
    await client.post(f"{base}/promote", headers=headers)
    await client.post(f"{base}/activation", headers=headers, json={"active": True})

    # Customer active, country still not — the second half of the gate.
    country_refused = await client.post(
        f"{V}/transport/claims/{claim['id']}/submit", headers=headers, json=submit_body
    )
    assert country_refused.status_code == 409
    assert country_refused.json()["code"] == "country_not_activated"

    await client.post(f"{base}/countries/LV/request", headers=headers)
    await client.post(f"{base}/countries/LV/activation", headers=headers, json={"active": True})
    ok = await client.post(
        f"{V}/transport/claims/{claim['id']}/submit", headers=headers, json=submit_body
    )
    assert ok.status_code == 200, ok.text
    assert ok.json()["status"] == "submitted"


# --------------------------------------------------------------------------- #
# Filing artifacts (G2.12) — real byte parses
# --------------------------------------------------------------------------- #


@pytest.mark.asyncio
async def test_wo77_workbook_download_parses_and_totals_the_frozen_claim(client, db_session):
    from openpyxl import load_workbook

    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    claim, _entity = await _submitted_claim(client, db_session, headers, org_id)

    r = await client.get(f"{V}/transport/claims/{claim['id']}/workbook", headers=headers)
    assert r.status_code == 200, r.text
    assert (
        r.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in r.headers["content-disposition"]
    assert f"claim-{claim['id']}-workbook.xlsx" in r.headers["content-disposition"]
    assert r.headers["x-content-type-options"] == "nosniff"

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Claim", "Lines"]
    header = {row[0]: row[1] for row in wb["Claim"].iter_rows(values_only=True)}
    assert header["Refund country"] == "LV"
    assert header["Claim period"] == "2026-Q2"
    assert header["Line count"] == 1

    rows = list(wb["Lines"].iter_rows(values_only=True))
    assert rows[0][0] == "Supplier" and rows[0][6] == "VAT EUR"
    total = rows[-1]
    assert total[0] == "TOTAL"
    # §4.10 — the TOTAL row equals the claim's frozen VAT base, exactly.
    assert Decimal(str(total[6])) == Decimal(claim["vat_eur"]) == Decimal("420.00")


@pytest.mark.asyncio
async def test_wo77_evidence_pack_download_parses_with_a_matching_manifest(client, db_session):
    import csv
    import hashlib

    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    claim, _entity = await _submitted_claim(client, db_session, headers, org_id)

    r = await client.get(f"{V}/transport/claims/{claim['id']}/evidence", headers=headers)
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == "application/zip"
    assert f"claim-{claim['id']}-evidence.zip" in r.headers["content-disposition"]
    assert r.headers["x-content-type-options"] == "nosniff"

    zf = zipfile.ZipFile(io.BytesIO(r.content))
    names = zf.namelist()
    assert any(n.endswith("claim-workbook.xlsx") for n in names)
    assert any(n.endswith("manifest.csv") for n in names)
    assert len(names) >= 3  # workbook + >=1 vaulted invoice document + manifest

    manifest_name = next(n for n in names if n.endswith("manifest.csv"))
    reader = list(csv.reader(io.StringIO(zf.read(manifest_name).decode())))
    assert reader[0] == ["file", "sha256", "size_bytes"]
    assert len(reader) >= 3  # header + workbook + document
    for path, sha, size in reader[1:]:
        data = zf.read(path)
        assert hashlib.sha256(data).hexdigest() == sha
        assert int(size) == len(data)


@pytest.mark.asyncio
async def test_wo77_artifacts_refuse_a_draft_claim_with_claim_not_frozen(client, db_session):
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    claim, _entity = await _draft_claim(client, db_session, headers, org_id)

    for path in ("workbook", "evidence"):
        r = await client.get(f"{V}/transport/claims/{claim['id']}/{path}", headers=headers)
        assert r.status_code == 409, f"{path}: {r.text}"
        assert r.json()["code"] == "claim_not_frozen"


@pytest.mark.asyncio
async def test_wo77_artifacts_refuse_a_drifted_header_and_a_synthetic_frozen_line(
    client, db_session
):
    """§4.10 / R3 over HTTP — the two fail-CLOSED refusals a filing artifact
    must never paper over."""
    from sqlalchemy import select

    from app.models.transport.vat_claim import VatRefundClaim, VatRefundClaimLine

    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    claim, _entity = await _submitted_claim(client, db_session, headers, org_id)

    db_session.expire_all()
    row = await db_session.scalar(select(VatRefundClaim).where(VatRefundClaim.id == claim["id"]))
    row.vat_eur = Decimal("999.99")  # simulated corruption of the frozen base
    await db_session.commit()

    drift = await client.get(f"{V}/transport/claims/{claim['id']}/workbook", headers=headers)
    assert drift.status_code == 409, drift.text
    assert drift.json()["code"] == "claim_totals_drift"

    row.vat_eur = Decimal("420.00")  # restore, then tamper the LINE instead
    line = await db_session.scalar(
        select(VatRefundClaimLine).where(VatRefundClaimLine.claim_id == claim["id"])
    )
    line.vat_id = "INPUT"  # the direct tamper the G2.12 suite uses (R3's subject)
    await db_session.commit()

    synthetic = await client.get(f"{V}/transport/claims/{claim['id']}/workbook", headers=headers)
    assert synthetic.status_code == 409, synthetic.text
    assert synthetic.json()["code"] == "synthetic_line_in_pack"


@pytest.mark.asyncio
async def test_wo77_evidence_refuses_missing_vault_bytes_while_the_workbook_still_renders(
    client, db_session
):
    """A dangling vault reference refuses the BUNDLE (an incomplete filing
    bundle is worse than none) — the workbook carries figures, not documents,
    and still downloads."""
    from sqlalchemy import select

    from app.models.extraction_run import ExtractionRun

    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    claim, _entity = await _submitted_claim(client, db_session, headers, org_id)

    db_session.expire_all()
    run = await db_session.scalar(select(ExtractionRun).where(ExtractionRun.org_id == org_id))
    run.source_sha256 = "a" * 64  # bytes were never stored under this digest
    await db_session.commit()

    wb = await client.get(f"{V}/transport/claims/{claim['id']}/workbook", headers=headers)
    assert wb.status_code == 200, wb.text

    ev = await client.get(f"{V}/transport/claims/{claim['id']}/evidence", headers=headers)
    assert ev.status_code == 409, ev.text
    assert ev.json()["code"] == "evidence_document_unavailable"


# --------------------------------------------------------------------------- #
# Module gate / authorization / tenancy
# --------------------------------------------------------------------------- #


def _all_new_calls(some_id: str, entity_id: str) -> list[tuple[str, str, dict | None]]:
    """Every route WO-77 adds, with a syntactically-valid body."""
    return [
        ("GET", f"{V}/transport/claims/{some_id}/waivers", None),
        ("POST", f"{V}/transport/claims/{some_id}/waivers", {"supplier": "Q8"}),
        ("DELETE", f"{V}/transport/claims/{some_id}/waivers/Q8", None),
        ("POST", f"{V}/transport/claims/{some_id}/status-code", {"code": "2A"}),
        ("GET", f"{V}/transport/claims/{some_id}/workbook", None),
        ("GET", f"{V}/transport/claims/{some_id}/evidence", None),
        ("GET", f"{V}/transport/status-codes", None),
        ("GET", f"{V}/transport/checklist-rules", None),
        ("POST", f"{V}/transport/checklist-rules/seed", None),
        ("POST", f"{V}/transport/checklist-rules/customer_data/active", {"active": False}),
        ("GET", f"{V}/transport/cadences", None),
        ("PUT", f"{V}/transport/cadences/Q8", {"cadence": "monthly"}),
        ("DELETE", f"{V}/transport/cadences/Q8", None),
        ("GET", f"{V}/transport/receipt-controls?period=2026-05", None),
        ("POST", f"{V}/transport/receipt-controls/{some_id}/override", {"waived": True}),
        ("GET", f"{V}/transport/note-overrides", None),
        (
            "PUT",
            f"{V}/transport/note-overrides",
            {
                "entity_id": entity_id,
                "supplier": "Q8",
                "refund_country": "LV",
                "invoice_ref": "REF-1",
                "target_invoice_id": some_id,
            },
        ),
        ("GET", f"{V}/transport/tie-out-expectations?period=2026-05", None),
        (
            "PUT",
            f"{V}/transport/tie-out-expectations",
            {
                "entity_id": entity_id,
                "supplier": "Q8",
                "period": "2026-05",
                "currency": "EUR",
                "expected_lines": 1,
            },
        ),
        (
            "DELETE",
            f"{V}/transport/tie-out-expectations"
            f"?entity_id={entity_id}&supplier=Q8&period=2026-05&currency=EUR",
            None,
        ),
        ("GET", f"{V}/transport/customers/{entity_id}/lifecycle", None),
        ("POST", f"{V}/transport/customers/{entity_id}/prospect", None),
        ("POST", f"{V}/transport/customers/{entity_id}/promote", None),
        ("POST", f"{V}/transport/customers/{entity_id}/activation", {"active": True}),
        ("POST", f"{V}/transport/customers/{entity_id}/inactive", None),
        ("POST", f"{V}/transport/customers/{entity_id}/countries/LV/request", None),
        (
            "POST",
            f"{V}/transport/customers/{entity_id}/countries/LV/activation",
            {"active": True},
        ),
    ]


async def _call(client, method: str, path: str, body, headers):
    if method == "GET":
        return await client.get(path, headers=headers)
    if method == "DELETE":
        return await client.request("DELETE", path, headers=headers)
    if method == "PUT":
        return await client.put(path, headers=headers, json=body)
    return await client.post(path, headers=headers, json=body)


@pytest.mark.asyncio
async def test_wo77_module_disabled_is_403_module_not_enabled_on_every_route(client, db_session):
    """ADR-P3 rule 3: an org that has not activated the vertical gets 403
    with the SERVICE's `module_not_enabled` code on every new route — the
    owner holds every permission, so this is purely the entitlement."""
    headers, _org_id = await _register_org(client)  # transport NOT enabled
    some_id, entity_id = str(uuid.uuid4()), str(uuid.uuid4())
    for method, path, body in _all_new_calls(some_id, entity_id):
        r = await _call(client, method, path, body, headers)
        assert r.status_code == 403, f"{method} {path}: {r.status_code} {r.text}"
        assert r.json()["code"] == "module_not_enabled", f"{method} {path}: {r.text}"


@pytest.mark.asyncio
async def test_wo77_missing_permission_is_403_structural(client, db_session, role_client):
    """Deny-by-default live on the wire: EMPLOYEE (no VAT_READ) is refused
    the reads; ACCOUNTANT (VAT_WRITE without VAT_SUBMIT) configures freely
    but cannot touch the claim-STATUS surface."""
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    claim, _entity = await _submitted_claim(client, db_session, headers, org_id)

    employee = await role_client("user")
    for path in (
        f"{V}/transport/checklist-rules",
        f"{V}/transport/cadences",
        f"{V}/transport/status-codes",
        f"{V}/transport/note-overrides",
    ):
        r = await employee.get(path)
        assert r.status_code == 403, f"{path}: {r.text}"

    acct = await _member_with_role(client, headers, db_session, "accountant")
    assert (await client.get(f"{V}/transport/checklist-rules", headers=acct)).status_code == 200
    cadence = await client.put(
        f"{V}/transport/cadences/DKV", headers=acct, json={"cadence": "monthly"}
    )
    assert cadence.status_code == 200, cadence.text  # VAT_WRITE granted
    seeded = await client.post(f"{V}/transport/checklist-rules/seed", headers=acct)
    assert seeded.status_code == 200, seeded.text

    # The claim-STATUS surface is VAT_SUBMIT — denied BEFORE any service work.
    denied = await client.post(
        f"{V}/transport/claims/{claim['id']}/status-code", headers=acct, json={"code": "2A"}
    )
    assert denied.status_code == 403, denied.text


@pytest.mark.asyncio
async def test_wo77_cross_tenant_ids_are_opaque_404s(client, db_session):
    """§4.4: org B (transport-enabled, full permissions) probing org A's
    claim / entity / control-row ids gets 404 — never 403."""
    headers_a, org_a = await _register_org(client)
    await _enable_transport(db_session, org_a)
    claim_a, entity_a = await _submitted_claim(client, db_session, headers_a, org_a)
    await receipt_control.run_receipt_control(db_session, org_a, "2026-05")
    await db_session.commit()
    grid_a = await client.get(
        f"{V}/transport/receipt-controls", headers=headers_a, params={"period": "2026-05"}
    )
    control_a = grid_a.json()[0]["id"]

    headers_b, org_b = await _register_org(client)
    await _enable_transport(db_session, org_b)

    cid = claim_a["id"]
    claim_scoped = [
        ("GET", f"{V}/transport/claims/{cid}/waivers", None, "claim_not_found"),
        ("POST", f"{V}/transport/claims/{cid}/waivers", {"supplier": "Q8"}, "claim_not_found"),
        ("DELETE", f"{V}/transport/claims/{cid}/waivers/Q8", None, "claim_not_found"),
        ("POST", f"{V}/transport/claims/{cid}/status-code", {"code": "2A"}, "claim_not_found"),
        ("GET", f"{V}/transport/claims/{cid}/workbook", None, "claim_not_found"),
        ("GET", f"{V}/transport/claims/{cid}/evidence", None, "claim_not_found"),
        (
            "GET",
            f"{V}/transport/customers/{entity_a.id}/lifecycle",
            None,
            "entity_not_found",
        ),
        (
            "POST",
            f"{V}/transport/customers/{entity_a.id}/prospect",
            None,
            "entity_not_found",
        ),
        (
            "POST",
            f"{V}/transport/receipt-controls/{control_a}/override",
            {"waived": True},
            "receipt_control_not_found",
        ),
    ]
    for method, path, body, code in claim_scoped:
        r = await _call(client, method, path, body, headers_b)
        assert r.status_code != 403, f"{method} {path}: 403 confirms the object exists (§4.4)"
        assert r.status_code == 404, f"{method} {path}: {r.status_code} {r.text}"
        assert r.json()["code"] == code, f"{method} {path}: {r.text}"

    # B's org-level listings show ZERO of A's rows.
    for path in (
        f"{V}/transport/checklist-rules",
        f"{V}/transport/cadences",
        f"{V}/transport/note-overrides",
    ):
        listing = await client.get(path, headers=headers_b)
        assert listing.status_code == 200 and listing.json() == [], path
    grid_b = await client.get(
        f"{V}/transport/receipt-controls", headers=headers_b, params={"period": "2026-05"}
    )
    assert grid_b.json() == []

    # A's own view is untouched by B's probing.
    still = await client.get(
        f"{V}/transport/receipt-controls", headers=headers_a, params={"period": "2026-05"}
    )
    assert {r["id"] for r in still.json()} == {r["id"] for r in grid_a.json()}


@pytest.mark.asyncio
async def test_wo77_reads_are_advisory_and_mutate_nothing(client, db_session):
    """§4.19 on the wire: every new READ leaves the claim exactly as it was —
    same status, same code, same frozen base."""
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    claim, entity = await _submitted_claim(client, db_session, headers, org_id)

    before = (await client.get(f"{V}/transport/claims/{claim['id']}", headers=headers)).json()
    for path in (
        f"{V}/transport/claims/{claim['id']}/waivers",
        f"{V}/transport/claims/{claim['id']}/workbook",
        f"{V}/transport/claims/{claim['id']}/evidence",
        f"{V}/transport/status-codes",
        f"{V}/transport/checklist-rules",
        f"{V}/transport/cadences",
        f"{V}/transport/note-overrides",
        f"{V}/transport/receipt-controls?period=2026-05",
        f"{V}/transport/tie-out-expectations?period=2026-05",
        f"{V}/transport/customers/{entity.id}/lifecycle",
    ):
        r = await client.get(path, headers=headers)
        assert r.status_code == 200, f"{path}: {r.text}"

    after = (await client.get(f"{V}/transport/claims/{claim['id']}", headers=headers)).json()
    assert after == before
