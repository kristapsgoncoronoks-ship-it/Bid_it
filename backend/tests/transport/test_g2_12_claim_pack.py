"""G2.12 (WO-74) — the Excel claim workbook + the ZIP evidence pack, both
rendered from ONE loaded pack, both refusing any synthetic line via THE single
`claim_gates.is_synthetic()` predicate (R3's fourth named consumer — "the
workbook builder"). See `app.services.transport.claim_pack`'s module docstring
for every fail-closed refusal this file proves.
"""

from __future__ import annotations

import inspect
import io
import zipfile
from datetime import date
from decimal import Decimal

import pytest
from sqlalchemy import select

from app.core.errors import AppError
from app.core.storage import sha256_hex
from app.models.invoice import Invoice
from app.models.transport.vat_claim import VatRefundClaim, VatRefundClaimLine
from app.models.vendor import Vendor
from app.services import documents, extraction, modules
from app.services.transport import claim as claim_svc
from app.services.transport import claim_lines, claim_pack, fuel_ingest, lock
from tests.factories.transport import synthetic_vehicle_ref
from tests.transport.conftest import activate_entity, enable_transport, make_entity, make_org

TODAY = date(2026, 8, 6)  # after 2026-Q2's period end — the R7 test seam


def _load_wb(data: bytes):
    from openpyxl import load_workbook

    return load_workbook(io.BytesIO(data))


async def _make_claim(db_session, org, entity, **overrides) -> VatRefundClaim:
    kwargs = dict(refund_country="LV", ref_period="2026-Q2")
    kwargs.update(overrides)
    return await claim_svc.get_or_create_claim(db_session, org.id, entity_id=entity.id, **kwargs)


async def _make_vendor(db_session, org, *, name: str = "Q8", country: str | None = "LV") -> Vendor:
    vendor = Vendor(org_id=org.id, name=name, country=country)
    db_session.add(vendor)
    await db_session.flush()
    return vendor


async def _make_invoice(db_session, org, vendor, *, number: str) -> Invoice:
    inv = Invoice(
        org_id=org.id,
        vendor_id=vendor.id,
        invoice_number=number,
        issue_date=date(2026, 5, 1),
        currency="EUR",
        subtotal=Decimal("2000.00"),
        tax_amount=Decimal("420.00"),
        total=Decimal("2420.00"),
    )
    db_session.add(inv)
    await db_session.flush()
    return inv


async def _make_txn(
    db_session,
    org,
    entity,
    *,
    supplier: str = "Q8",
    invoice_ref: str | None = "INV-0001",
    line_seq: int = 1,
    net_eur: Decimal = Decimal("2000.00"),
    vat_eur: Decimal = Decimal("420.00"),
):
    return await fuel_ingest.ingest_transaction(
        db_session,
        org.id,
        entity_id=entity.id,
        supplier=supplier,
        period="2026-05",
        line_seq=line_seq,
        country="LV",
        vehicle_ref=synthetic_vehicle_ref(),
        txn_date=date(2026, 5, 15),
        station="Demo Station Riga",
        product="DIESEL",
        qty=Decimal("100.000"),
        currency="EUR",
        net_local=net_eur,
        vat_local=vat_eur,
        gross_local=net_eur + vat_eur,
        net_eur=net_eur,
        vat_eur=vat_eur,
        invoice_ref=invoice_ref,
    )


async def _attach_document(
    db_session,
    org,
    invoice,
    *,
    data: bytes | None = b"%PDF-1.4 synthetic fixture document",
    sha256: str | None = None,
    filename: str = "q8-inv-0001.pdf",
) -> str:
    """Vault a capture for `invoice`. With `data`, the bytes really land in
    the (in-memory) object store; passing an explicit `sha256` with
    `data=None` records a DANGLING reference — the WO-74
    `evidence_document_unavailable` fixture."""
    if data is not None:
        sha, _size = await documents.store(documents.UPLOADS, org.id, data)
    else:
        assert sha256 is not None
        sha = sha256
    run = await extraction.record(
        db_session, org.id, filename=filename, sha256=sha, method="pdf-text", status="parsed"
    )
    await extraction.link_to_invoice(db_session, org.id, run.id, invoice.id)
    await db_session.commit()
    return sha


async def _submitted_claim(
    db_session,
    org,
    entity,
    *,
    vendor_name: str = "Q8",
    invoice_number: str = "INV-0001",
    doc_data: bytes | None = b"%PDF-1.4 synthetic fixture document",
    doc_sha256: str | None = None,
    doc_filename: str = "q8-inv-0001.pdf",
) -> tuple[VatRefundClaim, Invoice, str]:
    """The G2.12 happy-path fixture: one resolved, documented, frozen line
    (net 2000.00 / VAT 420.00 EUR — above LV's Art. 17 EUR-base quarterly
    minimum, so no override is needed)."""
    vendor = await _make_vendor(db_session, org, name=vendor_name)
    inv = await _make_invoice(db_session, org, vendor, number=invoice_number)
    claim = await _make_claim(db_session, org, entity)
    await db_session.commit()
    txn = await _make_txn(db_session, org, entity, supplier=vendor_name, invoice_ref=invoice_number)
    await db_session.commit()
    sha = await _attach_document(
        db_session, org, inv, data=doc_data, sha256=doc_sha256, filename=doc_filename
    )
    await claim_lines.build_claim_lines(db_session, org.id, claim.id)
    await db_session.commit()
    await lock.submit_claim(
        db_session,
        org.id,
        claim_id=claim.id,
        invoices=[(vendor_name, invoice_number, txn.id)],
        today=TODAY,
    )
    await db_session.commit()
    return claim, inv, sha


async def _ready_org(db_session):
    org = await make_org(db_session)
    await enable_transport(db_session, org.id)
    entity = await make_entity(db_session, org.id)
    await activate_entity(db_session, org.id, entity.id, "LV")
    await db_session.commit()
    return org, entity


# ---------------------------------------------------------------------------
# Refusals — every one fails CLOSED
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_g2_12_draft_claim_refuses_claim_not_frozen(db_session):
    org, entity = await _ready_org(db_session)
    claim = await _make_claim(db_session, org, entity)
    await db_session.commit()
    await _make_txn(db_session, org, entity)
    await db_session.commit()
    await claim_lines.build_claim_lines(db_session, org.id, claim.id)
    await db_session.commit()

    for builder in (claim_pack.build_workbook, claim_pack.build_evidence_pack):
        with pytest.raises(AppError) as exc:
            await builder(db_session, org.id, claim.id)
        assert exc.value.code == "claim_not_frozen"
        assert exc.value.status == 409


@pytest.mark.asyncio
async def test_g2_12_synthetic_unmatched_frozen_line_refuses_both_artifacts(db_session):
    """R3, the acceptance rule verbatim: 'a pack containing ANY synthetic
    line cannot be filed' — BOTH builders refuse with the same code.

    WO-75 fixture note: submission now applies the R3 synthetic lock gate
    itself (`lock.submit_claim` -> `claim_gates.enforce_no_synthetic_lines`),
    so a frozen UNMATCHED line can no longer be seeded through the legal
    path this test originally used. The corrupt state is seeded by DIRECT
    tamper on the frozen row instead (this file's own `line.vat_id = "INPUT"`
    precedent) — which is exactly what this gate now defends against:
    defence-in-depth, all four R3 surfaces block independently. The
    assertion set is unchanged."""
    org, entity = await _ready_org(db_session)
    claim, _inv, _sha = await _submitted_claim(db_session, org, entity)
    line = await db_session.scalar(
        select(VatRefundClaimLine).where(VatRefundClaimLine.claim_id == claim.id)
    )
    line.invoice_ref = "UNMATCHED"  # simulated post-freeze corruption
    line.invoice_id = None
    await db_session.commit()

    for builder in (claim_pack.build_workbook, claim_pack.build_evidence_pack):
        with pytest.raises(AppError) as exc:
            await builder(db_session, org.id, claim.id)
        assert exc.value.code == "synthetic_line_in_pack"
        assert "UNMATCHED" in exc.value.message


@pytest.mark.asyncio
async def test_g2_12_input_in_vat_id_refuses(db_session):
    """The second `is_synthetic()` input: a placeholder VAT id (`INPUT`)
    marks the line synthetic even when its ref looks real."""
    org, entity = await _ready_org(db_session)
    claim, _inv, _sha = await _submitted_claim(db_session, org, entity)
    line = await db_session.scalar(
        select(VatRefundClaimLine).where(VatRefundClaimLine.claim_id == claim.id)
    )
    line.vat_id = "INPUT"
    await db_session.commit()

    with pytest.raises(AppError) as exc:
        await claim_pack.build_workbook(db_session, org.id, claim.id)
    assert exc.value.code == "synthetic_line_in_pack"


@pytest.mark.asyncio
async def test_g2_12_tampered_header_total_refuses_claim_totals_drift(db_session):
    """Master-context §4.10 — the renderer recomputes the totals and refuses
    a frozen header its own lines cannot reproduce."""
    org, entity = await _ready_org(db_session)
    claim, _inv, _sha = await _submitted_claim(db_session, org, entity)
    claim.vat_eur = Decimal("999.99")  # simulated corruption of the frozen base
    await db_session.commit()

    with pytest.raises(AppError) as exc:
        await claim_pack.build_workbook(db_session, org.id, claim.id)
    assert exc.value.code == "claim_totals_drift"


@pytest.mark.asyncio
async def test_g2_12_missing_storage_bytes_refuse_evidence_document_unavailable(db_session):
    """A dangling vault reference (run recorded, object-store bytes gone)
    refuses the BUNDLE — an incomplete filing bundle is worse than none. The
    workbook alone still renders (it carries figures, not documents)."""
    org, entity = await _ready_org(db_session)
    claim, _inv, _sha = await _submitted_claim(
        db_session, org, entity, doc_data=None, doc_sha256="a" * 64
    )

    assert await claim_pack.build_workbook(db_session, org.id, claim.id)  # figures still render
    with pytest.raises(AppError) as exc:
        await claim_pack.build_evidence_pack(db_session, org.id, claim.id)
    assert exc.value.code == "evidence_document_unavailable"


@pytest.mark.asyncio
async def test_g2_12_module_disabled_refuses(db_session):
    """The module-entitlement gate is checked FIRST — an org that turned the
    vertical off cannot render artifacts for an existing claim either (the
    transport-wide service-gate pattern; the route batch adds role authz)."""
    org, entity = await _ready_org(db_session)
    claim, _inv, _sha = await _submitted_claim(db_session, org, entity)
    await modules.set_enabled(db_session, org.id, "transport", False)
    await db_session.commit()

    for builder in (claim_pack.build_workbook, claim_pack.build_evidence_pack):
        with pytest.raises(AppError) as exc:
            await builder(db_session, org.id, claim.id)
        assert exc.value.code == "module_not_enabled"


@pytest.mark.asyncio
async def test_g2_12_cross_tenant_claim_id_is_an_opaque_404(db_session):
    """Master-context §4.4 — org B's claim id under org A yields the same
    404 an unknown id would, never a 403."""
    org_a, entity_a = await _ready_org(db_session)
    org_b, entity_b = await _ready_org(db_session)
    claim_b, _inv, _sha = await _submitted_claim(db_session, org_b, entity_b)
    assert entity_a.id  # sanity: org A exists and is fully set up

    with pytest.raises(AppError) as exc:
        await claim_pack.build_workbook(db_session, org_a.id, claim_b.id)
    assert exc.value.code == "claim_not_found"
    assert exc.value.status == 404


# ---------------------------------------------------------------------------
# Workbook correctness (financial)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_g2_12_workbook_lines_and_total_equal_the_frozen_figures(db_session):
    org, entity = await _ready_org(db_session)
    claim, inv, _sha = await _submitted_claim(db_session, org, entity)
    await db_session.refresh(claim)
    assert claim.vat_eur == Decimal("420.00")  # the frozen base this must equal

    wb = _load_wb(await claim_pack.build_workbook(db_session, org.id, claim.id))
    ws = wb["Lines"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == (
        "Supplier",
        "Invoice number",
        "Invoice date",
        "Goods code",
        "Product group",
        "Net EUR",
        "VAT EUR",
        "Net local",
        "VAT local",
        "Currency",
    )
    assert len(rows) == 3  # header + one R2-grain line + TOTAL
    line = rows[1]
    assert line[0] == "Q8"
    assert line[1] == inv.invoice_number
    assert line[2] == inv.issue_date.isoformat()
    assert line[3] == "1"  # Art. 9: DIESEL -> goods code 1, never 9 (R11)
    assert Decimal(str(line[5])) == Decimal("2000.00")
    assert Decimal(str(line[6])) == Decimal("420.00")
    total = rows[2]
    assert total[0] == "TOTAL"
    assert Decimal(str(total[6])) == claim.vat_eur
    assert Decimal(str(total[8])) == claim.vat_local

    header = {r[0]: r[1] for r in wb["Claim"].iter_rows(values_only=True)}
    assert Decimal(str(header["VAT total EUR"])) == claim.vat_eur
    assert header["Refund country"] == "LV"
    assert header["Claim period"] == "2026-Q2"
    assert header["Period label"] == "Q2"
    assert header["Status"] == "submitted"
    assert header["Line count"] == 1


@pytest.mark.asyncio
async def test_g2_12_row_order_is_deterministic(db_session):
    """Two invoices, two product groups -> rows sorted by (invoice_ref,
    product_group), identical across renders."""
    org, entity = await _ready_org(db_session)
    vendor = await _make_vendor(db_session, org)
    inv_a = await _make_invoice(db_session, org, vendor, number="INV-0001")
    inv_b = await _make_invoice(db_session, org, vendor, number="INV-0002")
    claim = await _make_claim(db_session, org, entity)
    await db_session.commit()
    txn = await _make_txn(db_session, org, entity, invoice_ref="INV-0002", line_seq=1)
    await _make_txn(
        db_session,
        org,
        entity,
        invoice_ref="INV-0001",
        line_seq=2,
        net_eur=Decimal("500.00"),
        vat_eur=Decimal("105.00"),
    )
    await db_session.commit()
    await _attach_document(db_session, org, inv_a, filename="a.pdf")
    await _attach_document(db_session, org, inv_b, data=b"%PDF-1.4 second doc", filename="b.pdf")
    await claim_lines.build_claim_lines(db_session, org.id, claim.id)
    await db_session.commit()
    await lock.submit_claim(
        db_session, org.id, claim_id=claim.id, invoices=[("Q8", "INV-0002", txn.id)], today=TODAY
    )
    await db_session.commit()

    wb1 = _load_wb(await claim_pack.build_workbook(db_session, org.id, claim.id))
    wb2 = _load_wb(await claim_pack.build_workbook(db_session, org.id, claim.id))
    rows1 = list(wb1["Lines"].iter_rows(values_only=True))
    rows2 = list(wb2["Lines"].iter_rows(values_only=True))
    assert rows1 == rows2
    assert [r[1] for r in rows1[1:-1]] == ["INV-0001", "INV-0002"]


@pytest.mark.asyncio
async def test_g2_12_withdrawn_claim_still_renders_its_frozen_pack(db_session):
    """WO-74 design decision: withdraw releases LOCKS, never frozen lines —
    reproducing what WAS filed is audit-trail behaviour."""
    org, entity = await _ready_org(db_session)
    claim, _inv, _sha = await _submitted_claim(db_session, org, entity)
    await lock.withdraw_claim(db_session, org.id, claim.id)
    await db_session.commit()

    wb = _load_wb(await claim_pack.build_workbook(db_session, org.id, claim.id))
    header = {r[0]: r[1] for r in wb["Claim"].iter_rows(values_only=True)}
    assert header["Status"] == "withdrawn"
    assert Decimal(str(header["VAT total EUR"])) == Decimal("420.00")


# ---------------------------------------------------------------------------
# Formula-injection safety (CWE-1236) — the ONE shared sanitize_cell
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_g2_12_free_text_cells_are_formula_injection_safe(db_session):
    org, entity = await _ready_org(db_session)
    claim, _inv, _sha = await _submitted_claim(
        db_session, org, entity, vendor_name="=EvilVendor", invoice_number="=INV0002EVIL"
    )

    wb = _load_wb(await claim_pack.build_workbook(db_session, org.id, claim.id))
    line = list(wb["Lines"].iter_rows(values_only=True))[1]
    assert line[0] == "'=EvilVendor"  # supplier neutralised
    assert line[1] == "'=INV0002EVIL"  # invoice number neutralised
    # money cells stay NUMERIC — never quote-prefixed (a leading '-' there is
    # a negative amount, not an injection; core.csv_safety's own rule).
    assert isinstance(line[5], int | float)
    assert isinstance(line[6], int | float)


@pytest.mark.asyncio
async def test_g2_12_manifest_filenames_are_formula_injection_safe(db_session):
    org, entity = await _ready_org(db_session)
    claim, _inv, _sha = await _submitted_claim(
        db_session, org, entity, vendor_name="=EvilVendor", invoice_number="=INV0002EVIL"
    )

    data = await claim_pack.build_evidence_pack(db_session, org.id, claim.id)
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        manifest_name = next(n for n in zf.namelist() if n.endswith("manifest.csv"))
        manifest = zf.read(manifest_name).decode()
    for row in manifest.splitlines()[1:]:
        assert not row.startswith(("=", "+", "-", "@"))


# ---------------------------------------------------------------------------
# The evidence pack — bundle shape + the G2.12 acceptance (identity)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_g2_12_bundle_paths_follow_the_vault_tree(db_session):
    """§3.K K6 / §3.M M1 — `<Customer> <RegNo>/<Year>/<Country>/<Period>/`."""
    org, entity = await _ready_org(db_session)
    claim, _inv, _sha = await _submitted_claim(db_session, org, entity)

    data = await claim_pack.build_evidence_pack(db_session, org.id, claim.id)
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        names = zf.namelist()

    prefix = f"{entity.legal_name} {entity.registration_number}/2026/LV/Q2/"
    assert all(n.startswith(prefix) for n in names)
    assert f"{prefix}claim-workbook.xlsx" in names
    assert f"{prefix}manifest.csv" in names
    assert any(n.endswith("q8-inv-0001.pdf") for n in names)


@pytest.mark.asyncio
async def test_g2_12_document_bytes_and_manifest_hashes_are_real(db_session):
    org, entity = await _ready_org(db_session)
    payload = b"%PDF-1.4 synthetic fixture document"
    claim, _inv, sha = await _submitted_claim(db_session, org, entity, doc_data=payload)
    assert sha == sha256_hex(payload)

    data = await claim_pack.build_evidence_pack(db_session, org.id, claim.id)
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        doc_name = next(n for n in zf.namelist() if n.endswith(".pdf"))
        assert zf.read(doc_name) == payload  # byte-identical to what was vaulted
        manifest_name = next(n for n in zf.namelist() if n.endswith("manifest.csv"))
        manifest = zf.read(manifest_name).decode()
        # every non-manifest entry appears in the manifest with its REAL hash
        for name in zf.namelist():
            if name == manifest_name:
                continue
            assert sha256_hex(zf.read(name)) in manifest
    assert sha in manifest  # the vaulted document's hash, verifiable offline


@pytest.mark.asyncio
async def test_g2_12_same_document_bytes_are_bundled_once(db_session):
    """Two captures of the SAME bytes on one invoice dedupe by SHA-256 —
    one file in the bundle, not two."""
    org, entity = await _ready_org(db_session)
    claim, inv, _sha = await _submitted_claim(db_session, org, entity)
    await _attach_document(db_session, org, inv, filename="recapture.pdf")  # same default bytes

    data = await claim_pack.build_evidence_pack(db_session, org.id, claim.id)
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        doc_names = [n for n in zf.namelist() if n.endswith(".pdf")]
    assert len(doc_names) == 1


@pytest.mark.asyncio
async def test_g2_12_workbook_and_evidence_pack_agree_line_for_line(db_session):
    """THE G2.12 acceptance (`ARCH_plan.md`): 'the workbook and the evidence
    pack agree line-for-line' — every Lines-sheet cell, including the TOTAL
    row, identical between the standalone workbook and the workbook inside
    the bundle."""
    org, entity = await _ready_org(db_session)
    vendor = await _make_vendor(db_session, org)
    inv_a = await _make_invoice(db_session, org, vendor, number="INV-0001")
    inv_b = await _make_invoice(db_session, org, vendor, number="INV-0002")
    claim = await _make_claim(db_session, org, entity)
    await db_session.commit()
    txn = await _make_txn(db_session, org, entity, invoice_ref="INV-0001", line_seq=1)
    await _make_txn(
        db_session,
        org,
        entity,
        invoice_ref="INV-0002",
        line_seq=2,
        net_eur=Decimal("500.00"),
        vat_eur=Decimal("105.00"),
    )
    await db_session.commit()
    await _attach_document(db_session, org, inv_a, filename="a.pdf")
    await _attach_document(db_session, org, inv_b, data=b"%PDF-1.4 second doc", filename="b.pdf")
    await claim_lines.build_claim_lines(db_session, org.id, claim.id)
    await db_session.commit()
    await lock.submit_claim(
        db_session, org.id, claim_id=claim.id, invoices=[("Q8", "INV-0001", txn.id)], today=TODAY
    )
    await db_session.commit()

    standalone = _load_wb(await claim_pack.build_workbook(db_session, org.id, claim.id))
    bundle = await claim_pack.build_evidence_pack(db_session, org.id, claim.id)
    with zipfile.ZipFile(io.BytesIO(bundle)) as zf:
        wb_name = next(n for n in zf.namelist() if n.endswith("claim-workbook.xlsx"))
        embedded = _load_wb(zf.read(wb_name))

    rows_standalone = list(standalone["Lines"].iter_rows(values_only=True))
    rows_embedded = list(embedded["Lines"].iter_rows(values_only=True))
    assert rows_standalone == rows_embedded  # identical lines AND totals
    assert rows_standalone[-1][0] == "TOTAL"
    assert Decimal(str(rows_standalone[-1][6])) == Decimal("525.00")  # 420 + 105


# ---------------------------------------------------------------------------
# Vault-tree sanitization + period labels (pure units)
# ---------------------------------------------------------------------------


def test_g2_12_period_label_maps_the_two_claim_shapes():
    assert claim_pack.period_label("2026-Q2") == "Q2"
    assert claim_pack.period_label("2026-Q4") == "Q4"
    assert claim_pack.period_label("2026-YEAR") == "Annual"


def test_g2_12_segment_sanitization_is_windows_sharepoint_ftp_safe():
    s = claim_pack._sanitize_segment
    assert s('Bad<>:"/\\|?*Name') == "Bad_________Name"
    assert s("Trailing dots...") == "Trailing dots"
    assert s("Ctrl\x00\x1fChars") == "Ctrl__Chars"
    assert s("...") == "_"  # never an empty segment (would collapse the tree)
    assert s("") == "_"
    assert s("Plain Name") == "Plain Name"


@pytest.mark.asyncio
async def test_g2_12_illegal_entity_name_chars_never_reach_zip_paths(db_session):
    org = await make_org(db_session)
    await enable_transport(db_session, org.id)
    entity = await make_entity(db_session, org.id)
    entity.legal_name = 'Evil/Entity:Name*?"'
    await activate_entity(db_session, org.id, entity.id, "LV")
    await db_session.commit()
    claim, _inv, _sha = await _submitted_claim(db_session, org, entity)

    data = await claim_pack.build_evidence_pack(db_session, org.id, claim.id)
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        names = zf.namelist()
    # 4 path segments + filename — the '/' inside the name did NOT add a level
    assert all(len(n.split("/")) == 5 for n in names)
    assert all(not any(c in seg for c in '<>:"\\|?*') for n in names for seg in n.split("/"))


# ---------------------------------------------------------------------------
# Structural — R3's single-predicate discipline
# ---------------------------------------------------------------------------


def test_g2_12_r3_module_reuses_the_one_synthetic_predicate():
    """R3: the workbook builder consumes THE centralized
    `claim_gates.is_synthetic` — and defines no rival predicate."""
    src = inspect.getsource(claim_pack)
    assert "from app.services.transport.claim_gates import is_synthetic" in src
    assert "def is_synthetic" not in src  # no second implementation, ever
    # the predicate is actually CALLED on both of its inputs
    assert "is_synthetic(ln.invoice_ref, ln.vat_id)" in src


def test_g2_12_generation_is_read_only():
    """No artifact generation path writes: the module never flushes, commits,
    adds to the session, or records an audit event (read-only export — the
    `erp_export`/`report_writers` precedent; §4.16 audits mutations only)."""
    src = inspect.getsource(claim_pack)
    for forbidden in ("db.add(", "db.flush(", "db.commit(", "audit.record("):
        assert forbidden not in src
