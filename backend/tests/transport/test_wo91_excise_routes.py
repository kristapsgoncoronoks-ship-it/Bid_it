"""WO-91 — transport routes slice 8: the diesel excise-duty refund over HTTP.

Proves `api/routes/transport/excise.py` end to end — the report and the rate
registry — plus the
route matrix every transport slice ships: the structural permission pair
(EMPLOYEE denied / ACCOUNTANT granted), the read/write split (an AUDITOR reads
but cannot type a rate), the module entitlement (ADR-P3 rule 3), an org-scoped
read over deliberately OVERLAPPING data, the service's own refusal codes on the
wire, and Decimal-as-string (§4.9).

It also asserts what R42's acceptance line requires — *"The UI shows the
indicative-rate and eligibility caveats on **every surface that shows the
number**"* — as a property of the RESPONSE rather than of a screen: the rate
table is one of those surfaces, so it carries `eligibility`,
`eligibility_asserted: false`, `rate_caveat` and `legal_framing`.

Fixture strategy is the WO-82/WO-84/WO-87 suites', unchanged: orgs/tokens ride
the real HTTP register flow, transport enablement is direct service setup (the
module is in no billing plan, so the HTTP toggle would 402), and every assertion
about what the API returned goes through the API.
"""

from __future__ import annotations

import uuid
from datetime import date
from decimal import Decimal

import pytest
from httpx import AsyncClient

from app.core.money import q2
from app.services import modules
from app.services.transport import excise, fuel_ingest
from tests.factories.transport import synthetic_vehicle_ref
from tests.transport.conftest import make_entity

pytestmark = pytest.mark.asyncio

V = "/api/v1"
REPORT = f"{V}/transport/excise"
PACKET = f"{V}/transport/excise/packet"
RATES = f"{V}/transport/excise/rates"
PERIOD = "2026-05"
DAY = date(2026, 5, 12)

_SEQ = {"n": 0}


async def _diesel(db_session, org_id: str, entity_id: str, *, country: str, qty: str) -> None:
    """One validated diesel line, written through the REAL ingestion path — the
    only writer of `fuel_transactions` (WO-50/G1.2)."""
    _SEQ["n"] += 1
    net_eur = Decimal("1500.00")
    vat_eur = q2(net_eur * Decimal("0.21"))
    await fuel_ingest.ingest_transaction(
        db_session,
        org_id,
        entity_id=entity_id,
        supplier="Q8",
        period=PERIOD,
        line_seq=_SEQ["n"],
        country=country,
        vehicle_ref=synthetic_vehicle_ref(),
        txn_date=DAY,
        station="Demo Station Riga",
        product="DIESEL",
        qty=Decimal(qty),
        currency="EUR",
        net_local=net_eur,
        vat_local=vat_eur,
        gross_local=net_eur + vat_eur,
        net_eur=net_eur,
        vat_eur=vat_eur,
    )
    await db_session.commit()


async def _register_org(client: AsyncClient):
    suffix = uuid.uuid4().hex[:8]
    r = await client.post(
        f"{V}/auth/register",
        json={
            "organization_name": f"WO91 Org {suffix}",
            "name": "Owner",
            "email": f"owner-{suffix}@wo91.example.io",
            "password": "supersecret",
        },
    )
    assert r.status_code == 201, r.text
    body = r.json()
    return {"Authorization": f"Bearer {body['token']['access_token']}"}, body["organization"]["id"]


async def _enable_transport(db_session, org_id: str) -> None:
    await modules.set_enabled(db_session, org_id, "transport", True)


async def _member_with_role(client: AsyncClient, owner_headers, db_session, stored_role: str):
    from sqlalchemy import update

    from app.models.user import User, UserRole

    email = f"{stored_role}-{uuid.uuid4().hex[:8]}@wo91.example.io"
    inv = await client.post(
        f"{V}/team/invites", headers=owner_headers, json={"email": email, "role": "admin"}
    )
    assert inv.status_code == 201, inv.text
    acc = await client.post(
        f"{V}/auth/accept-invite",
        json={"token": inv.json()["token"], "name": "Member", "password": "supersecret"},
    )
    assert acc.status_code == 201, acc.text
    body = acc.json()
    await db_session.execute(
        update(User)
        .where(User.id == body["user"]["id"])
        .values(role=UserRole(stored_role), is_expense_approver=False)
    )
    await db_session.commit()
    return {"Authorization": f"Bearer {body['token']['access_token']}"}


def _by_country(payload: dict) -> dict[str, dict]:
    return {row["country"]: row for row in payload["rates"]}


# --------------------------------------------------------------------------- #
# The read surface
# --------------------------------------------------------------------------- #


async def test_wo91_rates_list_every_refund_state_at_the_harvested_default(client, db_session):
    """An organization that has typed nothing still sees the rate its figure
    WILL be computed at — the EUR 30.00 placeholder — rather than an empty
    list, and every row says it is a placeholder."""
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)

    r = await client.get(RATES, headers=headers)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["countries"] == ["BE", "ES", "FR", "HR", "HU", "IT", "SI"]
    assert body["default_rate_eur_per_1000l"] == "30.0000"
    rows = _by_country(body)
    assert set(rows) == set(body["countries"])
    assert all(row["is_override"] is False for row in rows.values())
    assert all(row["rate_eur_per_1000l"] == "30.0000" for row in rows.values())


async def test_wo91_every_rate_surface_carries_the_caveats(client, db_session):
    """R42's acceptance, asserted on the wire: the caveats are REQUIRED fields
    sourced from the service's single constants, so a client cannot receive the
    rate without them and cannot render a softened wording."""
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)

    body = (await client.get(RATES, headers=headers)).json()
    assert body["eligibility"] == excise.ELIGIBILITY_STATEMENT
    assert body["eligibility_asserted"] is False
    assert body["rate_caveat"] == excise.RATE_CAVEAT
    assert body["legal_framing"] == excise.LEGAL_FRAMING
    assert body["currency"] == "EUR"
    assert all(row["rate_caveat"] == excise.RATE_CAVEAT for row in body["rates"])


async def test_wo91_rates_cross_the_wire_as_decimal_strings(client, db_session):
    """§4.9 — a rate is a `Decimal` end to end. `27.1234` must survive verbatim;
    a float round-trip would be visible as a JSON number."""
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)

    put = await client.put(
        RATES, headers=headers, json={"country": "HU", "rate_eur_per_1000l": "27.1234"}
    )
    assert put.status_code == 200, put.text
    row = _by_country(put.json())["HU"]
    assert row["rate_eur_per_1000l"] == "27.1234"
    assert isinstance(row["rate_eur_per_1000l"], str)


# --------------------------------------------------------------------------- #
# The write surface
# --------------------------------------------------------------------------- #


async def test_wo91_setting_and_removing_an_override_round_trips(client, db_session):
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)

    put = await client.put(
        RATES, headers=headers, json={"country": "FR", "rate_eur_per_1000l": "22.5000"}
    )
    assert put.status_code == 200, put.text
    rows = _by_country(put.json())
    assert rows["FR"] == {
        "country": "FR",
        "rate_eur_per_1000l": "22.5000",
        "is_override": True,
        "rate_caveat": excise.RATE_CAVEAT,
    }
    assert rows["IT"]["is_override"] is False

    delete = await client.delete(RATES, headers=headers, params={"country": "FR"})
    assert delete.status_code == 200, delete.text
    after = _by_country(delete.json())["FR"]
    assert after["is_override"] is False
    assert after["rate_eur_per_1000l"] == "30.0000"


async def test_wo91_the_service_refusals_reach_the_wire_verbatim(client, db_session):
    """This module maps nothing: the codes are the service's own, so the wire
    vocabulary cannot drift from the service layer (§4.20)."""
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)

    unsupported = await client.put(
        RATES, headers=headers, json={"country": "LV", "rate_eur_per_1000l": "30.0000"}
    )
    assert unsupported.status_code == 422, unsupported.text
    assert unsupported.json()["code"] == "excise_country_not_supported"

    # `gt=0` is caught by the schema (422 without a service code); the SERVICE
    # refusal is proven on the boundary the schema cannot express — a value that
    # parses and is positive is accepted, and the service's own code appears for
    # a state it rejects. Both halves of "schemas catch shape, services catch
    # business rules" (master-context DoD 3).
    tiny = await client.put(
        RATES, headers=headers, json={"country": "FR", "rate_eur_per_1000l": "0.0001"}
    )
    assert tiny.status_code == 200, tiny.text

    zero = await client.put(
        RATES, headers=headers, json={"country": "FR", "rate_eur_per_1000l": "0"}
    )
    assert zero.status_code == 422, zero.text

    bad_delete = await client.delete(RATES, headers=headers, params={"country": "LV"})
    assert bad_delete.status_code == 422
    assert bad_delete.json()["code"] == "excise_country_not_supported"


# --------------------------------------------------------------------------- #
# The report
# --------------------------------------------------------------------------- #


async def test_wo91_the_report_returns_the_figure_as_decimal_strings(client, db_session):
    """HAND-COMPUTED on the wire: 1,234.567 L in FR at the harvested
    EUR 30.0000/1,000 L is EUR 37.04. Every figure — litres included — crosses
    as an exact decimal STRING (§4.9); a float round-trip of the litre total
    would be visible in the euro."""
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    entity = await make_entity(db_session, org_id, country="LV", seed=41)
    await db_session.commit()
    await _diesel(db_session, org_id, entity.id, country="FR", qty="1234.567")

    r = await client.get(REPORT, headers=headers, params={"period": PERIOD})
    assert r.status_code == 200, r.text
    body = r.json()
    assert len(body["rows"]) == 1
    row = body["rows"][0]
    assert row["litres"] == "1234.567"
    assert row["rate_eur_per_1000l"] == "30.0000"
    assert row["indicative_excise_eur"] == "37.04"
    assert row["rate_is_override"] is False
    assert row["country"] == "FR"
    assert body["indicative_excise_eur"] == "37.04"
    assert body["currency"] == "EUR"
    assert body["product_group"] == "Diesel"
    for value in (row["litres"], row["indicative_excise_eur"], body["indicative_excise_eur"]):
        assert isinstance(value, str)


async def test_wo91_every_report_response_carries_the_eligibility_denial(client, db_session):
    """R42's acceptance on the surface that actually shows the euro. The
    statement and the boolean are the service's own single constants, so a
    softened wording cannot appear downstream."""
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    entity = await make_entity(db_session, org_id, country="LV", seed=42)
    await db_session.commit()
    await _diesel(db_session, org_id, entity.id, country="FR", qty="1000.000")

    body = (await client.get(REPORT, headers=headers, params={"period": PERIOD})).json()
    assert body["eligibility"] == excise.ELIGIBILITY_STATEMENT
    assert body["eligibility_asserted"] is False
    assert body["rate_caveat"] == excise.RATE_CAVEAT
    assert body["legal_framing"] == excise.LEGAL_FRAMING
    assert body["filed_with"] == excise.FILED_WITH
    assert body["litre_basis"] == excise.LITRE_BASIS


async def test_wo91_a_state_outside_the_regime_is_skipped_not_zeroed_on_the_wire(
    client, db_session
):
    """ "We hold no rate for this state" and "this state refunds you nothing" are
    different facts, and the wire keeps them different: the skipped entry has
    litres and a line count and no euro field at all."""
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    entity = await make_entity(db_session, org_id, country="LV", seed=43)
    await db_session.commit()
    await _diesel(db_session, org_id, entity.id, country="LV", qty="9000.000")

    body = (await client.get(REPORT, headers=headers, params={"period": PERIOD})).json()
    assert body["rows"] == []
    assert body["indicative_excise_eur"] == "0.00"
    assert body["skipped_countries"] == [
        {"country": "LV", "litres": "9000.000", "lines": 1},
    ]


async def test_wo91_an_admin_override_moves_the_reported_figure(client, db_session):
    """The two surfaces meet: a rate typed through `PUT /excise/rates` changes
    the euro `GET /excise` reports, and the row says the rate was verified."""
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    entity = await make_entity(db_session, org_id, country="LV", seed=44)
    await db_session.commit()
    await _diesel(db_session, org_id, entity.id, country="FR", qty="1000.000")

    before = (await client.get(REPORT, headers=headers, params={"period": PERIOD})).json()
    assert before["rows"][0]["indicative_excise_eur"] == "30.00"

    put = await client.put(
        RATES, headers=headers, json={"country": "FR", "rate_eur_per_1000l": "22.5000"}
    )
    assert put.status_code == 200, put.text

    after = (await client.get(REPORT, headers=headers, params={"period": PERIOD})).json()
    assert after["rows"][0]["indicative_excise_eur"] == "22.50"
    assert after["rows"][0]["rate_is_override"] is True


async def test_wo91_a_malformed_period_refuses_invalid_period_on_the_wire(client, db_session):
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)

    r = await client.get(REPORT, headers=headers, params={"period": "2026-13"})
    assert r.status_code == 422, r.text
    body = r.json()
    assert body["code"] == "invalid_period"
    assert "YYYY-MM" in body["detail"]


async def test_wo91_an_empty_period_is_a_200_with_no_rows(client, db_session):
    """Never a 404: "you burned no diesel in a refunding state this month" is an
    answer, and the caveats still ride it."""
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)

    r = await client.get(REPORT, headers=headers, params={"period": PERIOD})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["rows"] == [] and body["skipped_countries"] == []
    assert body["indicative_excise_eur"] == "0.00"
    assert body["eligibility_asserted"] is False


async def test_wo91_another_orgs_entity_id_is_an_opaque_empty_scope(client, db_session):
    """§4.4 — org B naming org A's entity must learn nothing. This route's
    `entity_id` is a SCOPE filter rather than a fetch-by-id, so the honest
    answer is an empty report: A's litres are never returned, and B cannot tell
    a foreign entity from one of its own that had no diesel."""
    a_headers, a_org = await _register_org(client)
    b_headers, b_org = await _register_org(client)
    await _enable_transport(db_session, a_org)
    await _enable_transport(db_session, b_org)
    a_entity = await make_entity(db_session, a_org, country="LV", seed=45)
    await db_session.commit()
    await _diesel(db_session, a_org, a_entity.id, country="FR", qty="1000.000")

    mine = await client.get(REPORT, headers=a_headers, params={"period": PERIOD})
    assert mine.json()["indicative_excise_eur"] == "30.00"

    theirs = await client.get(
        REPORT, headers=b_headers, params={"period": PERIOD, "entity_id": a_entity.id}
    )
    assert theirs.status_code != 403, "a 403 would confirm the entity exists (§4.4)"
    assert theirs.status_code == 200, theirs.text
    assert theirs.json()["rows"] == []
    assert theirs.json()["indicative_excise_eur"] == "0.00"

    unscoped = await client.get(REPORT, headers=b_headers, params={"period": PERIOD})
    assert unscoped.json()["rows"] == []


# --------------------------------------------------------------------------- #
# The customs packet
# --------------------------------------------------------------------------- #


async def test_wo91_the_packet_downloads_as_a_workbook_matching_the_report(client, db_session):
    """R42's *"Excel packet for customs"* over HTTP, and the one-source rule on
    the wire: the workbook's TOTAL is the same euro the JSON route returns."""
    import io

    from openpyxl import load_workbook

    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    entity = await make_entity(db_session, org_id, country="LV", seed=46)
    await db_session.commit()
    await _diesel(db_session, org_id, entity.id, country="FR", qty="1234.567")

    r = await client.get(PACKET, headers=headers, params={"period": PERIOD})
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "excise-2026-05-customs-packet.xlsx" in r.headers["content-disposition"]
    assert r.headers["x-content-type-options"] == "nosniff"

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == [excise.PACKET_SHEET_COVER, excise.PACKET_SHEET_LITRES]
    sheet = wb[excise.PACKET_SHEET_LITRES]
    labels = [c.value for c in sheet[1]]
    total = [c.value for c in sheet[3]]
    assert total[0] == excise.TOTAL_LABEL
    packet_eur = Decimal(str(total[labels.index("Indicative excise EUR")]))

    body = (await client.get(REPORT, headers=headers, params={"period": PERIOD})).json()
    assert packet_eur == Decimal(body["indicative_excise_eur"]) == Decimal("37.04")

    # R42's acceptance again — the caveats are on the sheet that shows the
    # number, not only on its cover.
    text = "\n".join(str(c.value) for row in sheet.iter_rows() for c in row if c.value is not None)
    assert excise.ELIGIBILITY_STATEMENT in text
    assert excise.RATE_CAVEAT in text


async def test_wo91_an_empty_period_refuses_the_packet_on_the_wire(client, db_session):
    """422 `no_excise_findings`, fail-CLOSED. Handing back an empty workbook
    would hand back a document that looks like a customs filing."""
    headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)

    r = await client.get(PACKET, headers=headers, params={"period": PERIOD})
    assert r.status_code == 422, r.text
    assert r.json()["code"] == "no_excise_findings"


async def test_wo91_the_packet_is_a_read_not_a_write(client, db_session):
    """`TRANSPORT_READ` alone. Generating an artifact persists nothing, so an
    AUDITOR — who holds no `VAT_WRITE` — can still download it."""
    owner_headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    entity = await make_entity(db_session, org_id, country="LV", seed=47)
    await db_session.commit()
    await _diesel(db_session, org_id, entity.id, country="FR", qty="1000.000")

    auditor = await _member_with_role(client, owner_headers, db_session, "auditor")
    assert (await client.get(PACKET, headers=auditor, params={"period": PERIOD})).status_code == 200

    employee = await _member_with_role(client, owner_headers, db_session, "user")
    assert (
        await client.get(PACKET, headers=employee, params={"period": PERIOD})
    ).status_code == 403


# --------------------------------------------------------------------------- #
# Authorization — the structural permission pair, and the read/write split
# --------------------------------------------------------------------------- #


async def test_wo91_an_employee_is_denied_and_an_accountant_is_granted(client, db_session):
    """`TRANSPORT_READ`, declared on the router (ADR-0024). One role that must
    be refused and one that must be allowed — the definition-of-done §5 pair."""
    owner_headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)

    employee = await _member_with_role(client, owner_headers, db_session, "user")
    accountant = await _member_with_role(client, owner_headers, db_session, "accountant")

    assert (await client.get(RATES, headers=employee)).status_code == 403
    granted = await client.get(RATES, headers=accountant)
    assert granted.status_code == 200, granted.text


async def test_wo91_an_auditor_can_read_a_rate_but_not_type_one(client, db_session):
    """The per-route `VAT_WRITE` override. An AUDITOR holds `TRANSPORT_READ` and
    not `VAT_WRITE`, which is exactly the split configuring master data needs —
    and the SAME existing permission `overcharges.py` gates a contract term
    with. No permission member was invented for excise (§10)."""
    owner_headers, org_id = await _register_org(client)
    await _enable_transport(db_session, org_id)
    auditor = await _member_with_role(client, owner_headers, db_session, "auditor")

    assert (await client.get(RATES, headers=auditor)).status_code == 200
    assert (
        await client.put(
            RATES, headers=auditor, json={"country": "FR", "rate_eur_per_1000l": "22.5000"}
        )
    ).status_code == 403
    assert (
        await client.delete(RATES, headers=auditor, params={"country": "FR"})
    ).status_code == 403


async def test_wo91_an_anonymous_request_is_refused(client):
    assert (await client.get(RATES)).status_code == 401


async def test_wo91_module_disabled_refuses_403_module_not_enabled(client, db_session):
    """ADR-P3 rule 3 — the service gate, fail-CLOSED, surfaced verbatim."""
    headers, _org_id = await _register_org(client)
    r = await client.get(RATES, headers=headers)
    assert r.status_code == 403, r.text
    assert r.json()["code"] == "module_not_enabled"

    w = await client.put(RATES, headers=headers, json={"country": "FR", "rate_eur_per_1000l": "1"})
    assert w.status_code == 403
    assert w.json()["code"] == "module_not_enabled"


# --------------------------------------------------------------------------- #
# Tenancy — identical-looking data in two orgs
# --------------------------------------------------------------------------- #


async def test_wo91_an_org_never_sees_another_orgs_rate(client, db_session):
    """§4.1. Both orgs override the SAME state, so a missing tenant filter
    cannot pass by accident: the values are the only discriminator, and each
    org must read back exactly the rate it typed."""
    a_headers, a_org = await _register_org(client)
    b_headers, b_org = await _register_org(client)
    await _enable_transport(db_session, a_org)
    await _enable_transport(db_session, b_org)

    for headers, rate in ((a_headers, "22.5000"), (b_headers, "27.7500")):
        put = await client.put(
            RATES, headers=headers, json={"country": "FR", "rate_eur_per_1000l": rate}
        )
        assert put.status_code == 200, put.text

    for headers, mine, theirs in (
        (a_headers, "22.5000", "27.7500"),
        (b_headers, "27.7500", "22.5000"),
    ):
        rows = _by_country((await client.get(RATES, headers=headers)).json())
        assert rows["FR"]["rate_eur_per_1000l"] == mine
        assert rows["FR"]["rate_eur_per_1000l"] != theirs
