"""WO-91 — R42's Excel packet for CUSTOMS, and the one-source proof.

R42, verbatim: *"…with an **Excel packet for customs**"*, over a regime
`BA_fleet_fuel.md` §2.4 calls *"Separate regime, filed with **CUSTOMS**"*.

WHAT THIS FILE IS DELIBERATELY STRICT ABOUT
---------------------------------------------
* **the workbook is parsed and compared cell for cell against the analysis** —
  a packet whose litres or euro differ from the API's is a packet whose figures
  a customs officer can contradict with a screenshot;
* **the identity is STRUCTURAL, not coincidental** — asserted by the shape of
  the renderer (a sync function over a loaded report, holding no session) and
  by both projections coming from the ONE `_COLUMNS` spec, the WO-74/WO-83
  precedent;
* **the caveats are on the NUMBERS sheet too**, because a sheet detached from
  its cover page is one of R42's *"surfaces that show the number"*;
* **an empty period REFUSES** rather than emitting a document that looks like a
  filing and supports nothing;
* **free text is formula-injection-safe** (CWE-1236), numeric cells are not
  quoted (a customs officer must be able to re-add the column).
"""

from __future__ import annotations

import inspect
import io
from datetime import date
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from app.core.errors import ValidationError
from app.core.money import q2
from app.services.transport import excise, fuel_ingest
from tests.factories.transport import synthetic_vehicle_ref
from tests.transport.conftest import enable_transport, make_entity, make_org

PERIOD = "2026-05"
DAY = date(2026, 5, 12)

_SEQ = {"n": 0}


async def _org_entity(db_session, *, name: str = "WO91 Packet Org", seed: int = 51):
    org = await make_org(db_session, name=name)
    await enable_transport(db_session, org.id)
    entity = await make_entity(db_session, org.id, country="LV", seed=seed)
    await db_session.commit()
    return org.id, entity.id


async def _diesel(
    db_session, org_id: str, entity_id: str, *, country: str, qty: Decimal, station: str = "Riga"
) -> None:
    _SEQ["n"] += 1
    net_eur = Decimal("1500.00")
    vat_eur = q2(net_eur * Decimal("0.21"))
    await fuel_ingest.ingest_transaction(
        db_session,
        org_id,
        entity_id=entity_id,
        supplier="Q8",
        period=PERIOD,
        line_seq=_SEQ["n"],
        country=country,
        vehicle_ref=synthetic_vehicle_ref(),
        txn_date=DAY,
        station=station,
        product="DIESEL",
        qty=qty,
        currency="EUR",
        net_local=net_eur,
        vat_local=vat_eur,
        gross_local=net_eur + vat_eur,
        net_eur=net_eur,
        vat_eur=vat_eur,
    )
    await db_session.commit()


def _sheets(data: bytes):
    wb = load_workbook(io.BytesIO(data))
    return wb, wb[excise.PACKET_SHEET_COVER], wb[excise.PACKET_SHEET_LITRES]


# --------------------------------------------------------------------------- #
# The one-source proof — structural, not "they happen to agree today"
# --------------------------------------------------------------------------- #


def test_wo91_the_renderer_is_sync_and_holds_no_session():
    """`build_evidence_workbook` takes a LOADED `ExciseReport` and nothing else.
    It is not a coroutine and has no `AsyncSession` parameter, so it cannot
    query a second source even by a future accident — the `overcharge_pack`
    point-3 discipline, applied here."""
    fn = excise.build_evidence_workbook
    assert not inspect.iscoroutinefunction(fn)
    signature = inspect.signature(fn)
    assert list(signature.parameters) == ["report"]
    assert signature.parameters["report"].annotation in ("ExciseReport", excise.ExciseReport)


def test_wo91_the_loader_is_the_analysis_itself():
    """`evidence_workbook` calls `excise_report` — the SAME function the JSON
    route calls — and then renders. There is no second query anywhere in the
    packet path, so the two artifacts cannot look at different rows."""
    source = inspect.getsource(excise.evidence_workbook)
    assert "await excise_report(" in source
    assert "queries." not in source, "the packet loader issues a query of its own"
    assert "build_evidence_workbook(report)" in source


def test_wo91_both_projections_come_from_the_one_column_spec():
    """`packet_row` IS the projection the renderer uses, so a test can assert
    the workbook against it rather than against a re-typed expectation — and a
    column added to `_COLUMNS` reaches both at once."""
    source = inspect.getsource(excise.build_evidence_workbook)
    assert "packet_row(cell_row, report)" in source
    assert "_COLUMNS" in source


# --------------------------------------------------------------------------- #
# Cell for cell against the analysis
# --------------------------------------------------------------------------- #


async def test_wo91_the_workbook_reproduces_the_analysis_cell_for_cell(db_session):
    """R42's packet and the JSON report are one source. Two entities across two
    states, one of them rate-overridden, so every column varies between rows."""
    org = await make_org(db_session, name="WO91 Packet Grid")
    await enable_transport(db_session, org.id)
    a = await make_entity(db_session, org.id, country="LV", seed=61)
    b = await make_entity(db_session, org.id, country="EE", seed=62)
    await db_session.commit()
    await _diesel(db_session, org.id, a.id, country="FR", qty=Decimal("1234.567"))
    await _diesel(db_session, org.id, a.id, country="IT", qty=Decimal("2000.000"))
    await _diesel(db_session, org.id, b.id, country="FR", qty=Decimal("500.500"))
    await excise.set_rate(db_session, org.id, country="IT", rate_eur_per_1000l=Decimal("22.5000"))
    await db_session.commit()

    report = await excise.excise_report(db_session, org.id, period=PERIOD)
    data = await excise.evidence_workbook(db_session, org.id, period=PERIOD)
    _wb, _cover, sheet = _sheets(data)

    headers = [c.value for c in sheet[1]]
    assert headers == [c.label for c in excise._COLUMNS]

    body = [
        [c.value for c in row] for row in sheet.iter_rows(min_row=2, max_row=1 + len(report.rows))
    ]
    expected = [excise.packet_row(cell, report) for cell in report.rows]
    assert len(body) == 3
    for actual, want in zip(body, expected, strict=True):
        for got, exp in zip(actual, want, strict=True):
            # openpyxl round-trips a Decimal through the XLSX number type, so
            # compare on the DECIMAL value rather than the Python object.
            if isinstance(exp, Decimal):
                assert Decimal(str(got)) == exp
            else:
                assert got == exp

    total = [c.value for c in sheet[2 + len(report.rows)]]
    assert total[0] == excise.TOTAL_LABEL
    assert Decimal(str(total[headers.index("Litres")])) == report.litres
    assert Decimal(str(total[headers.index("Indicative excise EUR")])) == (
        report.indicative_excise_eur
    )
    # §4.10 — the printed total is the figure the SERVICE computed, and the
    # detail rows reproduce it EXACTLY. HAND-COMPUTED, cell by cell:
    #   a/FR  1234.567 L @ 30.0000  -> 37.03701 -> 37.04
    #   a/IT  2000.000 L @ 22.5000  -> 45.00000 -> 45.00   (the override)
    #   b/FR   500.500 L @ 30.0000  -> 15.01500 -> 15.02   (ROUND_HALF_UP)
    #                                              -------
    #                                               97.06
    assert report.indicative_excise_eur == Decimal("97.06")
    assert (
        sum((Decimal(str(r[headers.index("Indicative excise EUR")])) for r in body), Decimal("0"))
        == report.indicative_excise_eur
    )


async def test_wo91_the_rate_source_column_distinguishes_verified_from_placeholder(db_session):
    """§9.2 item 13's whole point, on the sheet a customs officer reads: which
    of these rates did a human verify, and which is our placeholder."""
    org_id, entity_id = await _org_entity(db_session, name="WO91 Packet Rates", seed=63)
    await _diesel(db_session, org_id, entity_id, country="FR", qty=Decimal("1000.000"))
    await _diesel(db_session, org_id, entity_id, country="IT", qty=Decimal("1000.000"))
    await excise.set_rate(db_session, org_id, country="FR", rate_eur_per_1000l=Decimal("22.5000"))
    await db_session.commit()

    _wb, _cover, sheet = _sheets(await excise.evidence_workbook(db_session, org_id, period=PERIOD))
    headers = [c.value for c in sheet[1]]
    col = headers.index("Rate source")
    sources = {
        row[headers.index("Country of supply")].value: row[col].value
        for row in sheet.iter_rows(min_row=2, max_row=3)
    }
    assert sources == {"FR": excise.RATE_SOURCE_OVERRIDE, "IT": excise.RATE_SOURCE_DEFAULT}


# --------------------------------------------------------------------------- #
# The caveats — on BOTH sheets
# --------------------------------------------------------------------------- #


async def test_wo91_both_sheets_carry_the_eligibility_and_rate_caveats(db_session):
    """R42: the caveats on *"every surface that shows the number"*. A numbers
    sheet detached from its cover page is such a surface, so it carries them
    too — and both render the SAME constants, so no wording can drift."""
    org_id, entity_id = await _org_entity(db_session, name="WO91 Packet Caveats", seed=64)
    await _diesel(db_session, org_id, entity_id, country="FR", qty=Decimal("1000.000"))

    _wb, cover, sheet = _sheets(await excise.evidence_workbook(db_session, org_id, period=PERIOD))
    for ws in (cover, sheet):
        text = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
        assert excise.ELIGIBILITY_STATEMENT in text, f"{ws.title} lost the eligibility statement"
        assert excise.RATE_CAVEAT in text, f"{ws.title} lost the rate caveat"
        assert excise.LEGAL_FRAMING in text, f"{ws.title} lost R53's framing"
    cover_text = "\n".join(
        str(c.value) for row in cover.iter_rows() for c in row if c.value is not None
    )
    assert excise.FILED_WITH in cover_text


async def test_wo91_the_cover_states_the_litres_no_rate_is_held_for(db_session):
    """Skipped litres are on the COVER, not buried: an operator who could not
    see them would read the total as "all my foreign diesel". No euro is
    printed for them, because there is none."""
    org_id, entity_id = await _org_entity(db_session, name="WO91 Packet Skipped", seed=65)
    await _diesel(db_session, org_id, entity_id, country="FR", qty=Decimal("1000.000"))
    await _diesel(db_session, org_id, entity_id, country="LV", qty=Decimal("7000.000"))

    _wb, cover, _sheet = _sheets(await excise.evidence_workbook(db_session, org_id, period=PERIOD))
    rows = [[c.value for c in row] for row in cover.iter_rows()]
    flat = [str(v) for row in rows for v in row if v is not None]
    assert any("no rate is held" in v for v in flat)
    assert ["LV", Decimal("7000.000"), 1] in [
        [r[0], Decimal(str(r[1])), r[2]] for r in rows if r and r[0] == "LV"
    ]


# --------------------------------------------------------------------------- #
# Refusals and safety
# --------------------------------------------------------------------------- #


async def test_wo91_an_empty_period_refuses_rather_than_emitting_a_workbook(db_session):
    """A customs packet with no lines in it looks like a filing and supports
    nothing — the `overcharge_pack.no_overcharge_detected` posture, applied to
    the other regime. Fails CLOSED."""
    org_id, entity_id = await _org_entity(db_session, name="WO91 Packet Empty", seed=66)
    with pytest.raises(ValidationError) as exc:
        await excise.evidence_workbook(db_session, org_id, period=PERIOD)
    assert exc.value.code == "no_excise_findings"

    # Litres that exist but sit outside the regime refuse for the same reason:
    # there is nothing to file, and saying so beats handing over a blank sheet.
    await _diesel(db_session, org_id, entity_id, country="LV", qty=Decimal("5000.000"))
    with pytest.raises(ValidationError) as exc:
        await excise.evidence_workbook(db_session, org_id, period=PERIOD)
    assert exc.value.code == "no_excise_findings"


async def test_wo91_free_text_cells_are_formula_injection_safe(db_session):
    """CWE-1236 through the ONE shared `core.csv_safety.sanitize_cell`. The
    entity's own legal name is operator-supplied free text and reaches the
    sheet, so it is the vector under test."""
    org = await make_org(db_session, name="WO91 Packet Injection")
    await enable_transport(db_session, org.id)
    entity = await make_entity(db_session, org.id, country="LV", seed=67)
    entity.legal_name = "=cmd|' /c calc'!A0"
    await db_session.flush()
    await db_session.commit()
    await _diesel(db_session, org.id, entity.id, country="FR", qty=Decimal("1000.000"))

    _wb, _cover, sheet = _sheets(await excise.evidence_workbook(db_session, org.id, period=PERIOD))
    name = sheet.cell(row=2, column=1).value
    assert name.startswith("'="), f"the entity name reached the sheet unneutralised: {name!r}"


async def test_wo91_numeric_cells_are_never_quoted(db_session):
    """The other half of `csv_safety`'s rule: a money/litre/rate cell is written
    as a real number so a customs officer can re-add the column. Quoting them
    would silently turn the figures into text."""
    org_id, entity_id = await _org_entity(db_session, name="WO91 Packet Numbers", seed=68)
    await _diesel(db_session, org_id, entity_id, country="FR", qty=Decimal("1234.567"))

    _wb, _cover, sheet = _sheets(await excise.evidence_workbook(db_session, org_id, period=PERIOD))
    headers = [c.value for c in sheet[1]]
    row = sheet[2]
    for label in ("Litres", "Rate EUR / 1,000 L", "Indicative excise EUR"):
        value = row[headers.index(label)].value
        assert not isinstance(value, str), f"{label} was written as text: {value!r}"
    assert Decimal(str(row[headers.index("Indicative excise EUR")].value)) == Decimal("37.04")
