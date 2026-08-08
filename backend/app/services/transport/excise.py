"""The diesel excise-duty refund — the second recoverable-cash stream, and the
figure that asserts nothing (G4.6; R42, R53's THIRD framing, R51, R49;
`BA_fleet_fuel.md` §1.2, §1.3, §2.4, §3.L, §6.1, §9.2; ARCH_plan G4.6).

THE HARVESTED DEFINITION, VERBATIM
------------------------------------
`BA_fleet_fuel.md` §2.4's `/excise` row is the whole surface in one sentence:

    "7 countries (BE·FR·IT·SI·HU·ES·HR); `litres × rate/1,000L`; rate default
     **EUR 30/1,000 L is an explicit PLACEHOLDER** in the reported EUR 25-33
     band, admin-overridable. **Asserts NO eligibility.** Separate regime,
     filed with **CUSTOMS**."

§1.3's money-flow diagram gives it its own branch — *"[EXCISE REBATE]
per-country litres × rate/1,000 L -> claim to CUSTOMS (separate regime)"* — and
R42 states the requirement and its acceptance:

    "**Diesel excise refund** as a parallel claim engine over the same
     validated diesel lines, per (entity × country), `litres × rate/1,000 L`,
     with an Excel packet for customs. **Rates are admin-overridable and the
     figure asserts NO eligibility** — surfaced loudly."
    Acceptance: "The UI shows the indicative-rate and eligibility caveats on
     every surface that shows the number."

Appendix B pins the two constants (*"Countries: BE, FR, IT, SI, HU, ES, HR"* /
*"Default rate EUR 30.00 / 1,000 L (PLACEHOLDER in the reported EUR 25-33
band)"*) and §6.1 places the regime in law: *"Diesel excise ('professional
diesel') — a **different regime**, claimed from **CUSTOMS**, trucks **>= 7.5
t**, not harmonised, rates change quarterly | ETD 2003/96; national law |
`excise.py` — **explicitly advisory; asserts no eligibility**"*.

THE GOVERNING CONSTRAINT: WHAT THIS FIGURE DOES NOT ASSERT
------------------------------------------------------------
§3.L's advisory-seam table, verbatim: *"`excise.py` | **Asserts NO eligibility**
(vehicle >=7.5t / carrier registration not modelled); rates are indicative
defaults."* §9.2 item 14 files the whole question as an OPEN owner decision:
*"**Who confirms eligibility for excise** (vehicle >= 7.5 t, carrier
registration)? It is deliberately not modelled."*

Those are the two conditions the spec names, and they are the only two this
module names. The number here is therefore **"excise you may be able to reclaim
if you qualify"**, never "excise you are owed" — and a haulier that does not
qualify is owed exactly nothing of it. Presenting it as an entitlement would put
a false figure in front of a customs authority under the client's own name.

So the limitation is STRUCTURAL, not a sentence someone downstream can drop.
Four mechanisms, each asserted by
`tests/transport/test_wo91_excise_eligibility.py`:

1. **One server-side constant.** `ELIGIBILITY_STATEMENT` is defined once, here.
   Every result shape, every response schema and both sheets of the customs
   workbook render THAT string. There is nothing else to render, so it cannot
   be softened, shortened or localised away downstream.
2. **`eligibility_asserted` is a REQUIRED field and is a literal `False`.** A
   consumer cannot receive the number without receiving the denial beside it,
   and a boolean survives a UI that truncates prose.
3. **Vocabulary.** No field this module or `app/schemas/transport_excise.py`
   emits contains `recover`, `owed`, `owes`, `claim`, `demand`, `due`, `debt` or
   `payable` — WO-87's `CLAIM_WORDS`, imported by the test rather than re-typed
   so the two surfaces cannot drift. The euro is **`indicative_excise_eur`**:
   the qualification is inside the field name, where a renderer cannot lose it.
4. **A seeded-violation self-test** for each scanner (`WORK_ORDER_TEMPLATE.md`
   rule 6) — a scanner that cannot fail proves nothing.

R53's THIRD FRAMING — "INDICATIVE, VERIFY"
--------------------------------------------
R53, verbatim: *"Legal framing per analysis must not be flattened: contract
breach = 'money the supplier owes' (claim letter); same-day overpay =
'negotiation evidence, NOT a contractual claim-back' (printed on every sheet);
peer/excise/estimate = **'indicative, verify'**."* §2.4's framing table gives the
words this module uses: *"peer benchmark / excise / refund estimate ->
**"Indicative / advisory — verify before relying"**"*.

`contract_audit` owns the first framing and may say "owes". `savings` owns the
second and denies being a claim. This module owns the THIRD and is weaker than
both: it does not even assert that the money exists for this haulier. Like
`savings`, it shares no code path with the claim-back family — it imports
neither `contract_audit` nor `overcharge` nor `overcharge_pack`, and none of
them imports it, so no excise euro can reach `overcharge.detected_eur` (the
figure WO-83's 30-day demand letter prints) even by a future accident.

THE RATES ARE INDICATIVE DEFAULTS, AND THE DEFAULT IS A PLACEHOLDER
---------------------------------------------------------------------
`DEFAULT_RATE_EUR_PER_1000L` is EUR 30.00 because §2.4 and Appendix B say so,
and both call it a **PLACEHOLDER** in a reported EUR 25-33 band. §9.2 item 13
records why it is still one: *"The excise rates are a single EUR 30/1,000 L
placeholder for all seven countries. Who owns the real per-country statutory
rates, and how often do they change (quarterly, per the research)?"* — an open
owner question this module does not answer in code (master-context §10). It
ships the placeholder, labels it as one on every surface (`RATE_CAVEAT`), and
lets an operator replace it per country (`set_rate`).

`rate_for` resolution: an override row wins; otherwise the default, for a
country in `REFUND_COUNTRIES`; otherwise **`None`**. `None` means *no rate is
configured for this state*, and a cell with no rate produces **no finding at
all** — never a EUR 0.00 row. A zero would read as *"this state refunds you
nothing"*, which is a different and false claim from *"we hold no rate for this
state"*.

§4.14 / §4.15 — WHY THIS MODULE READS NO MONEY COLUMN AT ALL
---------------------------------------------------------------
The mechanism is `litres × rate/1,000 L`. Litres are litres and the rate is
already EUR, so **nothing is converted and nothing is summed across currencies —
because no currency amount is ever read.** `net_local`, `vat_local`,
`gross_local`, `net_eur`, `vat_eur` and `net_eur_eff` do not appear in this
module, and a structural test asserts their absence by name (the
`overcharge_pack` discipline).

That makes this module's §4.15 posture deliberately DIFFERENT from
`savings.py`, which refuses `fx_source == "unknown"` rows outright: a price
comparison is arithmetic ON euros, so an unconvertible euro poisons it, while a
litre count is not. A PLN-invoiced diesel line with no available rate
contributes its litres here exactly like a EUR one, and refusing it would drop
real litres from a customs packet for a reason that has nothing to do with
litres. Stated rather than inherited, and proven by test.

§4.9 — DECIMAL, AND WHERE THE ONE ROUNDING HAPPENS
----------------------------------------------------
Litres are summed EXACTLY and never money-quantized (§4.2 row 9: *"Deliberately
NOT money-quantized"*), then, once:

    indicative_excise_eur = q2(litres / 1000 * rate)

One `q2`, ROUND_HALF_UP, from the exact litre total and the exact rate — never
from a rounded intermediate, so the reported litres, the reported rate and the
reported euro reconcile on the sheet a haulier hands to customs.

§4.19 / §3.L — ADVISORY, STRUCTURALLY
---------------------------------------
§3.L requires every advisory seam to be *"structurally incapable of gating or
mutating a legal figure"*. `excise_report` and the workbook builder issue no
write of any kind — no `db.add`, no attribute assignment on a `FuelTransaction`,
no flush, no audit event — and a test compares every transaction row before and
after a run. They gate no claim, touch no VAT figure and hold no lock. The only
writers here are the RATE CRUD, which is an ordinary audited configuration
mutation (§4.16, old->new).

WHAT THIS MODULE DELIBERATELY DOES NOT CONTAIN
------------------------------------------------
* **An excise claim LIFECYCLE.** R42 calls it a *"parallel claim engine"* of the
  analysis kind; no excise state machine is harvested anywhere in the spec, and
  `overcharge.py`'s `detected -> ... -> recovered` ladder belongs to G4.5's
  contractual claim-back. Not invented (§10).
* **Eligibility.** §9.2 item 14, an open owner question. This module states what
  is not modelled; it does not model it.
* **Real statutory rates.** §9.2 item 13, an open owner question. This module
  ships the spec's placeholder and the override seam.
"""

from __future__ import annotations

import io
import re
from collections import defaultdict
from collections.abc import Callable
from dataclasses import dataclass
from decimal import Decimal

from sqlalchemy.ext.asyncio import AsyncSession

from app.core.csv_safety import sanitize_cell
from app.core.errors import PermissionError, ValidationError
from app.core.money import q2
from app.models.transport.excise_rate import VatExciseRate
from app.services import audit as audit_svc
from app.services import issuer, modules
from app.services.transport import queries

# `BA_fleet_fuel.md` §2.4 *"7 countries (BE·FR·IT·SI·HU·ES·HR)"* and Appendix B
# *"Countries: BE, FR, IT, SI, HU, ES, HR"* — the states the spec records as
# operating a commercial-diesel excise refund. Harvested, not chosen. Sorted so
# every surface lists them in one order.
REFUND_COUNTRIES: tuple[str, ...] = ("BE", "ES", "FR", "HR", "HU", "IT", "SI")

# §2.4 / Appendix B, verbatim: *"rate default EUR 30/1,000 L is an explicit
# PLACEHOLDER in the reported EUR 25-33 band, admin-overridable"*. The
# PLACEHOLDER word is the spec's own and `RATE_CAVEAT` carries it to the wire.
DEFAULT_RATE_EUR_PER_1000L = Decimal("30.0000")

# The reported band the placeholder sits in (§2.4 *"the reported EUR 25-33
# band"*), carried so a surface can show what the default is a placeholder FOR.
REPORTED_RATE_BAND_EUR = (Decimal("25.00"), Decimal("33.00"))

# The denominator of the harvested mechanism: `litres × rate/1,000 L`.
RATE_LITRES = Decimal("1000")

# §3.L / R42 / §9.2 item 14 — the ONE server-side statement of what this figure
# does not assert. Every result shape, every response schema and both sheets of
# the customs workbook render THIS string; there is no second wording anywhere,
# which is what makes the limitation impossible to flatten downstream.
ELIGIBILITY_STATEMENT = (
    "This figure asserts NO eligibility. The conditions that qualify a haulier for a "
    "commercial-diesel excise refund — vehicle weight (>= 7.5 t) and carrier registration — "
    "are deliberately not modelled by this product. Read it as excise you may be able to "
    "reclaim IF you qualify, never as excise you are owed. Confirm eligibility with the "
    "customs authority of the country of supply before filing."
)

# §2.4 / §6.1 / §9.2 item 13 — the rate half of the same honesty.
RATE_CAVEAT = (
    "Rates are indicative defaults, not statutory figures. The default of "
    "EUR 30.00 per 1,000 litres is an explicit placeholder in a reported EUR 25-33 band; "
    "the duty is not harmonised across the EU and national rates are revised (quarterly, "
    "per the underlying research). Verify the rate with customs and override it here."
)

# R53's THIRD framing, in §2.4's own words for this analysis family. NOT a
# softened version of either other framing — see the module docstring.
LEGAL_FRAMING = "Indicative / advisory — verify before relying"

# §2.4 / §1.2 / §6.1 — a SEPARATE regime with a different addressee from the
# VAT refund. Stated so no surface files this with a tax authority.
FILED_WITH = "Customs authority of the country of supply (a separate regime from the VAT refund)"

# `BA_fleet_fuel.md` §3.G G1 / R49 — *"this basis must be stated on any new
# report surface"*. Litres are the subject here rather than a price, so the
# basis line states the litre basis and the currency of the rate.
LITRE_BASIS = "Validated diesel litres as invoiced; the rate is NET EUR per 1,000 litres"

# The currency every figure in this module is expressed in, stated rather than
# assumed (§4.14). No document amount is read, so nothing is converted.
CURRENCY = "EUR"

_PERIOD_RE = re.compile(r"^\d{4}-(0[1-9]|1[0-2])$")
_COUNTRY_RE = re.compile(r"^[A-Z]{2}$")


# --------------------------------------------------------------------------- #
# Gates — module entitlement and input shape
# --------------------------------------------------------------------------- #


async def _require_module(db: AsyncSession, org_id: str) -> None:
    """`modules.is_enabled`, never `require_enabled` — a service signals with
    `AppError`, not `fastapi.HTTPException` (ADR-P3 rule 3). Fails CLOSED."""
    if not await modules.is_enabled(db, org_id, "transport"):
        m = modules.MODULES_BY_KEY["transport"]
        raise PermissionError(f"The {m.name} module is not activated.", code="module_not_enabled")


def validate_period(period: str) -> None:
    """`"YYYY-MM"` — `FuelTransaction.period`'s own shape, and the same refusal
    code every month-grained transport service raises.

    Fails CLOSED for the reason `savings`/`contract_audit` give: silently
    analysing an empty set for a typo'd month looks exactly like *"you burned no
    diesel in a refunding state"*, which is a materially different and far more
    dangerous answer than *"that is not a period"* — here it would be the
    difference between filing nothing and knowing you had nothing to file.
    """
    if not _PERIOD_RE.match(period):
        raise ValidationError(
            f"'{period}' is not a valid period — use YYYY-MM", code="invalid_period"
        )


def _validate_country(country: str) -> str:
    up = (country or "").upper()
    if not _COUNTRY_RE.match(up):
        raise ValidationError(
            f"'{country}' is not an ISO 3166-1 alpha-2 country", code="invalid_country"
        )
    return up


def _validate_refund_country(country: str) -> str:
    """A rate may be typed ONLY for one of the harvested seven states.

    DOCUMENTED INTERPRETATION, fail-CLOSED. `BA_fleet_fuel.md` §2.4 and Appendix
    B state the seven countries as *which states operate the regime*, not as a
    default list to extend. Accepting an eighth would make this platform assert
    that a commercial-diesel excise refund exists in a state where the spec
    records none — an invented legal claim, which is the one thing a figure a
    haulier files with customs must never contain (master-context §10).

    What that deliberately costs: a state whose regime lapses cannot be switched
    OFF here, only re-rated. Adding an `active` column would be inventing a
    lifecycle the spec does not harvest, so the gap is recorded in
    `docs/plan/plan-a/wo/WO-91-excise-refund.md` as a named follow-up instead.
    """
    up = _validate_country(country)
    if up not in REFUND_COUNTRIES:
        raise ValidationError(
            f"'{up}' is not one of the states this product records as operating a "
            f"commercial-diesel excise refund ({', '.join(REFUND_COUNTRIES)}). A rate is not "
            "accepted for it, because storing one would assert a refund regime exists there.",
            code="excise_country_not_supported",
        )
    return up


def _validate_rate(rate: Decimal) -> Decimal:
    """A rate must be a real, positive rate.

    A zero is refused rather than stored as "this state refunds nothing": the
    absence of the regime is already expressed by a country not being in
    `REFUND_COUNTRIES`, and a stored zero would make `excise_report` publish a
    EUR 0.00 finding — a materially different and false claim from *"we hold no
    rate for this state"*. The database CHECK refuses it too (defense in depth).
    """
    if not isinstance(rate, Decimal):  # pragma: no cover - the schema types it
        rate = Decimal(str(rate))
    if rate <= 0:
        raise ValidationError(
            f"'{rate}' is not a usable excise rate — a rate is EUR per 1,000 litres and must "
            "be positive. A state that refunds nothing is expressed by holding no rate for it, "
            "never by a zero.",
            code="invalid_excise_rate",
        )
    return rate


# --------------------------------------------------------------------------- #
# The rate registry — the ONLY writers in this module
# --------------------------------------------------------------------------- #


async def get_rate_row(db: AsyncSession, org_id: str, country: str) -> VatExciseRate | None:
    """The org's typed override for one country, or `None`. Org-scoped; a rate
    belonging to another tenant is simply not there (§4.1/§4.4)."""
    return await db.scalar(queries.excise_rates(org_id, country=country))


async def list_rate_rows(db: AsyncSession, org_id: str) -> list[VatExciseRate]:
    """The org's TYPED override rows only — no defaults folded in.

    Separate from `list_rates` because the two answer different questions and a
    surface needs both: *"what rate will my figure use?"* (the resolved view)
    and *"which of these did a human actually verify with customs?"* (this one).
    Collapsing them would make an unverified placeholder indistinguishable from
    a verified statutory rate, which is exactly the confusion §9.2 item 13 warns
    about.
    """
    await _require_module(db, org_id)
    return list(await db.scalars(queries.excise_rates(org_id)))


async def list_rates(db: AsyncSession, org_id: str) -> dict[str, Decimal]:
    """`{country: rate}` for every one of the harvested seven — the override
    where one has been typed, the indicative default everywhere else.

    Returns the RESOLVED view rather than only the stored rows, because the
    question an operator is asking is *"what rate will my figure use?"*, and
    answering it with an empty list for an org that has typed nothing would hide
    the placeholder the figure is actually computed from.
    """
    await _require_module(db, org_id)
    rows = list(await db.scalars(queries.excise_rates(org_id)))
    overrides = {r.country: Decimal(r.rate_eur_per_1000l) for r in rows}
    return {c: overrides.get(c, DEFAULT_RATE_EUR_PER_1000L) for c in REFUND_COUNTRIES}


def default_rate_for(country: str) -> Decimal | None:
    """The indicative default for a country, or `None` outside the seven. Pure —
    no I/O — so a caller that already holds the override map can resolve without
    a second query, and so the `None` branch is testable on its own."""
    return DEFAULT_RATE_EUR_PER_1000L if country in REFUND_COUNTRIES else None


async def rate_for(db: AsyncSession, org_id: str, country: str) -> Decimal | None:
    """The rate `excise_report` will apply to `country`: the typed override, else
    the harvested indicative default, else **`None`**.

    `None` is not zero and must never be rendered as one — see the module
    docstring. A country outside the seven resolves to `None` and produces no
    finding at all.

    A HELPER, not an entry point: it applies no module gate, because every
    caller (`excise_report`, and the rate CRUD) has already passed one. Nothing
    route-facing calls it directly, and nothing should — an ungated entry point
    is exactly what ADR-P3 rule 3 exists to prevent.
    """
    row = await get_rate_row(db, org_id, country)
    if row is not None:
        return Decimal(row.rate_eur_per_1000l)
    return default_rate_for(country)


async def set_rate(
    db: AsyncSession, org_id: str, *, country: str, rate_eur_per_1000l: Decimal
) -> VatExciseRate:
    """Type the rate a customs authority actually applies, replacing the
    harvested placeholder for that state — the ONLY writer of `VatExciseRate`
    alongside `remove_rate`.

    Gate order, fails CLOSED: module entitlement -> the country is one of the
    harvested seven (`excise_country_not_supported`) -> the rate is positive
    (`invalid_excise_rate`).

    Idempotent: a repeat call with the identical rate returns the existing row
    and audits nothing (the `contract_audit.set_term` / `checklist.set_active`
    no-op convention). Any change audits old->new IN THE SAME transaction as the
    mutation (§4.16) — this rate multiplies real litres into a figure a haulier
    takes to a customs authority, so a change to it always has an actor.
    """
    await _require_module(db, org_id)
    country = _validate_refund_country(country)
    rate = _validate_rate(rate_eur_per_1000l)

    existing = await get_rate_row(db, org_id, country)
    if existing is not None:
        old = Decimal(existing.rate_eur_per_1000l)
        if old == rate:
            return existing
        existing.rate_eur_per_1000l = rate
        await db.flush()
        await audit_svc.record(
            db,
            audit_svc.A.TRANSPORT_EXCISE_RATE_SET,
            target_type="vat_excise_rate",
            target_id=existing.id,
            meta={"country": country, "old": str(old), "new": str(rate)},
            org_id=org_id,
        )
        return existing

    row = VatExciseRate(org_id=org_id, country=country, rate_eur_per_1000l=rate)
    db.add(row)
    await db.flush()
    await audit_svc.record(
        db,
        audit_svc.A.TRANSPORT_EXCISE_RATE_SET,
        target_type="vat_excise_rate",
        target_id=row.id,
        meta={
            "country": country,
            "old": None,
            "new": str(rate),
            "replaced_default": str(DEFAULT_RATE_EUR_PER_1000L),
        },
        org_id=org_id,
    )
    return row


async def remove_rate(db: AsyncSession, org_id: str, *, country: str) -> bool:
    """Drop the typed override so the state falls back to the harvested
    indicative default. Returns `True` when a row was removed.

    Removing is NOT "this state refunds nothing" — it is "we no longer hold a
    verified rate for it", and the figure goes back to being explicitly
    placeholder-based. Audited old->new like any configuration change (§4.16).
    A country with no override is a no-op returning `False`, never a 404: there
    is nothing to hide and nothing to guess at.
    """
    await _require_module(db, org_id)
    country = _validate_refund_country(country)

    existing = await get_rate_row(db, org_id, country)
    if existing is None:
        return False
    old = Decimal(existing.rate_eur_per_1000l)
    target_id = existing.id
    await db.delete(existing)
    await db.flush()
    await audit_svc.record(
        db,
        audit_svc.A.TRANSPORT_EXCISE_RATE_REMOVE,
        target_type="vat_excise_rate",
        target_id=target_id,
        meta={
            "country": country,
            "old": str(old),
            "new": None,
            "restored_default": str(DEFAULT_RATE_EUR_PER_1000L),
        },
        org_id=org_id,
    )
    return True


# --------------------------------------------------------------------------- #
# The analysis — READ-ONLY from here down (§4.19 / §3.L)
# --------------------------------------------------------------------------- #


@dataclass(frozen=True)
class ExciseCell:
    """One (entity × country) cell of the period — R42's own grain.

    `litres` is the EXACT sum of the cell's validated diesel quantities, never
    money-quantized (§4.2 row 9). `rate_eur_per_1000l` is the rate that was
    actually applied and `rate_is_override` says whether a human verified it or
    it is the harvested placeholder — a customs packet that could not tell those
    apart would be presenting a placeholder as a statutory figure.

    `indicative_excise_eur` is named for what it is. See the module docstring:
    the qualification lives in the field name because a renderer can lose a
    caveat string and cannot lose an identifier.
    """

    entity_id: str
    entity_name: str
    country: str
    litres: Decimal
    rate_eur_per_1000l: Decimal
    rate_is_override: bool
    indicative_excise_eur: Decimal
    lines: int


@dataclass(frozen=True)
class SkippedCountry:
    """A country whose validated diesel litres are in scope but for which this
    product holds NO rate — reported rather than silently dropped.

    Silence would be the dangerous answer: an operator who saw only the seven
    refunding states could not tell "we hold no rate for PL" from "you burned no
    diesel in PL". The litres are stated so the size of what is NOT being
    reported is visible, and `indicative_excise_eur` is deliberately absent from
    this shape — there is no figure, which is not the same as a figure of zero.
    """

    country: str
    litres: Decimal
    lines: int


@dataclass(frozen=True)
class ExciseReport:
    """R42's analysis over one accounting month.

    Every caveat is a REQUIRED field, not optional metadata: R42's acceptance is
    *"The UI shows the indicative-rate and eligibility caveats on **every
    surface that shows the number**"*, and the only way to guarantee that is to
    make the number un-renderable without them. `eligibility_asserted` is always
    `False` and exists so the denial survives a surface that truncates prose.
    """

    period: str
    entity_id: str | None
    country: str | None
    currency: str
    product_group: str
    litre_basis: str
    legal_framing: str
    eligibility: str
    eligibility_asserted: bool
    rate_caveat: str
    filed_with: str
    rows: tuple[ExciseCell, ...]
    skipped_countries: tuple[SkippedCountry, ...]
    litres: Decimal
    indicative_excise_eur: Decimal
    lines_examined: int


async def _entity_names(db: AsyncSession, org_id: str, entity_ids: set[str]) -> dict[str, str]:
    """Display names for the claimant entities in scope, read through the issuer
    SERVICE (`issuer.get_by_id`) rather than a raw cross-domain join — ADR-P3
    rule 2, the path `claim.py`/`fuel.py`/`claim_pack.py` all take. An entity
    that cannot be resolved falls back to its id rather than to an invented
    name (§10)."""
    out: dict[str, str] = {}
    for entity_id in sorted(entity_ids):
        profile = await issuer.get_by_id(db, org_id, entity_id)
        name = None if profile is None else (profile.legal_name or profile.name)
        out[entity_id] = name or entity_id
    return out


async def excise_report(
    db: AsyncSession,
    org_id: str,
    *,
    period: str,
    entity_id: str | None = None,
    country: str | None = None,
) -> ExciseReport:
    """`BA_fleet_fuel.md` R42's diesel excise-duty refund, complete.

    *"over the same validated diesel lines, per (entity × country), `litres ×
    rate/1,000 L`"* — one accounting month, optionally scoped to one claimant
    entity or one country of supply.

    **This figure asserts NO eligibility** (§3.L): vehicle weight (>= 7.5 t) and
    carrier registration are deliberately not modelled, so it is excise that may
    be reclaimable IF the haulier qualifies, never excise it is owed. The
    statement rides the result as `eligibility`, and `eligibility_asserted` is
    `False`. Framing: **indicative / advisory — verify before relying** (R53's
    third framing). Filed with **CUSTOMS**, a separate regime from the VAT
    refund (§2.4, §1.2, §6.1).

    Gate order, fails CLOSED: module entitlement -> `period` shape -> `country`
    shape. There is no FX gate and there is nothing for one to do: this function
    reads `qty` and no currency amount at all, so nothing is converted and
    nothing is summed across currencies (§4.14/§4.15 — see the module
    docstring for why that is a deliberate divergence from `savings.py`).

    A country this product holds no rate for produces **no row** and is reported
    in `skipped_countries` with its litres. A cell whose litres are not positive
    produces no row either: a promo correction can legitimately net a cell to
    zero or below, and a EUR 0.00 line on a customs packet is noise at best and
    a misstatement at worst.

    Read-only (§4.19 / §3.L): no write of any kind, and a test compares every
    `fuel_transactions` row column by column before and after a run.
    """
    await _require_module(db, org_id)
    validate_period(period)
    scope_country = _validate_country(country) if country is not None else None

    rows = list(
        await db.scalars(
            queries.excise_transactions(
                org_id, period=period, entity_id=entity_id, country=scope_country
            )
        )
    )

    # The rate resolution is done ONCE per country, from the one registry query,
    # rather than per row: a rate that changed mid-loop would produce a report
    # whose own cells disagree.
    overrides = {r.country: Decimal(r.rate_eur_per_1000l) for r in await list_rate_rows(db, org_id)}

    litres: dict[tuple[str, str], Decimal] = defaultdict(lambda: Decimal("0"))
    lines: dict[tuple[str, str], int] = defaultdict(int)
    for row in rows:
        key = (row.entity_id, row.country)
        litres[key] += Decimal(row.qty)
        lines[key] += 1

    rated: list[tuple[str, str]] = []
    unrated_litres: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    unrated_lines: dict[str, int] = defaultdict(int)
    for key in sorted(litres):
        _entity, cell_country = key
        if overrides.get(cell_country, default_rate_for(cell_country)) is None:
            unrated_litres[cell_country] += litres[key]
            unrated_lines[cell_country] += lines[key]
            continue
        rated.append(key)

    names = await _entity_names(db, org_id, {entity for entity, _ in rated})

    cells: list[ExciseCell] = []
    for entity, cell_country in rated:
        cell_litres = litres[(entity, cell_country)]
        if cell_litres <= 0:
            continue
        override = overrides.get(cell_country)
        rate = override if override is not None else DEFAULT_RATE_EUR_PER_1000L
        cells.append(
            ExciseCell(
                entity_id=entity,
                entity_name=names[entity],
                country=cell_country,
                litres=cell_litres,
                rate_eur_per_1000l=rate,
                rate_is_override=override is not None,
                # §4.9 — ONE q2, ROUND_HALF_UP, from the EXACT litre total and
                # the exact rate. Never from a rounded intermediate.
                indicative_excise_eur=q2(cell_litres / RATE_LITRES * rate),
                lines=lines[(entity, cell_country)],
            )
        )

    return ExciseReport(
        period=period,
        entity_id=entity_id,
        country=scope_country,
        currency=CURRENCY,
        product_group=queries.DIESEL,
        litre_basis=LITRE_BASIS,
        legal_framing=LEGAL_FRAMING,
        eligibility=ELIGIBILITY_STATEMENT,
        eligibility_asserted=False,
        rate_caveat=RATE_CAVEAT,
        filed_with=FILED_WITH,
        rows=tuple(cells),
        skipped_countries=tuple(
            SkippedCountry(country=c, litres=unrated_litres[c], lines=unrated_lines[c])
            for c in sorted(unrated_litres)
        ),
        litres=sum((c.litres for c in cells), Decimal("0")),
        indicative_excise_eur=q2(sum((c.indicative_excise_eur for c in cells), Decimal("0"))),
        lines_examined=len(rows),
    )


# --------------------------------------------------------------------------- #
# The customs evidence packet — ONE source, and a second renderer over it
# --------------------------------------------------------------------------- #

# The sheet a haulier hands to customs, and the one that explains it. Two sheets
# rather than one because R42's acceptance requires the caveats on *"every
# surface that shows the number"* — so BOTH carry them, and the numbers sheet
# does not depend on someone having read the cover.
PACKET_SHEET_COVER = "Excise refund"
PACKET_SHEET_LITRES = "Litres"
TOTAL_LABEL = "TOTAL"


class _Fmt:
    """How a column's value is rendered. The only place the report and the
    workbook are allowed to differ, and only in FORMATTING, never in value."""

    TEXT = "text"  # free text -> sanitized (CWE-1236)
    MONEY = "money"  # EUR, 2 dp
    RATE = "rate"  # EUR / 1,000 L, 4 dp
    QTY = "qty"  # litres, 3 dp (`FuelTransaction.qty`'s own scale)
    COUNT = "count"  # a line count — a number, never sanitized into text


# The two words the "Rate source" column may print — the `is_override`
# distinction §9.2 item 13 exists for, in the words a customs officer reads.
RATE_SOURCE_OVERRIDE = "verified override"
RATE_SOURCE_DEFAULT = "indicative default"


@dataclass(frozen=True)
class _Column:
    """One column of the litres table: its heading, how to read it off a
    `(cell, report)` pair, and how to format it. The accessor takes BOTH
    because a couple of columns are properties of the report rather than of the
    cell — and a renderer that reached for `report` directly could drift from
    the spec, while an accessor cannot."""

    label: str
    value: Callable[[ExciseCell, ExciseReport], object]
    fmt: str
    totalled: bool = False


# THE one column spec. Every accessor reads an `ExciseCell` (or the report's own
# constant) and DERIVES nothing, so a cell in the workbook can only differ from
# the JSON report if the two are looking at different ROWS — which one loader
# makes impossible.
_COLUMNS: tuple[_Column, ...] = (
    _Column("Claimant entity", lambda c, r: c.entity_name, _Fmt.TEXT),
    _Column("Entity reference", lambda c, r: c.entity_id, _Fmt.TEXT),
    _Column("Country of supply", lambda c, r: c.country, _Fmt.TEXT),
    _Column("Product group", lambda c, r: r.product_group, _Fmt.TEXT),
    _Column("Litres", lambda c, r: c.litres, _Fmt.QTY, totalled=True),
    _Column("Lines", lambda c, r: c.lines, _Fmt.COUNT),
    _Column("Rate EUR / 1,000 L", lambda c, r: c.rate_eur_per_1000l, _Fmt.RATE),
    _Column(
        "Rate source",
        lambda c, r: RATE_SOURCE_OVERRIDE if c.rate_is_override else RATE_SOURCE_DEFAULT,
        _Fmt.TEXT,
    ),
    _Column(
        "Indicative excise EUR", lambda c, r: c.indicative_excise_eur, _Fmt.MONEY, totalled=True
    ),
)


def packet_row(cell: ExciseCell, report: ExciseReport) -> list[object]:
    """One typed row of the litres table. Money, rates and litres stay
    `Decimal` (§4.9) so the spreadsheet holds real numbers; formatting for a
    medium happens in the renderer, never here.

    Public so a test can assert the workbook cell-for-cell against the SAME
    projection the renderer uses, rather than against a re-typed expectation."""
    return [column.value(cell, report) for column in _COLUMNS]


def _excel_cell(column: _Column, value: object) -> object:
    """A cell rendered into a spreadsheet. Free text goes through the ONE shared
    `core.csv_safety.sanitize_cell` (CWE-1236); money, rate and litre cells are
    written as raw `Decimal`s — a leading `-` there is a negative quantity, not
    an injection (`csv_safety`'s own rule), and quoting them would turn the
    numbers a customs officer re-adds into text."""
    if column.fmt == _Fmt.TEXT:
        return sanitize_cell(value)
    return value


def build_evidence_workbook(report: ExciseReport) -> bytes:
    """R42's *"Excel packet for customs"*, rendered from an `ExciseReport`.

    ONE SOURCE, TWO RENDERERS — the WO-74/WO-83 precedent, made structural the
    same way `overcharge_pack` makes it structural:

    * this is a **sync** function taking a LOADED report. It holds no
      `AsyncSession`, so it cannot query a second source even by accident;
    * it projects `_COLUMNS` off the report's own `ExciseCell` objects and
      derives nothing, so a cell can only differ from the JSON response if the
      two are looking at different rows — which one loader makes impossible;
    * the TOTAL row prints `report.indicative_excise_eur`, the figure the
      service already computed, never a second summation (§4.10).

    Both sheets carry `ELIGIBILITY_STATEMENT`, `RATE_CAVEAT`, `LEGAL_FRAMING`
    and `FILED_WITH`, because R42's acceptance is *"the caveats on **every
    surface that shows the number**"* and a numbers sheet detached from its
    cover page is exactly such a surface. R53's *"each workbook carries its
    framing text"* is the same requirement from the other side.

    A report with no rows RAISES rather than rendering (`no_excise_findings`) —
    see `evidence_workbook`.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = PACKET_SHEET_COVER
    header_rows: list[tuple[str, object]] = [
        ("Period", sanitize_cell(report.period)),
        ("Product group", sanitize_cell(report.product_group)),
        ("Currency", sanitize_cell(report.currency)),
        ("Litre basis", sanitize_cell(report.litre_basis)),
        ("Total litres", report.litres),
        ("Indicative excise EUR", report.indicative_excise_eur),
        ("Cells", len(report.rows)),
        ("Lines examined", report.lines_examined),
        ("File with", sanitize_cell(report.filed_with)),
        ("Framing", sanitize_cell(report.legal_framing)),
        ("Eligibility", sanitize_cell(report.eligibility)),
        ("Rate basis", sanitize_cell(report.rate_caveat)),
    ]
    for label, value in header_rows:
        ws.append([label, value])
    for row in ws.iter_rows(min_row=1, max_row=len(header_rows), max_col=1):
        row[0].font = Font(bold=True)

    if report.skipped_countries:
        # Stated on the COVER, not buried: an operator must see how many litres
        # are outside the figure, or they will read the total as "all my
        # foreign diesel". No euro is printed for these — there is none.
        ws.append([])
        ws.append(["Litres in scope for which no rate is held (no figure is reported)"])
        ws[ws.max_row][0].font = Font(bold=True)
        ws.append(["Country of supply", "Litres", "Lines"])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        for skipped in report.skipped_countries:
            ws.append([sanitize_cell(skipped.country), skipped.litres, skipped.lines])

    sheet = wb.create_sheet(PACKET_SHEET_LITRES)
    sheet.append([c.label for c in _COLUMNS])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for cell_row in report.rows:
        values = packet_row(cell_row, report)
        sheet.append([_excel_cell(_COLUMNS[i], v) for i, v in enumerate(values)])

    total_row: list[object] = []
    for idx, column in enumerate(_COLUMNS):
        if idx == 0:
            total_row.append(TOTAL_LABEL)
        elif column.label == "Litres":
            total_row.append(report.litres)
        elif column.totalled:
            total_row.append(report.indicative_excise_eur)
        else:
            total_row.append("")
    sheet.append(total_row)
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)

    # R42 / R53 on the sheet that carries the numbers, not only on the cover.
    sheet.append([])
    sheet.append([report.legal_framing])
    sheet.append([report.eligibility])
    sheet.append([report.rate_caveat])
    sheet.append([report.filed_with])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def evidence_workbook(
    db: AsyncSession,
    org_id: str,
    *,
    period: str,
    entity_id: str | None = None,
    country: str | None = None,
) -> bytes:
    """R42's Excel packet for **customs**, for one accounting month.

    Loads through the SAME `excise_report` the JSON surface returns — one
    source, so the packet and the API can never disagree — and renders it.

    REFUSES rather than emitting an empty document (422 `no_excise_findings`).
    A customs packet with no litres in it is not an empty report: it is a
    document that looks like a filing and supports nothing, and the honest
    answer to *"there is nothing to file"* is to say so rather than to hand
    someone a sheet they might send. The `overcharge_pack.no_overcharge_detected`
    precedent, applied to the other regime.

    Read-only: nothing is persisted, mutated, audited or committed (§4.19).
    """
    report = await excise_report(db, org_id, period=period, entity_id=entity_id, country=country)
    if not report.rows:
        raise ValidationError(
            f"No diesel litres in {period} fall in a state this product holds an excise rate "
            "for — refusing to produce a customs packet with no lines in it",
            code="no_excise_findings",
        )
    return build_evidence_workbook(report)


__all__ = [
    "CURRENCY",
    "DEFAULT_RATE_EUR_PER_1000L",
    "ELIGIBILITY_STATEMENT",
    "FILED_WITH",
    "LEGAL_FRAMING",
    "LITRE_BASIS",
    "RATE_CAVEAT",
    "RATE_LITRES",
    "REFUND_COUNTRIES",
    "PACKET_SHEET_COVER",
    "PACKET_SHEET_LITRES",
    "RATE_SOURCE_DEFAULT",
    "RATE_SOURCE_OVERRIDE",
    "REPORTED_RATE_BAND_EUR",
    "TOTAL_LABEL",
    "ExciseCell",
    "ExciseReport",
    "SkippedCountry",
    "build_evidence_workbook",
    "default_rate_for",
    "evidence_workbook",
    "excise_report",
    "get_rate_row",
    "packet_row",
    "list_rate_rows",
    "list_rates",
    "rate_for",
    "remove_rate",
    "set_rate",
    "validate_period",
]
