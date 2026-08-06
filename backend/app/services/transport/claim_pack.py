"""The two filing artifacts of a frozen VAT refund claim — the Excel claim
workbook and the ZIP evidence pack (G2.12; `BA_fleet_fuel.md` §2.2 line 204,
§3.K K6, §3.M M1; WO-74).

WHAT THE TWO ARTIFACTS ARE, AND FOR WHOM
--------------------------------------------
The **workbook** (`build_workbook`) is the figures document the operator files
with the refunding member state: the claim header plus one row per frozen
claim line at the R2 grain (one row per (invoice, product code), never an
`ALL:` aggregate) with its Art. 9 goods code, and a TOTAL row equal to the
claim's frozen VAT base. The **evidence pack** (`build_evidence_pack`) is the
audit-defence *filing bundle* (§3.K K6): a ZIP assembling that same workbook
PLUS every vaulted source document of the claim's invoices, laid out under the
human-navigable vault tree `<Customer> <RegNo> / <Year> / <Country> /
<Claim period> / <file>` (§3.M M1 — "every segment is sanitized for
Windows/SharePoint/FTP safety"), plus a `manifest.csv` naming each file's
SHA-256 (the same manifest discipline the spec's integrity section, §3.K K4,
establishes for backups).

WHY BOTH ARTIFACTS COME FROM ONE LOADER (`_load_pack`)
----------------------------------------------------------
`ARCH_plan.md`'s G2.12 acceptance is "the workbook and the evidence pack agree
line-for-line" — the same both-artifacts-one-source rule the harvested R41
states for the overcharge packet ("built from the SAME line source ... Both
artifacts ... show identical lines and totals"). Both public builders here are
thin renderers over ONE `_load_pack()` result, and the evidence pack's
embedded workbook is produced by the SAME `_render_workbook()` over the SAME
loaded pack — identity is structural, never two queries that happen to agree
today.

WHY FROZEN LINES ONLY (fail-closed `claim_not_frozen`)
----------------------------------------------------------
Fleet Fuel derived claim lines LIVE forever and protected the filed claim only
by a physically separate legal database; ADR-P3 translated that into
materialize-and-freeze at submission (G2.5, `freeze.py`: "after submission,
editing or re-closing the underlying period changes nothing about the filed
claim"). The FILING artifacts must therefore render exactly the frozen,
locked claim set — a draft claim's mutable lines could drift between the
moment the workbook is downloaded and the moment the claim locks, which is
precisely the divergence the freeze exists to make impossible. No frozen
lines -> refuse. A WITHDRAWN claim keeps its frozen lines (WO-51's
`withdraw_claim` deletes lock rows, never lines), so its artifacts stay
reproducible — deliberate: reproducing what WAS filed is audit-trail
behaviour, not a leak.

WHY THE SYNTHETIC REFUSAL LIVES HERE TOO (R3's fourth consumer)
-------------------------------------------------------------------
`BA_fleet_fuel.md` C2: "A pack containing ANY synthetic line CANNOT be filed",
enforced by THE single `claim_gates.is_synthetic()` predicate "used by the
lock gate, the checklist gate, the readiness check AND the workbook builder
... so they all block the same set of synthetic refs". This module is the
workbook builder — the fourth named consumer — and it REUSES WO-49's one
predicate, never a second implementation. Fails CLOSED: a synthetic ref means
"we cannot prove which real invoice this euro amount belongs to", and filing
it risks the whole claim (see `claim_gates`'s own docstring).

WHY THE TOTALS-DRIFT CHECK REFUSES (fail-closed `claim_totals_drift`)
-------------------------------------------------------------------------
Master-context §4.10 — the server recomputes every total. The renderer sums
the frozen lines with `q2` and refuses if the result differs from the claim's
frozen `vat_eur`/`vat_local` header: a filing artifact must never paper over
a corrupted frozen set with a header figure its own lines cannot reproduce.

WHY MISSING DOCUMENT BYTES REFUSE (fail-closed `evidence_document_unavailable`)
----------------------------------------------------------------------------------
R10 (WO-58) guaranteed every resolved invoice had >=1 vaulted capture at
submission. If the object-store bytes are GONE at assembly time
(`core.storage.StorageError` — a plain `Exception`, not an `AppError`, so it
is caught and re-raised in the wire shape), the bundle refuses rather than
shipping a silently-incomplete filing bundle — an audit answer missing an
invoice original is a rejected claim, and a rejected claim's invoices stay
locked (R5): the money would be gone, not delayed.

WHAT THIS MODULE NEVER DOES
-------------------------------
Persists nothing, mutates nothing, audits nothing (read-only artifact
generation — the `erp_export`/`report_writers` precedent; master-context
§4.16 requires audit on *mutations*). No route exists yet — the
`api/routes/transport/*` batch serves and permission-gates these bytes later.
Every figure is read from the frozen `Numeric(14,2)` columns as `Decimal` and
summed via `q2` ROUND_HALF_UP (R14, §4.9) — never recomputed from
`fuel_transactions`, never through float. Free-text cells go through the ONE
shared `core.csv_safety.sanitize_cell` (CWE-1236); numeric money cells are
deliberately NOT sanitized (a leading `-` is a negative amount, not an
injection — `csv_safety`'s own rule).
"""

from __future__ import annotations

import csv
import io
import zipfile
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.csv_safety import sanitize_cell
from app.core.errors import ConflictError, NotFoundError, PermissionError
from app.core.money import q2
from app.core.storage import StorageError, sha256_hex
from app.models.transport.vat_claim import VatRefundClaim, VatRefundClaimLine
from app.services import documents, extraction, modules, vendors
from app.services import invoices as ap_invoices
from app.services import issuer as issuer_svc
from app.services.transport.claim_gates import is_synthetic

WORKBOOK_FILENAME = "claim-workbook.xlsx"
MANIFEST_FILENAME = "manifest.csv"

# Characters Windows/SharePoint/FTP path segments must never contain (§3.M M1:
# "every segment is sanitized for Windows/SharePoint/FTP safety") — the
# Windows-forbidden set plus control characters, handled in _sanitize_segment.
_FORBIDDEN_SEGMENT_CHARS = '<>:"/\\|?*'

_LINES_HEADER = (
    "Supplier",
    "Invoice number",
    "Invoice date",
    "Goods code",
    "Product group",
    "Net EUR",
    "VAT EUR",
    "Net local",
    "VAT local",
    "Currency",
)


@dataclass(frozen=True)
class PackLine:
    """One frozen claim line, resolved for rendering — the R2 grain (one row
    per (invoice, product code)) with the supplier/invoice context the frozen
    line's own columns don't carry."""

    supplier: str
    invoice_number: str
    invoice_date: date | None
    goods_code: str | None
    product_group: str | None
    net_eur: Decimal
    vat_eur: Decimal
    net_local: Decimal | None
    vat_local: Decimal | None
    currency: str | None
    invoice_id: str


@dataclass(frozen=True)
class ClaimPack:
    """Everything both renderers need, loaded ONCE — see module docstring
    "WHY BOTH ARTIFACTS COME FROM ONE LOADER"."""

    claim_id: str
    entity_name: str
    entity_vat_number: str | None
    entity_registration_number: str | None
    refund_country: str
    ref_period: str
    status: str
    status_code: str | None
    submitted_date: date | None
    vat_eur: Decimal
    vat_local: Decimal
    currency: str | None
    net_eur_total: Decimal
    net_local_total: Decimal
    lines: tuple[PackLine, ...]


def period_label(ref_period: str) -> str:
    """§3.M M1's `period_label()` for the two shapes a claim can carry
    (`claim.validate_ref_period` + the DB CHECK constraint already guarantee
    them): `"YYYY-Qn"` -> `"Qn"`, `"YYYY-YEAR"` -> `"Annual"`."""
    tail = ref_period.split("-", 1)[1]
    return "Annual" if tail == "YEAR" else tail


def _sanitize_segment(segment: str) -> str:
    """One vault-tree path segment, made Windows/SharePoint/FTP-safe (§3.M
    M1): forbidden characters and control characters become `_`, trailing
    dots/spaces are stripped (Windows rejects them), and an empty result
    degrades to `"_"` so a pathological entity name never produces an empty
    path segment (which would silently collapse the tree one level)."""
    cleaned = "".join(
        "_" if (ch in _FORBIDDEN_SEGMENT_CHARS or ord(ch) < 32) else ch for ch in segment
    )
    cleaned = cleaned.rstrip(". ").strip()
    return cleaned or "_"


def vault_prefix(pack: ClaimPack) -> str:
    """The human-navigable bundle root (§3.K K6 / §3.M M1):
    `<Customer> <RegNo>/<Year>/<Country>/<Claim period>/`."""
    customer = pack.entity_name
    if pack.entity_registration_number:
        customer = f"{customer} {pack.entity_registration_number}"
    year = pack.ref_period.split("-", 1)[0]
    segments = [customer, year, pack.refund_country, period_label(pack.ref_period)]
    return "/".join(_sanitize_segment(s) for s in segments)


async def _get_claim(db: AsyncSession, org_id: str, claim_id: str) -> VatRefundClaim:
    """Org-scoped lookup — a cross-tenant claim id is indistinguishable from
    an unknown one (master-context §4.4: opaque 404, never 403)."""
    claim = await db.scalar(
        select(VatRefundClaim).where(VatRefundClaim.id == claim_id, VatRefundClaim.org_id == org_id)
    )
    if claim is None:
        raise NotFoundError("Claim not found", code="claim_not_found")
    return claim


async def _load_pack(db: AsyncSession, org_id: str, claim_id: str) -> ClaimPack:
    """The ONE loader both artifacts render from. Applies, in order: the
    module-entitlement gate, the org-scoped fetch, the frozen-lines-only
    rule, the R3 synthetic refusal, the totals-drift check, and the
    entity/invoice/vendor resolution — every refusal fails CLOSED (module
    docstring states why for each)."""
    if not await modules.is_enabled(db, org_id, "transport"):
        m = modules.MODULES_BY_KEY["transport"]
        raise PermissionError(f"The {m.name} module is not activated.", code="module_not_enabled")

    claim = await _get_claim(db, org_id, claim_id)

    lines = list(
        await db.scalars(
            select(VatRefundClaimLine)
            .where(
                VatRefundClaimLine.org_id == org_id,
                VatRefundClaimLine.claim_id == claim.id,
                VatRefundClaimLine.frozen_at.is_not(None),
            )
            .order_by(VatRefundClaimLine.invoice_ref, VatRefundClaimLine.product_group)
        )
    )
    if not lines:
        raise ConflictError(
            f"Claim is '{claim.status}' with no frozen lines — the filing artifacts "
            "render the frozen claim set only (submit the claim first)",
            code="claim_not_frozen",
        )

    # R3 — the workbook builder is the fourth named consumer of THE single
    # `is_synthetic()` predicate ("a pack containing ANY synthetic line
    # CANNOT be filed"). Name every offender so the operator fixes them all
    # in one pass, the same message discipline the other gates use.
    synthetic = sorted({ln.invoice_ref for ln in lines if is_synthetic(ln.invoice_ref, ln.vat_id)})
    if synthetic:
        raise ConflictError(
            f"Cannot build filing artifacts — synthetic line(s) in pack: {', '.join(synthetic)}",
            code="synthetic_line_in_pack",
        )

    # Defensive twin of `freeze._sum_lines`'s refusal (§4.14): a frozen set
    # spanning currencies cannot exist via `submit_claim`, but this module
    # must not ASSUME its input's provenance — refuse rather than mislabel.
    currencies = {ln.currency for ln in lines if ln.currency is not None}
    if len(currencies) > 1:
        raise ConflictError(
            f"Frozen claim lines span multiple currencies {sorted(currencies)} — "
            "refusing to total a filing artifact across them",
            code="claim_currency_mismatch",
        )

    # §4.10 — recompute the totals the artifact will print and refuse if the
    # frozen header no longer equals its own lines (module docstring "WHY THE
    # TOTALS-DRIFT CHECK REFUSES"). All Decimal, q2 ROUND_HALF_UP (R14).
    vat_eur_total = q2(sum((ln.vat_eur for ln in lines), Decimal("0")))
    net_eur_total = q2(sum((ln.net_eur for ln in lines), Decimal("0")))
    vat_local_total = q2(
        sum((ln.vat_local for ln in lines if ln.vat_local is not None), Decimal("0"))
    )
    net_local_total = q2(
        sum((ln.net_local for ln in lines if ln.net_local is not None), Decimal("0"))
    )
    if claim.vat_eur is None or q2(claim.vat_eur) != vat_eur_total:
        raise ConflictError(
            f"Frozen claim header vat_eur={claim.vat_eur} does not equal the sum of its "
            f"frozen lines ({vat_eur_total}) — refusing to render a drifted filing artifact",
            code="claim_totals_drift",
        )
    if claim.vat_local is not None and q2(claim.vat_local) != vat_local_total:
        raise ConflictError(
            f"Frozen claim header vat_local={claim.vat_local} does not equal the sum of its "
            f"frozen lines ({vat_local_total}) — refusing to render a drifted filing artifact",
            code="claim_totals_drift",
        )

    entity = await issuer_svc.get_by_id(db, org_id, claim.entity_id)
    entity_name = (entity.legal_name or entity.name or "").strip() if entity else ""

    # Vendor names via the vendors SERVICE (id -> name map), never a
    # cross-domain model import (`test_boundaries.py::
    # test_transport_services_do_not_import_other_domain_models`).
    vendor_names = {v.id: v.name for v in await vendors.list_vendors(db, org_id)}

    pack_lines: list[PackLine] = []
    for ln in lines:
        # After the synthetic refusal every line resolved to a real invoice
        # (`build_claim_lines`: a non-UNMATCHED ref always carries the
        # matched invoice's id) — belt-and-braces guard regardless, since a
        # filing artifact must never print an unattributable row.
        if ln.invoice_id is None:
            raise ConflictError(
                f"Frozen line '{ln.invoice_ref}' has no resolved invoice — cannot attribute "
                "it in a filing artifact",
                code="synthetic_line_in_pack",
            )
        inv = await ap_invoices.get_by_id(db, org_id, ln.invoice_id)
        supplier = vendor_names.get(inv.vendor_id, "") if inv is not None else ""
        pack_lines.append(
            PackLine(
                supplier=supplier,
                invoice_number=inv.invoice_number if inv is not None else ln.invoice_ref,
                invoice_date=inv.issue_date if inv is not None else None,
                goods_code=ln.goods_code,
                product_group=ln.product_group,
                net_eur=q2(ln.net_eur),
                vat_eur=q2(ln.vat_eur),
                net_local=q2(ln.net_local) if ln.net_local is not None else None,
                vat_local=q2(ln.vat_local) if ln.vat_local is not None else None,
                currency=ln.currency,
                invoice_id=ln.invoice_id,
            )
        )

    return ClaimPack(
        claim_id=claim.id,
        entity_name=entity_name,
        entity_vat_number=entity.vat_number if entity else None,
        entity_registration_number=entity.registration_number if entity else None,
        refund_country=claim.refund_country,
        ref_period=claim.ref_period,
        status=claim.status,
        status_code=claim.status_code,
        submitted_date=claim.submitted_date,
        vat_eur=vat_eur_total,
        vat_local=vat_local_total,
        currency=claim.currency,
        net_eur_total=net_eur_total,
        net_local_total=net_local_total,
        lines=tuple(pack_lines),
    )


def _render_workbook(pack: ClaimPack) -> bytes:
    """Render the claim workbook from a loaded pack. Two sheets (`Claim`
    header + `Lines` at the R2 grain with a TOTAL row). Free-text cells go
    through `sanitize_cell`; money cells are written as raw Decimals — a
    leading `-` there is a negative amount, not an injection
    (`core.csv_safety`'s own rule)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()

    ws = wb.active
    ws.title = "Claim"
    header_rows: list[tuple[str, object]] = [
        ("Claimant entity", sanitize_cell(pack.entity_name)),
        ("Entity VAT number", sanitize_cell(pack.entity_vat_number)),
        ("Entity registration number", sanitize_cell(pack.entity_registration_number)),
        ("Refund country", sanitize_cell(pack.refund_country)),
        ("Claim period", sanitize_cell(pack.ref_period)),
        ("Period label", sanitize_cell(period_label(pack.ref_period))),
        ("Status", sanitize_cell(pack.status)),
        ("Workflow code", sanitize_cell(pack.status_code)),
        ("Submitted date", pack.submitted_date.isoformat() if pack.submitted_date else ""),
        ("Line count", len(pack.lines)),
        ("VAT total EUR", pack.vat_eur),
        ("Net total EUR", pack.net_eur_total),
        ("VAT total local", pack.vat_local),
        ("Net total local", pack.net_local_total),
        ("Local currency", sanitize_cell(pack.currency)),
        ("Generated at (UTC)", datetime.now(UTC).isoformat()),
    ]
    for label, value in header_rows:
        ws.append([label, value])
    for row in ws["A1:A16"]:
        row[0].font = Font(bold=True)

    lines_ws = wb.create_sheet("Lines")
    lines_ws.append(list(_LINES_HEADER))
    for cell in lines_ws[1]:
        cell.font = Font(bold=True)
    for ln in pack.lines:
        lines_ws.append(
            [
                sanitize_cell(ln.supplier),
                sanitize_cell(ln.invoice_number),
                ln.invoice_date.isoformat() if ln.invoice_date else "",
                sanitize_cell(ln.goods_code),
                sanitize_cell(ln.product_group),
                ln.net_eur,
                ln.vat_eur,
                ln.net_local,
                ln.vat_local,
                sanitize_cell(ln.currency),
            ]
        )
    total_row = [
        "TOTAL",
        "",
        "",
        "",
        "",
        pack.net_eur_total,
        pack.vat_eur,
        pack.net_local_total,
        pack.vat_local,
        sanitize_cell(pack.currency),
    ]
    lines_ws.append(total_row)
    for cell in lines_ws[lines_ws.max_row]:
        cell.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def build_workbook(db: AsyncSession, org_id: str, claim_id: str) -> bytes:
    """The Excel claim workbook of a frozen claim (G2.12) — see the module
    docstring for the refusal rules (`claim_not_frozen`,
    `synthetic_line_in_pack`, `claim_totals_drift`; all fail CLOSED)."""
    pack = await _load_pack(db, org_id, claim_id)
    return _render_workbook(pack)


@dataclass(frozen=True)
class _BundleFile:
    path: str
    data: bytes


async def _collect_documents(db: AsyncSession, org_id: str, pack: ClaimPack) -> list[_BundleFile]:
    """Every vaulted capture of the pack's invoices, deduped by SHA-256
    (an invoice re-captured from the same bytes contributes one file), in
    deterministic (invoice_number, sha256) order. Missing object-store bytes
    refuse — module docstring "WHY MISSING DOCUMENT BYTES REFUSE"."""
    seen: set[str] = set()
    files: list[_BundleFile] = []
    by_invoice: dict[str, PackLine] = {}
    for ln in pack.lines:
        by_invoice.setdefault(ln.invoice_id, ln)

    for invoice_id, ln in sorted(by_invoice.items(), key=lambda kv: kv[1].invoice_number):
        runs = await extraction.list_for_invoice(db, org_id, invoice_id)
        for run in sorted(
            (r for r in runs if r.source_sha256 is not None), key=lambda r: r.source_sha256 or ""
        ):
            sha = run.source_sha256
            if sha is None or sha in seen:
                continue
            seen.add(sha)
            try:
                data = await documents.load(documents.UPLOADS, org_id, sha)
            except StorageError as exc:
                raise ConflictError(
                    f"Vaulted document {sha[:12]}… for invoice "
                    f"'{ln.invoice_number}' is missing from object storage — refusing to "
                    "assemble an incomplete filing bundle",
                    code="evidence_document_unavailable",
                ) from exc
            if data is None:
                raise ConflictError(
                    f"Vaulted document reference for invoice '{ln.invoice_number}' is empty — "
                    "refusing to assemble an incomplete filing bundle",
                    code="evidence_document_unavailable",
                )
            source_name = run.source_filename or f"{sha[:12]}.bin"
            filename = _sanitize_segment(f"{ln.invoice_number} - {source_name}")
            files.append(_BundleFile(path=filename, data=data))
    return files


def _manifest_csv(entries: list[tuple[str, bytes]]) -> str:
    """`manifest.csv` — file, sha256, size for every bundle entry (the §3.K
    K4 manifest discipline applied to the filing bundle). Filenames are
    free text -> sanitized; sizes/hashes are machine values, left raw."""
    out = io.StringIO()
    writer = csv.writer(out, lineterminator="\n")
    writer.writerow(["file", "sha256", "size_bytes"])
    for path, data in entries:
        writer.writerow([sanitize_cell(path), sha256_hex(data), len(data)])
    return out.getvalue()


async def build_evidence_pack(db: AsyncSession, org_id: str, claim_id: str) -> bytes:
    """The ZIP filing bundle of a frozen claim (G2.12, §3.K K6): the claim
    workbook (rendered from the SAME loaded pack — identical lines and
    totals by construction), every vaulted invoice document, and a
    `manifest.csv`, all under the human-navigable vault tree
    `<Customer> <RegNo>/<Year>/<Country>/<Claim period>/`. Same refusal
    rules as `build_workbook`, plus `evidence_document_unavailable` when a
    vaulted document's bytes are gone (fails CLOSED — module docstring)."""
    pack = await _load_pack(db, org_id, claim_id)
    workbook_bytes = _render_workbook(pack)
    doc_files = await _collect_documents(db, org_id, pack)

    prefix = vault_prefix(pack)
    entries: list[tuple[str, bytes]] = [(f"{prefix}/{WORKBOOK_FILENAME}", workbook_bytes)]
    entries.extend((f"{prefix}/{f.path}", f.data) for f in doc_files)

    manifest = _manifest_csv(entries)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path, data in entries:
            zf.writestr(path, data)
        zf.writestr(f"{prefix}/{MANIFEST_FILENAME}", manifest)
    return buf.getvalue()
