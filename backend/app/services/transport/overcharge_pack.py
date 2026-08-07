"""R41's two send-ready artifacts of a supplier overcharge claim-back — the
Excel EVIDENCE PACKET and the formal PDF CLAIM LETTER (G4.5; R41 + R53;
`BA_fleet_fuel.md` §2.4 line 221/236, §2.5 "Contract audit", §3.G G1 / R49;
WO-83).

WHAT THE TWO ARTIFACTS ARE, AND FOR WHOM
------------------------------------------
R41, verbatim: *"…with an **Excel evidence packet and a formal PDF claim letter
(with a credit/refund demand and a deadline) built from the SAME line source**,
and a `recovered_total` that feeds the north star. Read-only over the
analytics."* — with the acceptance *"Both artifacts for the same (supplier,
period) show identical lines and totals."*

The **packet** (`build_evidence_packet`) is the attachment a supplier's accounts
department reconciles against its own ledger: every breaching line with its
flag, its agreed and actual €/L, the gap, the litres and the recoverable euro,
plus the total. The **letter** (`build_claim_letter`) is the demand itself:
our letterhead, the supplier's own local legal entity, the same table, and
§2.4's *"formal PDF claim letter with a **30-day credit/refund demand**"*.

WHY BOTH ARTIFACTS COME FROM ONE LOADER AND ONE TABLE BUILDER
----------------------------------------------------------------
The binding precedent is `claim_pack.py` (G2.12/WO-74), whose own docstring
names this order's rule: *"the same both-artifacts-one-source rule the harvested
R41 states for the overcharge packet … identity is structural, never two queries
that happen to agree today."*

Here identity is structural THREE times over:

1. `_load_packet()` is the ONE loader. It calls `contract_audit.audit()` — the
   single line source, and the same one `overcharge.open_claim` snapshots — and
   carries its `Breach` objects into the packet UNCHANGED (not copied, not
   remapped).
2. `_COLUMNS` is the ONE column spec. Both renderers project the same columns,
   in the same order, from the same `Breach` objects, through the same
   accessors, so a cell can only differ if the two artifacts are looking at
   different lines — which (1) makes impossible.
3. Both renderers are SYNC functions over a loaded packet. They take no
   `AsyncSession` and cannot query anything, so neither can reach a second
   source even by accident.

The workbook additionally prints `_PROVENANCE_COLUMNS` (the claimant entity, the
as-invoiced and effective €/L, the line's document currency) to the RIGHT of the
shared block: a spreadsheet has room for lineage a one-page A4 letter does not,
and none of it is a figure — the shared block and the total are byte-identical
in both. R41's "identical lines and totals" is asserted cell-for-cell across the
parsed bytes of both artifacts (`test_wo83_both_artifacts_show_identical_lines_
and_totals`).

WHY A DRIFTED CLAIM-BACK REFUSES (fail-closed `overcharge_evidence_drift`)
-----------------------------------------------------------------------------
`overcharge.open_claim` FREEZES the detected euro (*"the euro the demand letter
quotes"*, `overcharge.py`) while `contract_audit.audit()` stays live. After a
re-ingest, a corrected line or an edited contract term the two can disagree —
which, for a REPORT, is information (`overcharge.py` says so). For a DEMAND it
is not: a letter claiming €8,000 whose attached evidence sums to €6,500 is a
letter a credit controller rejects on sight, and it is exactly the artifact R41's
acceptance exists to make impossible. So the pack refuses rather than silently
choosing one of the two figures — the fail-CLOSED twin of `claim_pack`'s
`claim_totals_drift` (*"a filing artifact must never paper over a corrupted
frozen set with a header figure its own lines cannot reproduce"*, §4.10). The
live figure stays one call away at `GET /transport/overcharges/audit`;
RE-SNAPSHOTTING a drifted claim-back is a lifecycle edge the spec never
harvested, so it is recorded in `docs/DECISIONS-NEEDED.md` §13 rather than
invented here (master-context §10).

WHY A TERMINAL CLAIM-BACK YIELDS THE PACKET BUT NOT A NEW LETTER
-------------------------------------------------------------------
`overcharge_claim_closed` (409), on the LETTER only. The packet is EVIDENCE and
stays reproducible in every state — WO-74's precedent exactly (*"a WITHDRAWN
claim keeps its frozen lines … reproducing what WAS filed is audit-trail
behaviour, not a leak"*). The letter is a LIVE demand carrying a payment
deadline: issuing one for a claim-back the org's own ledger records as
`recovered`, `rejected` or `written_off` would assert an open debt that is not
open. The terminal set is `overcharge.TRANSITIONS`' own empty tuples — no new
vocabulary is introduced.

WHY THE LETTER NEEDS A COMPLETE LETTERHEAD (fail-closed)
-----------------------------------------------------------
`issuer_profile_incomplete` (409, letter only), naming the missing fields.
`issuer.REQUIRED_FIELDS` is the EN 16931 / Art. 226 minimum this codebase
already enforces before an invoice may be issued; a formal demand from an
unnamed, unaddressed party is not a formal demand. The issuer is resolved
READ-ONLY (`list_issuers`/`get_by_id`) — never `issuer.get_or_create`, which
COMMITS a new row and would turn a read-only artifact into a writer.

THE FRAMING AND THE BASIS ARE PRINTED, NOT PARAPHRASED (R53, §3.G G1)
------------------------------------------------------------------------
R53: *"Legal framing per analysis must not be flattened: contract breach =
'money the supplier owes' (claim letter) … Each workbook carries its framing
text."* Both artifacts print `contract_audit.LEGAL_FRAMING` and
`contract_audit.PRICE_BASIS` — the SAME constants the JSON API already returns,
so a surface cannot drift from the wire. Basis, stated as §3.G G1 requires:
**NET EUR/L, final — VAT excluded, rebates applied.**

MONEY AND CURRENCY (§4.9, §4.10, §4.14)
------------------------------------------
Every euro is a `Decimal` produced by `contract_audit` (already `q2`-ed) and the
printed total is recomputed here with `money.q2` ROUND_HALF_UP, then checked
against BOTH the live audit total and the frozen `detected_eur` before anything
renders (§4.10 — the server recomputes every total). No float anywhere. The
artifacts total in **EUR only**: the three document-currency amount columns are
never read (a structural test asserts their absence by name), and a line's own
currency is printed in its own provenance column as exactly that — provenance,
never summed (§4.14; see `contract_audit.py`'s own §4.14 section for why a
supplier month spanning currencies is natively EUR-summable).

WHAT THIS MODULE NEVER DOES
------------------------------
Persists nothing, mutates nothing, audits nothing, commits nothing — the
`claim_pack`/`erp_export`/`report_writers` precedent (master-context §4.16
requires an audit row on *mutations*). Generating either artifact leaves the
claim-back and every fuel transaction column-for-column identical, which is
R41's *"read-only over the analytics"* and §4.19's advisory covenant.
Free-text cells go through the ONE shared `core.csv_safety.sanitize_cell`
(CWE-1236); numeric money cells deliberately do not, because a leading `-` there
is a negative amount, not an injection (`csv_safety`'s own rule). PDF text is
never evaluated as a formula, so the letter sanitizes nothing — it escapes,
because reportlab parses a mini-markup inside `Paragraph`.

NO NEW DEPENDENCY: `openpyxl` (claim_pack, report_writers) and `reportlab`
(invoice_pdf, report_writers) are both already pinned. If reportlab is somehow
absent the letter refuses 503 `pdf_renderer_unavailable` — the
`invoice_pdf.PdfUnavailable` → `issued.py` 503 behaviour, expressed as an
`AppError` because a service must not raise `HTTPException` (master-context §3)
and the route must stay a thin controller.
"""

from __future__ import annotations

import io
from collections.abc import Callable
from dataclasses import dataclass
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal
from xml.sax.saxutils import escape as xml_escape

from sqlalchemy.ext.asyncio import AsyncSession

from app.core.csv_safety import sanitize_cell
from app.core.errors import AppError, ConflictError, NotFoundError, PermissionError, ValidationError
from app.core.money import q2
from app.services import issuer as issuer_svc
from app.services import modules
from app.services.transport import contract_audit, overcharge, supplier_entity
from app.services.transport.contract_audit import Breach

# `BA_fleet_fuel.md` §2.4, verbatim twice: a *"formal PDF claim letter with a
# 30-day credit/refund demand"* (line 221) and *"a claim letter with a 30-day
# demand"* (the framing table, line 236). Harvested, not chosen.
DEMAND_DAYS = 30

# The demand is for a CREDIT NOTE **or** a REFUND — both words are §2.4's.
DEMAND_INSTRUMENTS = ("credit note", "refund")

PACKET_SHEET_CLAIM = "Claim-back"
PACKET_SHEET_EVIDENCE = "Evidence"
TOTAL_LABEL = "TOTAL"


class _Fmt:
    """How a column's value is rendered — the only place the two renderers are
    allowed to differ, and only in FORMATTING, never in value."""

    TEXT = "text"  # free text -> sanitized in Excel, escaped in the PDF
    MONEY = "money"  # EUR, 2 dp
    RATE = "rate"  # €/L, 4 dp (contract_audit.RATE_EXP)
    QTY = "qty"  # litres, 3 dp (FuelTransaction.qty's own scale)
    DATE = "date"  # ISO


@dataclass(frozen=True)
class _Column:
    """One column of the evidence table: its heading, how to read it off a
    `Breach`, and how to format it. Shared by BOTH renderers — see the module
    docstring's point 2."""

    label: str
    value: Callable[[Breach], object]
    fmt: str
    totalled: bool = False


# The SHARED demand table — identical, column for column, in both artifacts.
# Every accessor reads a `contract_audit.Breach` field; nothing is derived here.
_COLUMNS: tuple[_Column, ...] = (
    _Column("Date", lambda b: b.txn_date, _Fmt.DATE),
    _Column("Country", lambda b: b.country, _Fmt.TEXT),
    _Column("Station", lambda b: b.station, _Fmt.TEXT),
    _Column("Product group", lambda b: b.product_group, _Fmt.TEXT),
    _Column("Litres", lambda b: b.qty, _Fmt.QTY),
    _Column("Breach", lambda b: b.flag, _Fmt.TEXT),
    _Column("Agreed EUR/L", lambda b: b.agreed_eur_l, _Fmt.RATE),
    _Column("Actual EUR/L", lambda b: b.actual_eur_l, _Fmt.RATE),
    _Column("Gap EUR/L", lambda b: b.gap_eur_l, _Fmt.RATE),
    _Column("Recoverable EUR", lambda b: b.recover_eur, _Fmt.MONEY, totalled=True),
)

# Workbook-only LINEAGE columns. A spreadsheet has room for provenance a
# one-page A4 letter does not; none of these is a figure the demand rests on,
# and `Document currency` is provenance that is NEVER summed (§4.14).
_PROVENANCE_COLUMNS: tuple[_Column, ...] = (
    _Column("Claimant entity", lambda b: b.entity_id, _Fmt.TEXT),
    _Column("Invoiced EUR/L", lambda b: b.eur_l_doc, _Fmt.RATE),
    _Column("Effective EUR/L", lambda b: b.eur_l_eff, _Fmt.RATE),
    _Column("Document currency", lambda b: b.document_currency, _Fmt.TEXT),
    _Column("Fuel transaction id", lambda b: b.fuel_transaction_id, _Fmt.TEXT),
)


@dataclass(frozen=True)
class SupplierEntityLine:
    """The supplier's own local legal entity for one country of the evidence —
    R20/R21: the seller PRINTED on the invoice, resolved marker-only on the
    exact `(supplier, country)` key. `entity_name`/`vat_number` are `None` when
    no registration exists; nothing is ever inferred (master-context §10)."""

    country: str
    entity_name: str | None
    vat_number: str | None


@dataclass(frozen=True)
class OverchargePacket:
    """Everything both renderers need, loaded ONCE — see the module docstring
    "WHY BOTH ARTIFACTS COME FROM ONE LOADER AND ONE TABLE BUILDER".

    `lines` are `contract_audit.Breach` objects as detection produced them.
    `letterhead` is `issuer.seller_snapshot`'s plain dict (never an ORM object
    from another domain — ADR-P3 rule 2 / `test_boundaries.py`).
    """

    claim_id: str
    supplier: str
    period: str
    status: str
    currency: str
    price_basis: str
    legal_framing: str
    detected_eur: Decimal
    recovered_eur: Decimal | None
    lines: tuple[Breach, ...]
    total_eur: Decimal
    supplier_entities: tuple[SupplierEntityLine, ...]
    letterhead: dict | None
    letterhead_missing: tuple[str, ...]
    letter_date: date
    demand_due_date: date
    generated_at: datetime


async def _require_module(db: AsyncSession, org_id: str) -> None:
    """`modules.is_enabled`, never `require_enabled` — a service signals with
    `AppError`, not `fastapi.HTTPException` (ADR-P3 rule 3), and fails CLOSED."""
    if not await modules.is_enabled(db, org_id, "transport"):
        m = modules.MODULES_BY_KEY["transport"]
        raise PermissionError(f"The {m.name} module is not activated.", code="module_not_enabled")


async def _letterhead(db: AsyncSession, org_id: str, issuer_id: str | None):
    """OUR letterhead entity, READ-ONLY. The named company or the org's default
    (`issued.py`'s own resolution order, applied to a claim-back that carries no
    issuer of its own). Deliberately NOT `issuer.get_or_create`: that commits a
    new row, and a read-only artifact must never write one. An issuer id from
    another org is indistinguishable from an unknown one (§4.4, opaque 404)."""
    if issuer_id:
        profile = await issuer_svc.get_by_id(db, org_id, issuer_id)
        if profile is None:
            raise NotFoundError("Issuer company not found", code="issuer_not_found")
        return profile
    rows = await issuer_svc.list_issuers(db, org_id)
    return rows[0] if rows else None


async def _load_packet(
    db: AsyncSession, org_id: str, claim_id: str, *, issuer_id: str | None = None
) -> OverchargePacket:
    """The ONE loader both artifacts render from. Applies, in order: the module
    entitlement, the org-scoped claim-back fetch (opaque 404), the LIVE line
    source `contract_audit.audit()`, the empty-evidence refusal, the
    frozen-vs-live drift refusal, the per-country supplier entities and our
    letterhead. Every refusal fails CLOSED and the module docstring states why
    for each.

    Basis: **NET EUR/L, final — VAT excluded, rebates applied** (§3.G G1 / R49).
    Currency: **EUR only** (§4.14).
    """
    await _require_module(db, org_id)
    claim = await overcharge.get_claim(db, org_id, claim_id)

    # THE line source — the same call `overcharge.open_claim` snapshots. Both
    # artifacts render this one result; nothing else is queried for a figure.
    result = await contract_audit.audit(db, org_id, period=claim.period, supplier=claim.supplier)
    if not result.breaches:
        raise ValidationError(
            f"The contract audit finds no breach for {claim.supplier} in {claim.period} today — "
            "refusing to produce an evidence packet or a demand letter with no lines",
            code="no_overcharge_detected",
        )

    # §4.10 — recompute the total the artifacts will print, then require it to
    # reproduce the frozen figure the claim-back demands (module docstring,
    # "WHY A DRIFTED CLAIM-BACK REFUSES").
    total = q2(sum((b.recover_eur for b in result.breaches), Decimal("0")))
    detected = q2(claim.detected_eur)
    if total != detected:
        raise ConflictError(
            f"The claim-back demands {detected} EUR but its evidence now totals {total} EUR — "
            "refusing to send a demand its own attached lines do not reproduce "
            "(re-run the audit to see the current findings)",
            code="overcharge_evidence_drift",
        )

    entities: list[SupplierEntityLine] = []
    for country in sorted({b.country for b in result.breaches}):
        reg = await supplier_entity.get_registration(
            db, org_id, supplier=claim.supplier, country=country
        )
        entities.append(
            SupplierEntityLine(
                country=country,
                entity_name=reg.entity_name if reg is not None else None,
                vat_number=reg.vat_number if reg is not None else None,
            )
        )

    profile = await _letterhead(db, org_id, issuer_id)
    generated_at = datetime.now(UTC)
    letter_date = generated_at.date()

    return OverchargePacket(
        claim_id=claim.id,
        supplier=claim.supplier,
        period=claim.period,
        status=claim.status,
        currency=result.currency,
        price_basis=result.price_basis,
        legal_framing=result.legal_framing,
        detected_eur=detected,
        recovered_eur=None if claim.recovered_eur is None else q2(claim.recovered_eur),
        lines=result.breaches,
        total_eur=total,
        supplier_entities=tuple(entities),
        letterhead=None if profile is None else issuer_svc.seller_snapshot(profile),
        letterhead_missing=tuple(issuer_svc.missing_fields(profile)),
        letter_date=letter_date,
        demand_due_date=letter_date + timedelta(days=DEMAND_DAYS),
        generated_at=generated_at,
    )


# --------------------------------------------------------------------------- #
# The ONE table builder — both renderers project from here
# --------------------------------------------------------------------------- #


def _cell(column: _Column, breach: Breach) -> object:
    """One typed cell. Money and rates stay `Decimal` (§4.9); dates become their
    ISO string; text stays text. Formatting for a specific medium happens in the
    renderer, never here — so the two artifacts can only differ in APPEARANCE."""
    value = column.value(breach)
    if column.fmt == _Fmt.DATE:
        return value.isoformat() if isinstance(value, date) else ""
    return value


def _evidence_rows(packet: OverchargePacket, columns: tuple[_Column, ...]) -> list[list[object]]:
    """The evidence table's data rows, in detection order. `columns` is either
    the shared `_COLUMNS` (the letter, and the left block of the workbook) or
    `_COLUMNS + _PROVENANCE_COLUMNS` (the workbook). ONE builder, so a row can
    only differ between the artifacts if the LINES differ — which `_load_packet`
    makes impossible."""
    return [[_cell(c, b) for c in columns] for b in packet.lines]


def _total_row(packet: OverchargePacket, columns: tuple[_Column, ...]) -> list[object]:
    """The TOTAL row: the label in the first cell, the total under every
    `totalled` column, blanks elsewhere. The total is `packet.total_eur`, which
    `_load_packet` already recomputed AND reconciled against the claim-back's
    frozen demand (§4.10) — it is never re-summed differently here."""
    row: list[object] = []
    for idx, column in enumerate(columns):
        if idx == 0:
            row.append(TOTAL_LABEL)
        elif column.totalled:
            row.append(packet.total_eur)
        else:
            row.append("")
    return row


def _excel_cell(column: _Column, value: object) -> object:
    """A cell rendered into a spreadsheet. Free text goes through the ONE shared
    `sanitize_cell` (CWE-1236); money, rate and quantity cells are written as raw
    `Decimal`s so Excel holds real numbers — a leading `-` there is a negative
    amount, not an injection (`core.csv_safety`'s own rule)."""
    if column.fmt == _Fmt.TEXT or column.fmt == _Fmt.DATE:
        return sanitize_cell(value)
    return value


def _fmt_for_text(column: _Column, value: object) -> str:
    """A cell rendered as text (the PDF path, and every assertion that compares
    the two artifacts). Money 2 dp, rates 4 dp, litres 3 dp — the same scales
    `contract_audit` quantized them to."""
    if value == "" or value is None:
        return ""
    if column.fmt == _Fmt.MONEY:
        return f"{Decimal(str(value)):,.2f}"
    if column.fmt == _Fmt.RATE:
        return f"{Decimal(str(value)):.4f}"
    if column.fmt == _Fmt.QTY:
        return f"{Decimal(str(value)):,.3f}"
    return str(value)


# --------------------------------------------------------------------------- #
# Renderer 1 — the Excel evidence packet
# --------------------------------------------------------------------------- #


def _render_evidence_workbook(packet: OverchargePacket) -> bytes:
    """The Excel evidence packet: a `Claim-back` header sheet (carrying R53's
    legal framing and §3.G G1's price basis) and an `Evidence` sheet with the
    shared table, its workbook-only provenance columns and a bold TOTAL row.

    A SYNC function over a loaded packet — it holds no session and cannot query
    a second source (module docstring, point 3). Free-text cells go through the
    ONE shared `sanitize_cell`; money/rate/quantity cells are written as raw
    `Decimal`s, because a leading `-` there is a negative amount, not an
    injection (`core.csv_safety`'s own rule).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    columns = _COLUMNS + _PROVENANCE_COLUMNS

    wb = Workbook()
    ws = wb.active
    ws.title = PACKET_SHEET_CLAIM
    header_rows: list[tuple[str, object]] = [
        ("Supplier", sanitize_cell(packet.supplier)),
        ("Period", sanitize_cell(packet.period)),
        ("Claim-back status", sanitize_cell(packet.status)),
        ("Claim-back reference", sanitize_cell(packet.claim_id)),
        ("Currency", sanitize_cell(packet.currency)),
        ("Price basis", sanitize_cell(packet.price_basis)),
        ("Legal framing", sanitize_cell(packet.legal_framing)),
        ("Lines", len(packet.lines)),
        ("Recoverable total", packet.total_eur),
        ("Demanded (claim-back)", packet.detected_eur),
        ("Letter date", packet.letter_date.isoformat()),
        ("Payment due by", packet.demand_due_date.isoformat()),
        ("Generated at (UTC)", packet.generated_at.isoformat()),
    ]
    for label, value in header_rows:
        ws.append([label, value])
    for row in ws.iter_rows(min_row=1, max_row=len(header_rows), max_col=1):
        row[0].font = Font(bold=True)

    ws.append([])
    ws.append(["Supplier legal entity (per country of supply)"])
    ws[ws.max_row][0].font = Font(bold=True)
    ws.append(["Country", "Legal entity", "VAT number"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for ent in packet.supplier_entities:
        ws.append(
            [
                sanitize_cell(ent.country),
                sanitize_cell(ent.entity_name or ""),
                sanitize_cell(ent.vat_number or ""),
            ]
        )

    ev = wb.create_sheet(PACKET_SHEET_EVIDENCE)
    ev.append([c.label for c in columns])
    for cell in ev[1]:
        cell.font = Font(bold=True)
    for row in [*_evidence_rows(packet, columns), _total_row(packet, columns)]:
        ev.append([_excel_cell(columns[i], value) for i, value in enumerate(row)])
    for cell in ev[ev.max_row]:
        cell.font = Font(bold=True)

    # R53 / §3.G G1 on the sheet that carries the numbers, not only on the
    # header sheet — "each workbook carries its framing text".
    ev.append([])
    ev.append([packet.legal_framing])
    ev.append([packet.price_basis])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Renderer 2 — the formal PDF claim letter
# --------------------------------------------------------------------------- #


def _addressee_lines(packet: OverchargePacket) -> list[str]:
    """Who the demand is addressed to: the fuel-card network, then its own local
    legal entity per country of supply (R20 — a multi-entity network issues a
    different entity per country). A country with no registration prints nothing
    beyond its code: marker-only, never a guessed name (R21)."""
    out = [packet.supplier]
    for ent in packet.supplier_entities:
        if ent.entity_name and ent.vat_number:
            out.append(f"{ent.country}: {ent.entity_name} (VAT {ent.vat_number})")
        elif ent.entity_name:
            out.append(f"{ent.country}: {ent.entity_name}")
        else:
            out.append(f"{ent.country}: no registered legal entity on file")
    return out


def _letterhead_lines(packet: OverchargePacket) -> list[str]:
    """Our letterhead from the existing issuer registry. Only fields that are
    actually set are printed — an empty line is never invented."""
    head = packet.letterhead or {}
    lines = [str(head.get("legal_name") or "")]
    for key in ("address_line1", "address_line2"):
        if head.get(key):
            lines.append(str(head[key]))
    town = " ".join(str(head.get(k) or "") for k in ("postal_code", "city")).strip()
    if town:
        lines.append(town)
    if head.get("country"):
        lines.append(str(head["country"]))
    if head.get("vat_number"):
        lines.append(f"VAT {head['vat_number']}")
    if head.get("registration_number"):
        lines.append(f"Reg. {head['registration_number']}")
    for key in ("email", "phone"):
        if head.get(key):
            lines.append(str(head[key]))
    return [line for line in lines if line]


def demand_paragraph(packet: OverchargePacket) -> str:
    """§2.4's demand, in one sentence: a **credit note or refund** of the total,
    **within 30 days**. Public so a test asserts the harvested wording rather
    than a paraphrase of it."""
    return (
        f"We therefore request a {DEMAND_INSTRUMENTS[0]} or {DEMAND_INSTRUMENTS[1]} of "
        f"EUR {packet.total_eur:,.2f} within {DEMAND_DAYS} days of the date of this letter, "
        f"that is by {packet.demand_due_date.isoformat()}."
    )


def _render_claim_letter(packet: OverchargePacket) -> bytes:
    """The formal PDF claim letter — our letterhead, the supplier's local legal
    entity, the SAME evidence table, and §2.4's 30-day credit/refund demand.

    A SYNC function over a loaded packet: it holds no session and cannot query a
    second source. `Paragraph` text is XML-escaped because reportlab parses a
    mini-markup there; `Table` cells take plain strings and need no escaping.
    Nothing is sanitized for formula injection — PDF text is never evaluated as
    a formula (`report_writers.to_pdf`'s own note).
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:  # pragma: no cover - the dependency is pinned
        raise AppError(
            f"The PDF renderer is unavailable on this deployment: {exc}",
            code="pdf_renderer_unavailable",
            status=503,
        ) from exc

    ink = "#1e293b"
    muted = "#64748b"
    brand = "#2f57d4"
    line = "#e2e8f0"

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title", parent=styles["Title"], fontSize=16, textColor=colors.HexColor(brand)
    )
    body = ParagraphStyle(
        "body", parent=styles["Normal"], fontSize=9.5, leading=13, textColor=colors.HexColor(ink)
    )
    small = ParagraphStyle(
        "small", parent=styles["Normal"], fontSize=8, leading=10, textColor=colors.HexColor(muted)
    )

    def para(text: str, style: ParagraphStyle) -> Paragraph:
        return Paragraph(xml_escape(text), style)

    story: list = []
    for text in _letterhead_lines(packet):
        story.append(para(text, small))
    story.append(Spacer(1, 10))
    story.append(para(packet.letter_date.isoformat(), small))
    story.append(Spacer(1, 6))
    story.append(para("To:", small))
    for text in _addressee_lines(packet):
        story.append(para(text, body))
    story.append(Spacer(1, 12))
    story.append(
        para(f"Contract breach claim-back — {packet.supplier}, {packet.period}", title_style)
    )
    story.append(Spacer(1, 6))
    # R53 — the framing, printed verbatim from the same constant the API returns.
    story.append(para(packet.legal_framing, body))
    story.append(Spacer(1, 6))
    story.append(
        para(
            f"Our audit of the fuel and toll lines you invoiced for {packet.period} against the "
            f"commercial terms agreed with you identifies {len(packet.lines)} line(s) that were "
            f"charged outside those terms, totalling EUR {packet.total_eur:,.2f}. The lines are "
            "set out below and are enclosed in full as a spreadsheet.",
            body,
        )
    )
    story.append(Spacer(1, 10))

    header = [c.label for c in _COLUMNS]
    rows = [
        [_fmt_for_text(_COLUMNS[i], v) for i, v in enumerate(row)]
        for row in _evidence_rows(packet, _COLUMNS)
    ]
    total = [_fmt_for_text(_COLUMNS[i], v) for i, v in enumerate(_total_row(packet, _COLUMNS))]
    tbl = Table([header, *rows, total], repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(brand)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("ALIGN", (4, 0), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LINEBELOW", (0, 1), (-1, -2), 0.4, colors.HexColor(line)),
                ("LINEABOVE", (0, -1), (-1, -1), 0.8, colors.HexColor(ink)),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor(ink)),
            ]
        )
    )
    story.append(tbl)
    story.append(Spacer(1, 12))
    story.append(para(demand_paragraph(packet), body))

    head = packet.letterhead or {}
    if head.get("iban"):
        bic = f", BIC {head['bic']}" if head.get("bic") else ""
        story.append(Spacer(1, 4))
        story.append(para(f"Refunds to IBAN {head['iban']}{bic}.", body))
    story.append(Spacer(1, 10))
    story.append(
        para(
            "Please confirm within the period above which of the two you will issue. This letter "
            "and its enclosure are without prejudice to any further amounts arising from the same "
            "terms in other periods.",
            body,
        )
    )
    story.append(Spacer(1, 12))
    # §3.G G1 — the basis, stated on the report surface.
    story.append(para(f"Basis: {packet.price_basis}. All amounts in EUR.", small))
    story.append(para(f"Claim-back reference {packet.claim_id}.", small))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        title=f"Contract breach claim-back {packet.supplier} {packet.period}",
    )
    doc.build(story)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# The two public builders
# --------------------------------------------------------------------------- #


async def build_evidence_packet(
    db: AsyncSession, org_id: str, claim_id: str, *, issuer_id: str | None = None
) -> bytes:
    """R41's **Excel evidence packet** for one claim-back — the per-(supplier ×
    period) breach lines with their flags, gap, litres and recoverable euros,
    plus the total, carrying R53's legal framing and §3.G G1's price basis.

    Available in EVERY lifecycle state, including the terminal three: the packet
    is EVIDENCE and stays reproducible (the WO-74 precedent). Refuses
    `module_not_enabled` (403), `overcharge_claim_not_found` (404, opaque),
    `no_overcharge_detected` (422) and `overcharge_evidence_drift` (409) — all
    fail CLOSED; see the module docstring for why each.

    Read-only: nothing is persisted, mutated, audited or committed.
    """
    packet = await _load_packet(db, org_id, claim_id, issuer_id=issuer_id)
    return _render_evidence_workbook(packet)


async def build_claim_letter(
    db: AsyncSession, org_id: str, claim_id: str, *, issuer_id: str | None = None
) -> bytes:
    """R41's **formal PDF claim letter** for one claim-back — our letterhead, the
    supplier's own local legal entity, the SAME lines and the SAME total as the
    evidence packet (structurally, not coincidentally), and §2.4's 30-day
    credit/refund demand.

    Two refusals the packet does not have, both fail-CLOSED and both explained in
    the module docstring: `overcharge_claim_closed` (409) for a claim-back in a
    terminal state — a live demand must not assert a debt the ledger records as
    settled, rejected or written off — and `issuer_profile_incomplete` (409) when
    our own letterhead lacks an Art. 226 field.

    Read-only: nothing is persisted, mutated, audited or committed.
    """
    packet = await _load_packet(db, org_id, claim_id, issuer_id=issuer_id)
    if not overcharge.TRANSITIONS[packet.status]:
        raise ConflictError(
            f"This claim-back is '{packet.status}' — a closed matter takes no new "
            f"{DEMAND_DAYS}-day payment demand (the evidence packet stays available)",
            code="overcharge_claim_closed",
        )
    if packet.letterhead_missing:
        raise ConflictError(
            "The claim letter goes out on your own letterhead, and your issuer company is "
            "missing: " + ", ".join(packet.letterhead_missing),
            code="issuer_profile_incomplete",
        )
    return _render_claim_letter(packet)


__all__ = [
    "DEMAND_DAYS",
    "DEMAND_INSTRUMENTS",
    "PACKET_SHEET_CLAIM",
    "PACKET_SHEET_EVIDENCE",
    "TOTAL_LABEL",
    "OverchargePacket",
    "SupplierEntityLine",
    "build_claim_letter",
    "build_evidence_packet",
    "demand_paragraph",
]
